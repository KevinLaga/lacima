# empaques/views_inventory.py
from decimal import Decimal

from django.contrib import messages
from django.contrib.auth.decorators import login_required, permission_required
from django.db import transaction
from django.db.models import DecimalField, ExpressionWrapper, F, Q, Sum, Value as V
from django.db.models.functions import Coalesce
from django.http import HttpResponse, HttpResponseForbidden
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse

from .models import (
    EMPRESA_CHOICES, EMPRESAS_ALMACEN,
    InventoryItem, InventoryMovement,
    Pedimento, PedimentoItem,
    Remision, RemisionItem,
)
from .forms_inventory import (
    InventoryItemForm, InventoryMovementForm,
    PedimentoForm, PedimentoItemFormSet,
    RemisionForm, RemisionItemFormSet,
)


# ─────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────

def _stock_por_empresa(empresa):
    """
    Devuelve queryset con stock disponible por artículo para una empresa.
    stock = Σcantidad - Σconsumido en PedimentoItems de esa empresa.
    """
    return (
        PedimentoItem.objects
        .filter(pedimento__empresa=empresa)
        .values('articulo_id', 'articulo__sku', 'articulo__name', 'articulo__unit')
        .annotate(
            total=Coalesce(Sum('cantidad'), Decimal('0'),
                           output_field=DecimalField(max_digits=12, decimal_places=2)),
            consumido=Coalesce(Sum('consumido'), Decimal('0'),
                               output_field=DecimalField(max_digits=12, decimal_places=2)),
        )
        .annotate(
            stock=ExpressionWrapper(F('total') - F('consumido'),
                                    output_field=DecimalField(max_digits=12, decimal_places=2))
        )
        .order_by('articulo__name')
    )


def _aplicar_fifo(remision):
    """
    Para cada artículo de la remisión, descuenta de los PedimentoItems
    más antiguos de la misma empresa primero.
    Devuelve lista de errores si no hay stock suficiente.
    """
    errores = []
    for ri in remision.items.select_related('articulo'):
        pendiente = ri.cantidad
        lotes = (
            PedimentoItem.objects
            .filter(pedimento__empresa=remision.empresa, articulo=ri.articulo)
            .filter(consumido__lt=F('cantidad'))
            .order_by('pedimento__fecha', 'pedimento__id')
        )
        for lote in lotes:
            if pendiente <= 0:
                break
            disponible = lote.cantidad - lote.consumido
            tomar = min(disponible, pendiente)
            lote.consumido += tomar
            lote.save(update_fields=['consumido'])
            pendiente -= tomar
        if pendiente > 0:
            errores.append(
                f"{ri.articulo.name}: falta {pendiente} {ri.articulo.unit} (stock insuficiente)"
            )
    return errores


def _revertir_fifo(remision):
    """Revierte el FIFO de una remisión eliminando su consumo de los lotes."""
    for ri in remision.items.select_related('articulo'):
        por_revertir = ri.cantidad
        lotes = (
            PedimentoItem.objects
            .filter(pedimento__empresa=remision.empresa, articulo=ri.articulo)
            .filter(consumido__gt=0)
            .order_by('-pedimento__fecha', '-pedimento__id')
        )
        for lote in lotes:
            if por_revertir <= 0:
                break
            revertir = min(lote.consumido, por_revertir)
            lote.consumido -= revertir
            lote.save(update_fields=['consumido'])
            por_revertir -= revertir


# ─────────────────────────────────────────────
# Almacén (vista principal)
# ─────────────────────────────────────────────

@login_required
def almacen_list(request):
    empresa = request.GET.get('empresa', '').strip()
    if empresa not in EMPRESAS_ALMACEN:
        empresa = ''

    stock_rows = _stock_por_empresa(empresa) if empresa else []

    return render(request, 'empaques/almacen_list.html', {
        'empresa': empresa,
        'empresas': EMPRESAS_ALMACEN,
        'stock_rows': stock_rows,
    })


# ─────────────────────────────────────────────
# Artículos del catálogo
# ─────────────────────────────────────────────

@login_required
@permission_required("empaques.add_inventoryitem", raise_exception=True)
def almacen_item_new(request):
    if request.method == "POST":
        form = InventoryItemForm(request.POST)
        if form.is_valid():
            item = form.save(commit=False)
            if hasattr(item, "created_by_id") and not item.created_by_id:
                item.created_by = request.user
            item.save()
            messages.success(request, "Artículo creado correctamente.")
            return redirect(reverse("almacen_list"))
        messages.error(request, "Revisa los errores del formulario.")
    else:
        form = InventoryItemForm()
    return render(request, "empaques/almacen_item_form.html", {"form": form})


# ─────────────────────────────────────────────
# Pedimentos
# ─────────────────────────────────────────────

@login_required
def pedimento_list(request):
    empresa = request.GET.get('empresa', '').strip()
    if empresa not in EMPRESAS_ALMACEN:
        empresa = ''

    orden = request.GET.get('orden', '-fecha')
    if orden not in ('fecha', '-fecha'):
        orden = '-fecha'

    qs = Pedimento.objects.prefetch_related('items__articulo')
    if empresa:
        qs = qs.filter(empresa=empresa)
    qs = qs.order_by(orden)

    return render(request, 'empaques/pedimento_list.html', {
        'pedimentos': qs,
        'empresa': empresa,
        'empresas': EMPRESAS_ALMACEN,
        'orden': orden,
    })


@login_required
@permission_required("empaques.add_pedimento", raise_exception=True)
def pedimento_new(request):
    empresa_param = request.GET.get('empresa', '')

    if request.method == 'POST':
        form = PedimentoForm(request.POST)
        formset = PedimentoItemFormSet(request.POST)
        if form.is_valid() and formset.is_valid():
            with transaction.atomic():
                ped = form.save(commit=False)
                ped.created_by = request.user
                ped.save()
                formset.instance = ped
                formset.save()
            messages.success(request, f"Pedimento {ped.folio} creado.")
            return redirect('pedimento_detail', pk=ped.pk)
    else:
        initial = {}
        if empresa_param in EMPRESAS_ALMACEN:
            initial['empresa'] = empresa_param
        form = PedimentoForm(initial=initial)
        formset = PedimentoItemFormSet()

    return render(request, 'empaques/pedimento_form.html', {
        'form': form,
        'formset': formset,
        'titulo': 'Nuevo Pedimento',
        'empresa_param': empresa_param,
    })


@login_required
def pedimento_detail(request, pk):
    ped = get_object_or_404(Pedimento.objects.prefetch_related('items__articulo'), pk=pk)
    return render(request, 'empaques/pedimento_detail.html', {'ped': ped})


# ─────────────────────────────────────────────
# Remisiones
# ─────────────────────────────────────────────

@login_required
def remision_list(request):
    empresa = request.GET.get('empresa', '').strip()
    if empresa not in EMPRESAS_ALMACEN:
        empresa = ''

    orden = request.GET.get('orden', '-fecha')
    if orden not in ('fecha', '-fecha'):
        orden = '-fecha'

    qs = Remision.objects.prefetch_related('items__articulo')
    if empresa:
        qs = qs.filter(empresa=empresa)
    qs = qs.order_by(orden)

    return render(request, 'empaques/remision_list.html', {
        'remisiones': qs,
        'empresa': empresa,
        'empresas': EMPRESAS_ALMACEN,
        'orden': orden,
    })


@login_required
@permission_required("empaques.add_remision", raise_exception=True)
def remision_new(request):
    empresa_param = request.GET.get('empresa', '')

    if request.method == 'POST':
        form = RemisionForm(request.POST)
        formset = RemisionItemFormSet(request.POST)
        if form.is_valid() and formset.is_valid():
            with transaction.atomic():
                rem = form.save(commit=False)
                rem.created_by = request.user
                rem.save()
                formset.instance = rem
                formset.save()
                errores = _aplicar_fifo(rem)
                if errores:
                    # Revertir todo si no hay stock suficiente
                    _revertir_fifo(rem)
                    rem.delete()
                    for e in errores:
                        messages.error(request, e)
                    return render(request, 'empaques/remision_form.html', {
                        'form': form,
                        'formset': RemisionItemFormSet(request.POST),
                        'titulo': 'Nueva Remisión',
                        'empresa_param': empresa_param,
                    })
            messages.success(request, f"Remisión {rem.folio} creada y stock descontado (FIFO).")
            return redirect('remision_detail', pk=rem.pk)
    else:
        initial = {}
        if empresa_param in EMPRESAS_ALMACEN:
            initial['empresa'] = empresa_param
        form = RemisionForm(initial=initial)
        formset = RemisionItemFormSet()

    return render(request, 'empaques/remision_form.html', {
        'form': form,
        'formset': formset,
        'titulo': 'Nueva Remisión',
        'empresa_param': empresa_param,
    })


@login_required
def remision_detail(request, pk):
    rem = get_object_or_404(Remision.objects.prefetch_related('items__articulo'), pk=pk)

    # Traza FIFO: qué pedimentos proveyeron cada artículo
    traza = []
    for ri in rem.items.all():
        lotes = (
            PedimentoItem.objects
            .filter(pedimento__empresa=rem.empresa, articulo=ri.articulo)
            .filter(consumido__gt=0)
            .select_related('pedimento')
            .order_by('pedimento__fecha', 'pedimento__id')
        )
        traza.append({'item': ri, 'lotes': list(lotes)})

    return render(request, 'empaques/remision_detail.html', {
        'rem': rem,
        'traza': traza,
    })


# ─────────────────────────────────────────────
# Reporte Excel de inventario
# ─────────────────────────────────────────────

LOGO_MAP = {
    "La Cima Produce":      "la-cima-produce",
    "RC Organics":          "rc-organics",
    "GH Farms":             "gh-farms",
    "Gourmet Baja Farms":   "gourmet-baja-farms",
    "GBF Farms":            "gbf-farms",
    "AGRICOLA DH & G":      "AGRICOLA",
    "AGRICOLA DH&G GONZALO":"AGRICOLA",
}

@login_required
def almacen_reporte_xlsx(request):
    import os
    from io import BytesIO
    from datetime import date as _date
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, GradientFill
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.utils import get_column_letter
    from django.conf import settings as django_settings

    empresa = request.GET.get('empresa', '').strip()
    if empresa not in EMPRESAS_ALMACEN:
        return redirect(reverse('almacen_list'))

    # ── Stock actual ──────────────────────────────────────────────────
    rows = list(_stock_por_empresa(empresa))

    # ── Estilos ───────────────────────────────────────────────────────
    NAVY   = "1E3A5F"
    LIGHT  = "EBF2FA"
    WHITE  = "FFFFFF"
    GREEN  = "166534"
    RED    = "991B1B"
    GRAY   = "6B7280"

    title_font    = Font(name='Calibri', size=20, bold=True, color=NAVY)
    sub_font      = Font(name='Calibri', size=11,  italic=True, color=GRAY)
    th_font       = Font(name='Calibri', size=11,  bold=True,  color=WHITE)
    th_fill       = PatternFill("solid", fgColor=NAVY)
    row_fill_alt  = PatternFill("solid", fgColor=LIGHT)
    bold_navy     = Font(name='Calibri', size=11,  bold=True,  color=NAVY)
    ok_font       = Font(name='Calibri', size=11,  bold=True,  color=GREEN)
    zero_font     = Font(name='Calibri', size=11,  bold=True,  color=RED)
    normal_font   = Font(name='Calibri', size=11)
    gray_font     = Font(name='Calibri', size=10,  color=GRAY)

    thin  = Side(style='thin',   color='D1D5DB')
    thick = Side(style='medium', color=NAVY)
    border_th   = Border(left=thick, right=thick, top=thick, bottom=thick)
    border_cell = Border(left=thin,  right=thin,  top=thin,  bottom=thin)
    border_bot  = Border(left=thin,  right=thin,  top=thin,  bottom=thick)

    center = Alignment(horizontal='center', vertical='center')
    left   = Alignment(horizontal='left',   vertical='center')
    right  = Alignment(horizontal='right',  vertical='center')

    # ── Workbook ──────────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"
    ws.sheet_view.showGridLines = False

    # ── Logo ──────────────────────────────────────────────────────────
    logo_slug = LOGO_MAP.get(empresa)
    LOGO_ROWS = 6   # filas reservadas para el logo
    if logo_slug:
        logo_path = os.path.join(django_settings.BASE_DIR, "static", "logos", f"{logo_slug}.png")
        if os.path.exists(logo_path):
            img = XLImage(logo_path)
            target_h = 90
            scale    = target_h / img.height
            img.width  = int(img.width  * scale)
            img.height = int(img.height * scale)
            ws.add_image(img, "A1")

    # ── Título (columnas D-H) ─────────────────────────────────────────
    ws.merge_cells("D1:H3")
    tc = ws["D1"]
    tc.value     = f"Reporte de Inventario — {empresa}"
    tc.font      = title_font
    tc.alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("D4:H5")
    sc = ws["D4"]
    sc.value     = f"Generado el {_date.today().strftime('%d/%m/%Y')}   |   Stock disponible al momento de descarga"
    sc.font      = sub_font
    sc.alignment = Alignment(horizontal="left", vertical="center")

    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 25
    ws.row_dimensions[3].height = 25
    ws.row_dimensions[4].height = 20
    ws.row_dimensions[5].height = 20
    ws.row_dimensions[6].height = 10  # spacer

    # ── Encabezados de tabla ──────────────────────────────────────────
    HDR_ROW = 7
    COLS = ["Id", "Artículo", "Unidad", "Entradas totales", "Consumido", "Disponible"]
    COL_WIDTHS = [12, 38, 12, 18, 14, 14]

    for i, (h, w) in enumerate(zip(COLS, COL_WIDTHS), start=1):
        c = ws.cell(row=HDR_ROW, column=i, value=h)
        c.font      = th_font
        c.fill      = th_fill
        c.alignment = center
        c.border    = border_th
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[HDR_ROW].height = 26
    ws.freeze_panes = f"A{HDR_ROW + 1}"

    # ── Filas de datos ────────────────────────────────────────────────
    r = HDR_ROW + 1
    total_entradas  = Decimal('0')
    total_consumido = Decimal('0')
    total_disponible = Decimal('0')

    for idx, row in enumerate(rows):
        alt     = (idx % 2 == 1)
        fill    = row_fill_alt if alt else None
        stock   = row['stock']
        ent     = row['total']
        cons    = row['consumido']

        values = [
            row['articulo__sku'],
            row['articulo__name'],
            row['articulo__unit'],
            float(ent),
            float(cons),
            float(stock),
        ]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.border    = border_cell
            cell.alignment = center if col != 2 else left
            if fill:
                cell.fill = fill

            if col == 1:
                cell.font = gray_font
            elif col == 2:
                cell.font = bold_navy
            elif col == 6:
                cell.font = ok_font if stock > 0 else zero_font
            else:
                cell.font = normal_font

        ws.row_dimensions[r].height = 20
        total_entradas   += ent
        total_consumido  += cons
        total_disponible += stock
        r += 1

    # ── Fila de totales ───────────────────────────────────────────────
    if rows:
        ws.row_dimensions[r].height = 22
        ws.cell(row=r, column=1, value="TOTAL").font = Font(name='Calibri', size=11, bold=True, color=WHITE)
        ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=NAVY)
        ws.cell(row=r, column=1).alignment = center

        ws.cell(row=r, column=2, value="").fill = PatternFill("solid", fgColor=NAVY)
        ws.cell(row=r, column=3, value="").fill = PatternFill("solid", fgColor=NAVY)

        for col, val in [(4, float(total_entradas)), (5, float(total_consumido)), (6, float(total_disponible))]:
            cell      = ws.cell(row=r, column=col, value=val)
            cell.font = Font(name='Calibri', size=11, bold=True, color=WHITE)
            cell.fill = PatternFill("solid", fgColor=NAVY)
            cell.alignment = center
            cell.border    = border_bot
        r += 1

    # ── Pie de página ─────────────────────────────────────────────────
    

    # ── Respuesta HTTP ────────────────────────────────────────────────
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    from django.utils.text import slugify
    filename = f"inventario_{slugify(empresa)}_{_date.today().strftime('%Y%m%d')}.xlsx"
    resp = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp


# ─────────────────────────────────────────────
# Inventario inicial
# ─────────────────────────────────────────────

@login_required
@permission_required("empaques.add_pedimento", raise_exception=True)
def inventario_inicial(request):
    empresa = request.GET.get('empresa', '').strip()
    if empresa not in EMPRESAS_ALMACEN:
        empresa = ''

    articulos = InventoryItem.objects.order_by('name')
    errores = []

    if request.method == 'POST':
        empresa = request.POST.get('empresa', '').strip()
        fecha   = request.POST.get('fecha', '').strip()

        if empresa not in EMPRESAS_ALMACEN:
            errores.append("Selecciona una empresa válida.")
        if not fecha:
            errores.append("La fecha es obligatoria.")

        cantidades = {}
        for art in articulos:
            val = request.POST.get(f'qty_{art.pk}', '').strip()
            if val:
                try:
                    qty = Decimal(val)
                    if qty > 0:
                        cantidades[art] = qty
                except Exception:
                    errores.append(f"Cantidad inválida para {art.name}.")

        if not cantidades:
            errores.append("Captura al menos una cantidad mayor a 0.")

        if not errores:
            with transaction.atomic():
                from datetime import date as _date
                fecha_obj = _date.fromisoformat(fecha)
                ped = Pedimento.objects.create(
                    empresa=empresa,
                    fecha=fecha_obj,
                    notas="Inventario inicial",
                    created_by=request.user,
                )
                for art, qty in cantidades.items():
                    PedimentoItem.objects.create(
                        pedimento=ped,
                        articulo=art,
                        cantidad=qty,
                    )
            messages.success(request, f"Inventario inicial guardado como {ped.folio}.")
            return redirect(f"{reverse('inventario_inicial')}?empresa={empresa}")

    return render(request, 'empaques/inventario_inicial.html', {
        'empresa': empresa,
        'empresas': EMPRESAS_ALMACEN,
        'articulos': articulos,
        'errores': errores,
    })


# ─────────────────────────────────────────────
# Movimiento individual (legacy, mantener funcional)
# ─────────────────────────────────────────────

@login_required
@permission_required("empaques.add_inventorymovement", raise_exception=True)
def almacen_movement_new(request, tipo=None):
    if request.method == "POST":
        form = InventoryMovementForm(request.POST, user=request.user)
        if form.is_valid():
            m = form.save(commit=False)
            if m.type == "ADJ":
                agg = InventoryMovement.objects.filter(item=m.item).aggregate(
                    ent=Coalesce(Sum('quantity', filter=Q(type='IN'),
                                 output_field=DecimalField(max_digits=12, decimal_places=2)),
                                 V(0), output_field=DecimalField(max_digits=12, decimal_places=2)),
                    sal=Coalesce(Sum('quantity', filter=Q(type='OUT'),
                                 output_field=DecimalField(max_digits=12, decimal_places=2)),
                                 V(0), output_field=DecimalField(max_digits=12, decimal_places=2)),
                    adj=Coalesce(Sum('quantity', filter=Q(type='ADJ'),
                                 output_field=DecimalField(max_digits=12, decimal_places=2)),
                                 V(0), output_field=DecimalField(max_digits=12, decimal_places=2)),
                )
                current_stock = (agg["ent"] - agg["sal"] + agg["adj"]) or Decimal("0")
                desired_final = m.quantity
                delta = desired_final - current_stock
                m.quantity = delta
                if not m.notes:
                    m.notes = f"Ajuste a {desired_final} (Δ {delta})"
            m.save()
            messages.success(request, "Movimiento registrado.")
            return redirect("almacen_kardex", item_id=m.item_id)
        messages.error(request, "Revisa los errores del formulario.")
    else:
        initial = {}
        if tipo in ("IN", "OUT", "ADJ"):
            initial["type"] = tipo
        item_id = request.GET.get("item")
        if item_id:
            initial["item"] = item_id
        form = InventoryMovementForm(initial=initial)
    return render(request, "empaques/almacen_movement_form.html", {"form": form})


@login_required
def almacen_kardex(request, item_id):
    if not request.user.has_perm("empaques.view_inventorymovement"):
        return HttpResponseForbidden("No tienes permiso para ver movimientos.")
    item = get_object_or_404(InventoryItem, pk=item_id)
    movimientos = item.movements.select_related("created_by").order_by("date", "id")
    saldo = Decimal("0")
    rows = []
    for m in movimientos:
        delta = m.quantity if m.type in ("IN", "ADJ") else -m.quantity
        saldo += delta
        rows.append({"m": m, "saldo": saldo})
    stock_agg = item.movements.aggregate(
        ent=Coalesce(Sum('quantity', filter=Q(type='IN'),
                     output_field=DecimalField(max_digits=12, decimal_places=2)),
                     V(0), output_field=DecimalField(max_digits=12, decimal_places=2)),
        sal=Coalesce(Sum('quantity', filter=Q(type='OUT'),
                     output_field=DecimalField(max_digits=12, decimal_places=2)),
                     V(0), output_field=DecimalField(max_digits=12, decimal_places=2)),
        adj=Coalesce(Sum('quantity', filter=Q(type='ADJ'),
                     output_field=DecimalField(max_digits=12, decimal_places=2)),
                     V(0), output_field=DecimalField(max_digits=12, decimal_places=2)),
    )
    stock_qty = stock_agg["ent"] - stock_agg["sal"] + stock_agg["adj"]
    return render(request, "empaques/almacen_kardex.html", {
        "item": item, "rows": rows, "stock_qty": stock_qty,
    })
