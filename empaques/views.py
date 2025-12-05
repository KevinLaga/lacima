import os
import csv
from datetime import date
from collections import defaultdict
import collections

from django.conf import settings
from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.forms import inlineformset_factory
from django.http import HttpResponse, HttpResponseForbidden
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.shortcuts import redirect
from collections import defaultdict
from .models import ProductionDisplay 



from .models import Presentation, Shipment, ShipmentItem
from .forms import (
    ShipmentForm,
    ShipmentItemForm,
    BaseShipmentItemFormSet,
)

from django.contrib.auth.decorators import login_required, permission_required

import json
from datetime import datetime, timedelta
from pathlib import Path

from django.views.decorators.http import require_http_methods
from django.utils.dateparse import parse_date
from django.shortcuts import render, redirect
from django.http import HttpResponse, HttpResponseBadRequest

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from .models import Shipment, ShipmentItem
from .production_config import ALLOWED_COMBOS, ORDERED_ALIASES

# --- PRODUCCIÓN DEL DÍA (sin migraciones) ------------------------------------
import json
from datetime import date as _date
from django.utils.timezone import localdate
from django.views.decorators.http import require_http_methods
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.http import HttpResponse, Http404
from django.conf import settings
from django.utils.text import slugify
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.page import PageMargins

from .models import Shipment, ShipmentItem
from .production_config import ALLOWED_COMBOS, ORDERED_ALIASES
from .models import Presentation, ProductionDisplay
def _pd_combos_active():
    """
    Devuelve [(presentation_name, size), ...] en el orden de ProductionDisplay (activos).
    """
    return list(
        ProductionDisplay.objects
        .filter(is_active=True)
        .select_related("presentation")
        .order_by("order", "presentation__name", "size")
        .values_list("presentation__name", "size")
    )

import unicodedata, re
# ---- Lista de clientes ----
clientes = [ 
    "La Cima Produce",
    "RC Organics",
    "GH Farms",
    "Gourmet Baja Farms",
    "GBF Farms",
]
LEGAL_CLIENT_NAME = {
"La Cima Produce": "La Cima Produce, S.P.R. DE R.L",
"RC Organics": "Empaque N.1 S. DE R.L. DE C.V.",
"GH Farms": "Empaque N.1 S. DE R.L. DE C.V.",  
"Gourmet Baja Farms": "Gourmet Baja Farms S. DE R.L. DE C.V.",
"GBF Farms": "GBF Farms S. DE R.L. DE C.V.",
}
LOGO_SLUG = {
    'RC': 'rc-organics',
    'LACIMA': 'la-cima-produce',
    'GH': 'gh-farms',
    'GOURMET': 'gourmet-baja-farms',
    'GBF': 'gbf-farms',
}

clientes_slug = [(c, slugify(c)) for c in clientes]

def _canon_client(s: str) -> str:
    """
    Normaliza el nombre del cliente a uno de:
      RC / LACIMA / GH / GOURMET / GBF
    Acepta variantes como:
      "La Cima Produce", "RC Organics", "Gourmet Baja Farms",
      razones sociales largas (S. DE R.L. DE C.V., etc.), puntos, espacios raros, etc.
    """
    if not s:
        return ""

    # 1) quitar acentos y bajar a ascii
    t = ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')
    t = t.lower()

    # 2) reemplazar separadores raros por espacio, colapsar espacios
    t = t.replace('&', ' and ')
    t = t.replace('.', ' ')
    t = re.sub(r'[^a-z0-9]+', ' ', t)
    t = ' '.join(t.split())

    # 3) detección por palabras clave
    #    (no depende de que venga exactamente "la cima" o "rc")
    if 'cima' in t:                   # la cima produce, la cima spr de rl...
        return 'LACIMA'
    if re.search(r'(^| )rc( |$)', t) or 'rcorganics' in t:
        return 'RC'
    if 'gourmet' in t:                # gourmet baja farms, ...
        return 'GOURMET'
    if re.search(r'(^| )gh( |$)', t) or 'ghfarms' in t:
        return 'GH'
    if 'gbf' in t:                    # gbf farms, ...
        return 'GBF'

    # 4) fallback (no reconocido)
    return t.upper()




# --- Empresas soportadas (normalización) ---
COMPANY_CANON = {
    'rc': 'RC',
    'lacima': 'LACIMA', 'la cima': 'LACIMA', 'la_cima': 'LACIMA',
    'gh': 'GH',
    'gourmet': 'GOURMET',
    'gbf': 'GBF',
    'AGRICOLA DH & G': 'AGRICOLA DH & G',
}
COMPANY_CHOICES = ['RC', 'LACIMA', 'GH', 'GOURMET', 'GBF', 'AGRICOLA DH & G']
DEFAULT_COMPANY = 'LACIMA'  # elige el que prefieras por defecto

import unicodedata
import re

def _canon_size(sz: str) -> str:
    s = (sz or "").strip().upper()
    import unicodedata, re
    s = ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')
    s = s.replace('-', ' ').replace('_',' ')
    s = re.sub(r'\s+', ' ', s).strip()
    aliases = {
        "X LARGE": "XLARGE",
        "X-LARGE": "XLARGE",
        "XL": "XLARGE",
        "XLG": "XLARGE",
        "LGE": "LARGE",
        "LG": "LARGE",
        "STD": "STANDARD",
        "STANDAR": "STANDARD",
        "STANDART": "STANDARD",
    }
    s = aliases.get(s, s)
    return s.replace(" ", "")  # queda "XLARGE", "STANDARD", etc.
from decimal import Decimal, ROUND_HALF_UP

SPECIAL_EQ11_ROUND_CLIENTS = {"AGRICOLA DH & G", "BAJA MIST"}

def _canon_company(name: str | None) -> str | None:
    if not name:
        return None
    n = (name or "").strip().upper()
    if n in {"AGRICOLA DH & G", "BAJA MIST"}:
        return "AGRICOLA DH & G"
    return n

def _round_half_up_to_int(x) -> int:
    return int(Decimal(str(x)).quantize(Decimal('0'), rounding=ROUND_HALF_UP))

def _canon_pair(pres: str, size: str):
    """(PRESENTACIÓN en MAYÚSCULAS compacta, TAMAÑO canónico)"""
    return ((pres or "").strip().upper(), _canon_size(size))


def canon_company(s: str | None) -> str | None:
    if not s:
        return None
    k = (s or '').strip().lower().replace('_', ' ').replace('-', ' ')
    return COMPANY_CANON.get(k)

def company_slug(canon: str | None) -> str:
    return (canon or 'all').lower()



PROD_DIR = Path(settings.BASE_DIR) / "data" / "production"
PROD_DIR.mkdir(parents=True, exist_ok=True)
def _row_has_numbers(saved_row, per4):
    """Devuelve True si hay algún número > 0 ya sea en manuales o en las 4 columnas de embarques."""
    if not saved_row:
        saved_row = {}
    manuales = [
        saved_row.get("exist_prev", 0),
        saved_row.get("exist_almacen", 0),
        saved_row.get("debe", 0),
        saved_row.get("pago", 0),
        saved_row.get("presto", 0),
        saved_row.get("le_pagaron", 0),
        saved_row.get("prod_dia", 0),
    ]
    return any((x or 0) for x in manuales + list(per4 or []))


# === HELPERS DE DISEÑO PARA SEMANAL ===
def _week_single_company_sheet(wb, monday, sunday, empresa_label, rows_iter):
    """Hoja estilo 'anterior' para una sola empresa."""
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.utils import get_column_letter
    import os

    ws = wb.create_sheet(f"Semana - {empresa_label}")
    ws.title = f"Semana - {empresa_label}"

    # Estilos
    title_font   = Font(name='Calibri', size=18, bold=True, color="3C78D8")
    subtitle_font= Font(name='Calibri', size=11, italic=True, color="6D6D6D")
    th_font      = Font(name='Calibri', size=12, bold=True, color="FFFFFF")
    th_fill      = PatternFill("solid", fgColor="225577")
    thin_border  = Border(
        left=Side(style='thin', color='AAAAAA'),
        right=Side(style='thin', color='AAAAAA'),
        top=Side(style='thin', color='AAAAAA'),
        bottom=Side(style='thin', color='AAAAAA'),
    )

    # Logo
    LOGO_SLUG_MAP = {
        "La Cima Produce": "la-cima-produce",
        "RC Organics": "gh-farms",
        "GH Farms": "gh-farms",
        "Gourmet Baja Farms": "gourmet-baja-farms",
        "GBF Farms": "gbf-farms",
        "AGRICOLA DH & G": "agricola",
    }
    logo_slug = LOGO_SLUG_MAP.get(empresa_label)
    if logo_slug:
        logo_path = os.path.join(settings.BASE_DIR, "static", "logos", f"{logo_slug}.png")
        if os.path.exists(logo_path):
            img = XLImage(logo_path)
            target_h = 120
            scale = target_h / img.height
            img.width  = int(img.width * scale)
            img.height = int(img.height * scale)
            ws.add_image(img, "A1")

    # Título
    ws.merge_cells(start_row=1, start_column=4, end_row=2, end_column=10)  # D1:J2
    title = f"Resumen semanal {monday.strftime('%d/%m/%Y')} – {sunday.strftime('%d/%m/%Y')} – {empresa_label}"
    tcell = ws.cell(row=1, column=4, value=title)
    tcell.font = title_font
    tcell.alignment = Alignment(horizontal="left", vertical="center")

    ws.cell(row=3, column=4, value="Detalle de embarques y totales por semana") \
      .font = subtitle_font

    # Encabezados
    headers = ["N° EMBARQUE", "N° FACTURA", "FECHA", "PRESENTACIÓN", "TAMAÑO",
               "CANTIDAD", "EQUIV. 11 LBS", "IMPORTE ($)"]
    row = 7
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = th_font; c.fill = th_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border
    ws.row_dimensions[row].height = 22
    row += 1

    # Ancho cómodo (grande)
    for idx in range(1, len(headers)+1):
        ws.column_dimensions[get_column_letter(idx)].width = 24
    ws.freeze_panes = "A8"

    total_boxes = 0
    total_eq    = 0.0
    total_amt   = 0.0

    # rows_iter: iterable de dicts con keys: num_emb, inv, date, pres, size, qty, eq, amt
    for r in rows_iter:
        ws.cell(row=row, column=1, value=r["num_emb"])
        ws.cell(row=row, column=2, value=r["inv"])
        ws.cell(row=row, column=3, value=r["date"].strftime('%d/%m/%Y'))
        ws.cell(row=row, column=4, value=r["pres"])
        ws.cell(row=row, column=5, value=r["size"])
        ws.cell(row=row, column=6, value=r["qty"])
        ws.cell(row=row, column=7, value=round(r["eq"], 2))
        ws.cell(row=row, column=8, value=round(r["amt"], 2))
        for c in range(1, 9):
            ws.cell(row=row, column=c).border = thin_border
            if c in (6, 7, 8):
                ws.cell(row=row, column=c).alignment = Alignment(horizontal="right")
            else:
                ws.cell(row=row, column=c).alignment = Alignment(horizontal="center")
        total_boxes += r["qty"]; total_eq += r["eq"]; total_amt += r["amt"]
        row += 1

    # Totales
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.cell(row=row, column=1, value="TOTALES:") \
      .font = Font(name='Calibri', size=12, bold=True, color="225577")
    ws.cell(row=row, column=6, value=total_boxes).font = Font(bold=True)
    ws.cell(row=row, column=7, value=round(total_eq, 2)).font = Font(bold=True)
    ws.cell(row=row, column=8, value=round(total_amt, 2)).font = Font(bold=True)
    return ws


def _week_multi_company_sheet(wb, monday, sunday, company_sections, grand_totals):
    """Hoja multiempresa (nueva) con secciones por empresa y totales generales."""
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    ws = wb.create_sheet("Resumen por empresa")

    title_font   = Font(name='Calibri', size=18, bold=True, color="3C78D8")
    th_font      = Font(name='Calibri', size=12, bold=True, color="FFFFFF")
    th_fill      = PatternFill("solid", fgColor="225577")
    thin_border  = Border(
        left=Side(style='thin', color='AAAAAA'),
        right=Side(style='thin', color='AAAAAA'),
        top=Side(style='thin', color='AAAAAA'),
        bottom=Side(style='thin', color='AAAAAA'),
    )

    r = 1
    ws.cell(row=r, column=1, value=f"Resumen semanal {monday:%d/%m/%Y} – {sunday:%d/%m/%Y}").font = title_font
    r += 2

    # company_sections: iterable de (empresa_label, pres_info, totals, detalle_rows)
    for comp, write_func in company_sections:
        r = write_func(ws, r, thin_border, th_font, th_fill, comp)

    # Totales generales
    ws.cell(row=r, column=1, value="TOTALES GENERALES").font = Font(size=14, bold=True); r += 1
    headers_tot = ["Métrica", "Total"]
    for c, txt in enumerate(headers_tot, start=1):
        cell = ws.cell(row=r, column=c, value=txt)
        cell.font = th_font; cell.fill = th_fill
    r += 1
    for label, val in grand_totals:
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=val)
        r += 1
    return ws

def load_prod(d: date):
    p = _prod_path(d)
    if p.exists():
        with p.open("r", encoding="utf-8") as f:
            return json.load(f)
    return None

def save_prod(d: date, payload: dict):
    p = _prod_path(d)
    with p.open("w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)

def make_key(pres: str, size: str) -> str:
    return f"{slugify(pres)}__{slugify(size)}"

@login_required
def production_yesterday(request):
    today = date.today()
    yday = today - timedelta(days=1)
    data = load_prod(yday)

    ctx = {
        "prod_date": yday,
        "data": data,  # puede ser None si nunca se guardó ayer
    }
    return render(request, "empaques/production_yesterday.html", ctx)

def _ensure_prod_dir():
    PROD_DIR.mkdir(parents=True, exist_ok=True)

def _prod_path(d: _date, company: str | None = None) -> Path:
    subdir = PROD_DIR / company_slug(company)
    subdir.mkdir(parents=True, exist_ok=True)
    return subdir / f"{d.isoformat()}.json"

def _load_prod(d: _date, company: str | None = None):
    p = _prod_path(d, company)
    if p.exists():
        return json.loads(p.read_text(encoding="utf-8"))
    return None


def _day_increments(fday, empresa):
    """
    Devuelve (cajas_trabajadas, eq11_dia) calculados SOLO con datos del día y del día previo real.
    No usa ni confía en 'acum_*' guardados.
    """
    data = _load_prod(fday, empresa) or {}

    exist_piso_hoy  = int(data.get("exist_piso_hoy") or 0)
    recibidas       = int(data.get("cajas_campo_recibidas") or 0)

    prev = _load_last_before(fday, empresa) or {}
    exist_piso_ayer = int(prev.get("exist_piso_hoy") or 0)

    cajas_trabajadas = (exist_piso_ayer + recibidas) - exist_piso_hoy
    if cajas_trabajadas < 0:
        cajas_trabajadas = 0

    # Eq. 11 lb del día = sum(prod_dia * factor_presentación) por cada fila
    rows = (data.get("rows") or {})
    pres_cf = { (n or "").strip().upper(): float(cf or 1.0)
                for (n, cf) in Presentation.objects.values_list("name", "conversion_factor") }
    eq11_dia = 0.0
    for key, r in rows.items():
        pres = key.split("|", 1)[0] if "|" in key else key
        cf = pres_cf.get((pres or "").strip().upper(), 1.0)
        prod_dia = int(r.get("prod_dia", 0))
        eq11_dia += prod_dia * cf

    return int(cajas_trabajadas or 0), float(eq11_dia or 0.0)

def _season_acum_until(d, empresa):
    """
    Suma incremental estricta desde el inicio de temporada hasta 'd' INCLUIDO.
    Ignora por completo cualquier 'acum_*' guardado en los JSON.
    """
    start, _ = _season_bounds(d)
    subdir = PROD_DIR / company_slug(empresa)
    if not subdir.exists():
        return (0, 0.0)

    acum_cosechadas = 0
    acum_empacadas  = 0.0

    for fname in sorted(os.listdir(subdir)):
        if not fname.endswith(".json"):
            continue
        try:
            fday = _date.fromisoformat(fname[:-5])
        except ValueError:
            continue
        if fday < start or fday > d:
            continue

        ct, eq11 = _day_increments(fday, empresa)
        acum_cosechadas += ct
        acum_empacadas  += eq11

    return int(acum_cosechadas or 0), float(acum_empacadas or 0.0)



def _save_prod(d: _date, payload: dict, company: str | None = None):
    p = _prod_path(d, company)
    p.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def _combo_key(pres, size):
    """Alias para mantener compatibilidad: misma normalización que _canon_pair."""
    return _canon_pair(pres, size)

def _combos_from_db():
    """
    TODAS las combinaciones (presentación, tamaño) que existen en la BD,
    deduplicadas y normalizadas, ordenadas.
    """
    qs = (
        ShipmentItem.objects
        .select_related('presentation')
        .values_list('presentation__name', 'size')
        .distinct()
    )
    combos = {_canon_pair(p, s) for (p, s) in qs}
    return sorted(combos)  # lista de tuplas [(PRES, SIZE), ...]



def _group_shipments_by_combo(target_date, empresa=None):
    """
    Devuelve:
      - cols: ids (o trackings) de hasta 4 embarques del día para esa empresa
      - totals: {(PRES,SIZE): total_del_día}
      - per_ship: {(PRES,SIZE): [q1,q2,q3,q4]}
      - eq11_by_combo: {(PRES,SIZE): eq_11_lbs_total}
    """
    from django.db.models import Q

    qs = (Shipment.objects
          .filter(date=target_date)
          .order_by("id")
          .prefetch_related("items", "items__presentation"))

    ships = list(qs)
    if not ships:
        return [], {}, {}, {}

    emp = _canon_client(empresa) if empresa else None

    def shipment_matches_company(s):
        if not emp:
            return bool(s.items.all())
        # 1) ¿Algún item tiene cliente = empresa?
        item_match = any(_canon_client(it.cliente) == emp for it in s.items.all())
        if item_match:
            return True
        # 2) Respaldo: flags a nivel Shipment (si los usas)
        flag = False
        if emp == "LACIMA":
            flag = bool(getattr(s, "order_lacima", None))
        elif emp == "RC":
            flag = bool(getattr(s, "order_rc", None))
        elif emp == "GOURMET":
            flag = bool(getattr(s, "order_gourmet", None))
        elif emp == "GBF":
            flag = bool(getattr(s, "order_gbf", None))
        elif emp == "GH":
            flag = bool(getattr(s, "order_gh", None))
        return flag

    candidate_ids = [s.id for s in ships if shipment_matches_company(s)]
    if not candidate_ids:
        return [], {}, {}, {}

    cols = candidate_ids[:4]

    totals = {}
    per_ship = {}
    eq11_by_combo = {}

    for s in ships:
        col_idx = cols.index(s.id) if s.id in cols else None

        # ¿este embarque fue “asignado” a la empresa por el flag?
        assigned_to_emp = shipment_matches_company(s) if emp else True

        for it in s.items.select_related("presentation").all():
            if emp:
                # usa el item si coincide cliente, o si el item no tiene cliente
                # pero el embarque está asignado a esa empresa por flag
                if _canon_client(it.cliente) != emp and not (not it.cliente and assigned_to_emp):
                    continue

            pres = getattr(it.presentation, "name", "")
            size = it.size
            key  = _canon_pair(pres, size)  # <-- usa tamaño canónico
            qty  = int(it.quantity or 0)
            cf   = float(getattr(it.presentation, "conversion_factor", 1.0))

            totals[key] = totals.get(key, 0) + qty
            eq11_by_combo[key] = eq11_by_combo.get(key, 0.0) + qty * cf

            if key not in per_ship:
                per_ship[key] = [0, 0, 0, 0]
            if col_idx is not None:
                per_ship[key][col_idx] += qty

    return cols, totals, per_ship, eq11_by_combo

from datetime import date as _date
from django.utils import timezone

# Ajusta aquí tu inicio de temporada real
SEASON_START_MONTH = 4
SEASON_START_DAY   = 1

def _season_bounds(hoy=None):
    hoy = hoy or timezone.localdate()
    season_start = _date(hoy.year if hoy.month >= SEASON_START_MONTH else hoy.year - 1,
                         SEASON_START_MONTH, SEASON_START_DAY)
    return season_start, hoy

def _load_last_before(d, empresa):
    subdir = PROD_DIR / company_slug(empresa)
    if not subdir.exists():
        return {}
    best_day = None
    for fname in os.listdir(subdir):
        if not fname.endswith(".json"):
            continue
        try:
            fday = _date.fromisoformat(fname[:-5])
        except ValueError:
            continue
        if fday < d and (best_day is None or fday > best_day):
            best_day = fday
    return _load_prod(best_day, empresa) or {} if best_day else {}

def _effective_acum_base(day, empresa, saved_today):
    """
    Devuelve (base_cosechadas, base_empacadas) para 'hasta AYER'.
    Si 'use_manual_acum' está activo en el día, usa campos manuales del día.
    Si no, usa el último día previo con JSON. Si tampoco hay, usa los manuales del propio día (primer día).
    """
    prev = _load_prod(day - timedelta(days=1), empresa) or _load_last_before(day, empresa) or {}
    base_c_auto = prev.get("acum_cosechadas")
    base_e_auto = prev.get("acum_empacadas")

    man_c = saved_today.get("acum_cosechadas_ayer")
    man_e = saved_today.get("acum_empacadas_ayer")
    use_manual = bool(saved_today.get("use_manual_acum"))

    if use_manual:
        return int(man_c or 0), float(man_e or 0.0)

    base_c = base_c_auto if base_c_auto is not None else int(saved_today.get("acum_cosechadas_ayer", 0))
    base_e = base_e_auto if base_e_auto is not None else float(saved_today.get("acum_empacadas_ayer", 0.0))
    return int(base_c or 0), float(base_e or 0.0)







def _ordered_combos():
    # Ordena según ORDERED_ALIASES (si no está, lo manda al final)
    with_index = []
    for pres, size in ALLOWED_COMBOS:
        with_index.append((ORDERED_ALIASES.get((pres, size), 9999), pres, size))
    with_index.sort(key=lambda x: x[0])
    return [(p, s) for _, p, s in with_index]

from datetime import timedelta, date as _date
from django.utils.text import slugify
from django.http import Http404, HttpResponse
from django.views.decorators.http import require_http_methods
from django.contrib.auth.decorators import login_required


def _all_combos_from_db():
    pres_list = [ (p or "").strip().upper()
                  for p in Presentation.objects.values_list('name', flat=True)
                  if (p or "").strip() ]

    sizes_db = list(ShipmentItem.objects.values_list('size', flat=True).distinct())
    sizes_allowed = [s for (_, s) in ALLOWED_COMBOS]

    size_set = set()
    for s in sizes_db + sizes_allowed:
        cs = _canon_size(s)
        if cs:
            size_set.add(cs)

    default_size_order = ["JUMBO", "XLARGE", "LARGE", "STANDARD", "SMALL"]
    order_map = {name: i for i, name in enumerate(default_size_order, start=1)}
    ordered_sizes = sorted(size_set, key=lambda s: order_map.get(s, 999))

    combos = [ ( (p or "").strip().upper(), cs ) for p in pres_list for cs in ordered_sizes ]
    return combos
def _display_combos():
    """
    Devuelve la lista [(presentación, tamaño), ...] desde el catálogo admin.
    """
    qs = (ProductionDisplay.objects
          .filter(is_active=True)
          .select_related("presentation")
          .order_by("order", "presentation__name", "size"))
    return [(d.presentation.name, d.size) for d in qs]

# Helpers arriba del view:
from decimal import Decimal, ROUND_HALF_UP

SPECIAL_EQ11_ROUND_CLIENTS = {"AGRICOLA DH & G", "BAJA MIST"}

def _canon_company_label(name: str | None) -> str:
    if not name:
        return ""
    n = str(name).strip().lower()
    if "cima" in n:
        return "La Cima Produce"
    if n.startswith("rc"):
        return "RC Organics"
    if "gourmet" in n:
        return "Gourmet Baja Farms"
    if n == "gbf" or "gbf" in n:
        return "GBF Farms"
    if "gh" in n:
        return "GH Farms"
    if "agricola" in n or "dh & g" in n or "dhg" in n or "baja mist" in n:
        return "AGRICOLA DH & G"   # ← tu cliente especial
    return name.strip()

def _client_order_for(shipment, company_label):
    lbl = (company_label or "").lower()
    if "cima" in lbl:
        return getattr(shipment, "order_lacima", None)
    if lbl.startswith("rc"):
        return getattr(shipment, "order_rc", None)
    if "gh" in lbl and "rc" not in lbl:
        return getattr(shipment, "order_gh", None)
    if "gourmet" in lbl:
        return getattr(shipment, "order_gourmet", None)
    if "gbf" in lbl:
        return getattr(shipment, "order_gbf", None)
    if "agricola" in lbl or "dhg" in lbl:
        return getattr(shipment, "order_dhg", None)
    return None

def _round_half_up_to_int(x: float | Decimal) -> int:
    return int(Decimal(str(x)).quantize(Decimal('0'), rounding=ROUND_HALF_UP))

def _iter_company_items(embarques, empresa_filter: str | None):
    """
    Genera (empresa_normalizada, s, it) para cada item que encaje con el filtro de empresa.
    Si empresa_filter es None (o 'general'), entrega TODOS y agrupa por 'cliente'.
    """
    want_general = (not empresa_filter) or (empresa_filter.lower() == "general")
    for s in embarques:
        for it in s.items.all():
            comp = _canon_company_label(getattr(it, "cliente", None))
            if want_general:
                yield comp, s, it
            else:
                if _canon_company_label(empresa_filter) == comp:
                    yield comp, s, it

def _compute_company_summary(company_name, tuples_cs):
    """
    Recibe todos los (s, it) de LA empresa y regresa:
    - presentaciones_info: {(pres,size): {'cajas': int, 'dinero': float}}
    - totales: dict con 'total_cajas', 'total_eq11', 'total_dinero' (ajustado si es AGRICOLA)
    - detalle_items: lista [(fecha, tracking, factura, pres, size, qty, importe)]
    """
    from collections import defaultdict
    presentaciones_info = collections.defaultdict(lambda: {'cajas': 0, 'dinero': 0.0})
    detalle = []
    total_cajas = 0
    total_eq11  = 0.0
    total_dinero_normal = 0.0  # suma normal (precio * qty)

    for s, it in tuples_cs:
        pres = getattr(it.presentation, "name", "")
        size = it.size
        qty  = int(it.quantity or 0)
        price = float(getattr(it.presentation, "price", 0.0))
        cf    = float(getattr(it.presentation, "conversion_factor", 1.0))
        importe = qty * price

        presentaciones_info[(pres, size)]['cajas']  += qty
        presentaciones_info[(pres, size)]['dinero'] += importe

        total_cajas += qty
        total_eq11  += qty * cf
        total_dinero_normal += importe

        detalle.append((
            s.date.strftime("%d/%m/%Y"),
            s.tracking_number,
            s.invoice_number,
            pres, size, qty, round(importe, 2)
        ))

    # Ajuste SOLO para AGRICOLA: redondeo HALF-UP del total Eq.11 → * 3.40
    if _canon_company_label(company_name).upper() in SPECIAL_EQ11_ROUND_CLIENTS:
        eq11_billable = _round_half_up_to_int(total_eq11)
        total_dinero_final = float(Decimal('3.40') * Decimal(eq11_billable))
    else:
        total_dinero_final = total_dinero_normal

    return presentaciones_info, {
        'total_cajas': total_cajas,
        'total_eq11': round(total_eq11, 2),
        'total_dinero': round(total_dinero_final, 2),
        'total_dinero_normal': round(total_dinero_normal, 2),  # por si lo quieres mostrar en notas
    }, detalle

def _write_company_section(ws, start_row, border, th_font, th_fill, company_name,
                           presentaciones_info, totals, detalle):
    """
    Escribe: Título empresa, tabla de presentaciones, totales de empresa,
             tabla de detalle (opcional).
    Devuelve el 'r' siguiente.
    """
    from openpyxl.styles import Alignment, Font
    r = start_row
    title_font = Font(name="Calibri", size=14, bold=True)

    # Título empresa
    ws.cell(row=r, column=1, value=f"EMPRESA: {company_name}").font = title_font
    r += 1

    # Tabla de presentaciones
    headers_pres = ["Presentación", "Tamaño", "Total cajas", "Total dinero"]
    for c, txt in enumerate(headers_pres, start=1):
        cell = ws.cell(row=r, column=c, value=txt)
        cell.font = th_font; cell.fill = th_fill
        cell.alignment = Alignment(horizontal="center"); cell.border = border
    r += 1

    if presentaciones_info:
        for (n_pres, sz), info in sorted(presentaciones_info.items()):
            ws.cell(row=r, column=1, value=n_pres).border = border
            ws.cell(row=r, column=2, value=sz).border = border
            ws.cell(row=r, column=3, value=info['cajas']).border = border
            ws.cell(row=r, column=4, value=round(info['dinero'], 2)).border = border
            for c in range(1, 5):
                ws.cell(row=r, column=c).alignment = Alignment(horizontal="center")
            r += 1
    else:
        ws.cell(row=r, column=1, value="(Sin datos)").border = border
        r += 1

    # Totales de la empresa
    r += 1
    ws.cell(row=r, column=1, value="TOTAL CAJAS").border = border
    ws.cell(row=r, column=2, value=totals['total_cajas']).border = border
    r += 1
    ws.cell(row=r, column=1, value="TOTAL Eq. 11 lbs").border = border
    ws.cell(row=r, column=2, value=totals['total_eq11']).border = border
    r += 1
    ws.cell(row=r, column=1, value="TOTAL IMPORTE ").border = border
    ws.cell(row=r, column=2, value=totals['total_dinero']).border = border
    r += 2

    # (Opcional) Detalle
    headers_det = ["Fecha", "N° Embarque", "N° Factura", "Presentación", "Tamaño", "Cantidad", "Importe"]
    for c, txt in enumerate(headers_det, start=1):
        cell = ws.cell(row=r, column=c, value=txt)
        cell.font = th_font; cell.fill = th_fill
        cell.alignment = Alignment(horizontal="center"); cell.border = border
    r += 1

    for fila in sorted(detalle, key=lambda x: (x[0], x[1], x[3], x[4])):  # ordena por fecha/embarque/pres/size
        for c, v in enumerate(fila, start=1):
            ws.cell(row=r, column=c, value=v).border = border
            ws.cell(row=r, column=c).alignment = Alignment(horizontal="center")
        r += 1

    r += 2
    return r
def _weekly_general_build(weeks_qs):
    """Workbook con:
       - 'Resumen por empresa' (secciones por empresa)
       - 'GENERAL' (matriz una fila: Semana, Rango, columnas por empresa, TOTAL_EQ11)
    """
    wb = Workbook()
    th_font = Font(name='Calibri', size=12, bold=True, color="FFFFFF")
    th_fill = PatternFill("solid", fgColor="225577")
    thin = Border(left=Side(style='thin', color='AAAAAA'),
                  right=Side(style='thin', color='AAAAAA'),
                  top=Side(style='thin', color='AAAAAA'),
                  bottom=Side(style='thin', color='AAAAAA'))

    # 1) Hoja por empresa
    ws = wb.active
    ws.title = "Resumen por empresa"
    ws.cell(row=1, column=1, value="Resumen semanal – General por empresa").font = Font(size=18, bold=True, color="3C78D8")
    r = 3

    from collections import defaultdict
    company_pairs = defaultdict(list)  # lbl -> [(shipment,item),...]
    for comp, s, it in _iter_company_items(weeks_qs, None):
        company_pairs[comp].append((s, it))

    grand_cajas = grand_eq = grand_amt = 0.0
    for comp in sorted(company_pairs.keys(), key=lambda x: x.upper()):
        pres_info, totals, detalle = _compute_company_summary(comp, company_pairs[comp])
        r = _write_company_section(ws, r, thin, th_font, th_fill, comp, pres_info, totals, detalle)
        grand_cajas += totals["total_cajas"]
        grand_eq    += totals["total_eq11"]
        grand_amt   += totals["total_dinero"]

    ws.cell(row=r, column=1, value="TOTALES GENERALES").font = Font(size=14, bold=True); r += 1
    for label, val in [("TOTAL CAJAS", int(grand_cajas)),
                       ("TOTAL Eq. 11 lbs", round(grand_eq, 2)),
                       ("TOTAL IMPORTE", round(grand_amt, 2))]:
        ws.cell(row=r, column=1, value=label).border = thin
        ws.cell(row=r, column=2, value=val).border = thin
        r += 1

    # 2) Hoja matriz GENERAL
    ws2 = wb.create_sheet("GENERAL")
    empresas = sorted(company_pairs.keys(), key=lambda s: s.upper())
    headers = ["Semana", "Rango"] + empresas + ["TOTAL_EQ11"]
    for c, h in enumerate(headers, start=1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.font = th_font; cell.fill = th_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    if weeks_qs:
        d1 = min(s.date for s in weeks_qs)
        d2 = max(s.date for s in weeks_qs)
        week_label = f"{d1.isocalendar().week}"
        rango = f"{d1.strftime('%d/%m/%Y')} – {d2.strftime('%d/%m/%Y')}"
    else:
        week_label = ""; rango = ""

    # Importe por empresa + Eq11 total
    from collections import defaultdict as _dd
    per_company_amt = {e: 0.0 for e in empresas}
    comp_eq11 = _dd(float)
    total_eq11 = 0.0
    for comp, s, it in _iter_company_items(weeks_qs, None):
        qty = int(it.quantity or 0)
        cf  = float(getattr(it.presentation, "conversion_factor", 1.0))
        prc = float(getattr(it.presentation, "price", 0.0))
        total_eq11 += qty * cf
        comp_eq11[comp] += qty * cf
        per_company_amt[comp] += qty * prc

    # Ajuste especial por empresa
    for comp in empresas:
        if _canon_company_label(comp).upper() in SPECIAL_EQ11_ROUND_CLIENTS:
            per_company_amt[comp] = float(Decimal('3.40') * Decimal(_round_half_up_to_int(comp_eq11[comp])))

    row = 2
    ws2.cell(row=row, column=1, value=week_label)
    ws2.cell(row=row, column=2, value=rango)
    c = 3
    for emp in empresas:
        ws2.cell(row=row, column=c, value=round(per_company_amt[emp], 2)); c += 1
    ws2.cell(row=row, column=c, value=round(total_eq11, 2))

    # anchos
    ws2.column_dimensions['A'].width = 10
    ws2.column_dimensions['B'].width = 28
    for i in range(3, 3+len(empresas)+1):
        ws2.column_dimensions[chr(ord('A')+i-1)].width = 16

    return wb
def _month_weeks_iter(year:int, month:int):
    """Devuelve lista de (monday, sunday) que cubren TODAS las semanas que tocan ese mes."""
    import calendar
    from datetime import timedelta, date
    first = date(year, month, 1)
    last  = date(year, month, calendar.monthrange(year, month)[1])
    # ir al lunes de la semana del 'first'
    monday = first - timedelta(days=(first.weekday()))  # weekday: 0=Mon
    buckets = []
    while monday <= last:
        sunday = monday + timedelta(days=6)
        # si la semana toca el mes
        if monday.month == month or sunday.month == month:
            buckets.append((monday, sunday))
        monday += timedelta(days=7)
    return buckets

def _matrix_sheet_for_month(wb, year, month, empresa_filter):
    """Agrega hoja 'GENERAL SEMANAS' con filas por semana del mes (solo GENERAL)."""
    from openpyxl.styles import Font, Alignment, PatternFill
    from collections import defaultdict

    ws = wb.create_sheet("GENERAL SEMANAS")
    th_font = Font(name='Calibri', size=12, bold=True, color="FFFFFF")
    th_fill = PatternFill("solid", fgColor="225577")

    # Descubrimos empresas presentes en el mes
    # (usamos _canon_company_label para normalizar nombres)
    companies = set()
    for s in Shipment.objects.filter(date__year=year, date__month=month).prefetch_related('items','items__presentation'):
        for it in s.items.all():
            companies.add(_canon_company_label(getattr(it, "cliente", None)))
    empresas = sorted(companies, key=lambda x: x.upper())

    headers = ["Semana", "Rango"] + empresas + ["TOTAL_EQ11"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = th_font; cell.fill = th_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    row = 2
    for monday, sunday in _month_weeks_iter(year, month):
        qs = (Shipment.objects
              .filter(date__range=(monday, sunday))
              .order_by('date','id')
              .prefetch_related('items','items__presentation'))

        per_company_amt = {e: 0.0 for e in empresas}
        comp_eq11 = defaultdict(float)
        total_eq11 = 0.0

        for s in qs:
            for it in s.items.all():
                comp = _canon_company_label(getattr(it, "cliente", None))
                qty  = int(it.quantity or 0)
                cf   = float(getattr(it.presentation, "conversion_factor", 1.0))
                prc  = float(getattr(it.presentation, "price", 0.0))
                total_eq11 += qty * cf
                per_company_amt[comp] += qty * prc
                comp_eq11[comp] += qty * cf

        # Ajuste BAJA MIST/AGRICOLA
        for comp in empresas:
            if _canon_company_label(comp).upper() in SPECIAL_EQ11_ROUND_CLIENTS:
                per_company_amt[comp] = float(Decimal('3.40') * Decimal(_round_half_up_to_int(comp_eq11[comp])))

        week_no = monday.isocalendar().week
        ws.cell(row=row, column=1, value=week_no)
        ws.cell(row=row, column=2, value=f"{monday.strftime('%d/%m/%Y')} – {sunday.strftime('%d/%m/%Y')}")
        c = 3
        for emp in empresas:
            ws.cell(row=row, column=c, value=round(per_company_amt[emp], 2)); c += 1
        ws.cell(row=row, column=c, value=round(total_eq11, 2))
        row += 1

    # anchos
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 28
    for i in range(3, 3+len(empresas)+1):
        ws.column_dimensions[chr(ord('A')+i-1)].width = 16

def _matrix_sheet_for_year(wb, year):
    """Agrega hoja 'GENERAL MESES' con filas por MES del año (solo GENERAL)."""
    from collections import defaultdict
    from openpyxl.styles import Font, Alignment, PatternFill
    ws = wb.create_sheet("GENERAL MESES")
    th_font = Font(name='Calibri', size=12, bold=True, color="FFFFFF")
    th_fill = PatternFill("solid", fgColor="225577")

    # Empresas observadas en todo el año
    companies = set()
    qs_all = (Shipment.objects.filter(date__year=year)
              .order_by('date','id')
              .prefetch_related('items','items__presentation'))
    for s in qs_all:
        for it in s.items.all():
            companies.add(_canon_company_label(getattr(it, "cliente", None)))
    empresas = sorted(companies, key=lambda x: x.upper())

    headers = ["Mes", "Rango"] + empresas + ["TOTAL_EQ11"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = th_font; cell.fill = th_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    from datetime import date as _d, timedelta as _td
    row = 2
    for month in range(1,13):
        # rango del mes
        import calendar
        d1 = _d(year, month, 1)
        d2 = _d(year, month, calendar.monthrange(year, month)[1])
        qs = (Shipment.objects.filter(date__range=(d1, d2))
              .order_by('date','id')
              .prefetch_related('items','items__presentation'))

        per_company_amt = {e: 0.0 for e in empresas}
        comp_eq11 = defaultdict(float)
        total_eq11 = 0.0
        for s in qs:
            for it in s.items.all():
                comp = _canon_company_label(getattr(it, "cliente", None))
                qty  = int(it.quantity or 0)
                cf   = float(getattr(it.presentation, "conversion_factor", 1.0))
                prc  = float(getattr(it.presentation, "price", 0.0))
                total_eq11 += qty * cf
                per_company_amt[comp] += qty * prc
                comp_eq11[comp] += qty * cf

        for comp in empresas:
            if _canon_company_label(comp).upper() in SPECIAL_EQ11_ROUND_CLIENTS:
                per_company_amt[comp] = float(Decimal('3.40') * Decimal(_round_half_up_to_int(comp_eq11[comp])))

        ws.cell(row=row, column=1, value=month)
        ws.cell(row=row, column=2, value=f"{d1.strftime('%d/%m/%Y')} – {d2.strftime('%d/%m/%Y')}")
        c = 3
        for emp in empresas:
            ws.cell(row=row, column=c, value=round(per_company_amt[emp], 2)); c += 1
        ws.cell(row=row, column=c, value=round(total_eq11, 2))
        row += 1

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 28
    for i in range(3, 3+len(empresas)+1):
        ws.column_dimensions[chr(ord('A')+i-1)].width = 16



@login_required
@require_http_methods(["GET", "POST"])
def production_today(request):
    from .models import Presentation

    raw_empresa = request.GET.get("empresa") or request.POST.get("empresa")
    empresa = canon_company(raw_empresa) or DEFAULT_COMPANY

    # Fecha
    qdate = request.GET.get("date") or localdate().isoformat()
    try:
        prod_date = _date.fromisoformat(qdate)
    except ValueError:
        raise Http404("Fecha inválida")

    # Estado del día actual + 'ayer'
    saved_today = _load_prod(prod_date, empresa) or {}
    saved_yday  = _load_prod(prod_date - timedelta(days=1), empresa) or {}
    rows_yday   = saved_yday.get("rows", {})
    order_number = (saved_today.get("order_number") or "")

    # Embarques del día (filtrados por empresa)
    ship_cols, totals, per_ship, eq11_map = _group_shipments_by_combo(prod_date, empresa)

    # Combos a mostrar
    combos = _display_combos() or (_all_combos_from_db() or _ordered_combos())

    # Factores por presentación
    pres_cf = { (n or "").strip().upper(): float(cf or 1.0)
                for n, cf in Presentation.objects.values_list("name", "conversion_factor") }

    # --- Filas y total eq. 11 lb de la PRODUCCIÓN DEL DÍA ---
    rows = []
    total_eq11_produccion = 0.0

    for pres, size in combos:
        k_totals = _combo_key(pres, size)
        k_saved  = f"{pres}|{size}"
        k_form   = f"{slugify(pres)}__{slugify(size)}"

        exist_prev = int(rows_yday.get(k_saved, {}).get("exist_almacen", 0))
        row_today  = (saved_today.get("rows") or {}).get(k_saved, {})

        exist_almacen = int(row_today.get("exist_almacen", 0))
        debe          = int(row_today.get("debe", 0))
        pago          = int(row_today.get("pago", 0))
        presto        = int(row_today.get("presto", 0))
        le_pagaron    = int(row_today.get("le_pagaron", 0))

        per4       = per_ship.get(k_totals, [0, 0, 0, 0])
        total_emb  = totals.get(k_totals, 0)
        eq11_today = round(eq11_map.get(k_totals, 0.0), 2)

        prod_dia = exist_almacen - exist_prev - debe + pago + presto - le_pagaron + total_emb

        rows.append({
            "pres": pres, "size": size,
            "form_prefix": k_form,
            "per_ship": per4,
            "total_emb": total_emb,
            "exist_prev": exist_prev,
            "exist_almacen": exist_almacen,
            "debe": debe, "pago": pago, "presto": presto, "le_pagaron": le_pagaron,
            "prod_dia": prod_dia,
            "eq11": eq11_today,
        })

        cf = pres_cf.get((pres or "").strip().upper(), 1.0)
        total_eq11_produccion += (prod_dia or 0) * cf

    # --- Bloque inferior del día ---
    exist_piso_hoy  = int((saved_today or {}).get("exist_piso_hoy", 0))
    cajas_campo_recibidas = int((saved_today or {}).get("cajas_campo_recibidas", 0))
    exist_piso_ayer = int((saved_yday  or {}).get("exist_piso_hoy", 0))
    total_cajas_trabajar   = exist_piso_ayer + cajas_campo_recibidas
    cajas_campo_trabajadas = total_cajas_trabajar - exist_piso_hoy

    # --- Base efectiva "hasta AYER" ---
    base_c_hayer, base_e_hayer = _effective_acum_base(prod_date, empresa, saved_today)

    # --- Acumulados mostrados (con HOY) ---
    total_cosechadas_acumulado = int(base_c_hayer) + int(cajas_campo_trabajadas or 0)
    total_empacadas_acumulado  = float(base_e_hayer) + float(total_eq11_produccion or 0.0)
    factor_global = (total_empacadas_acumulado / total_cosechadas_acumulado) if total_cosechadas_acumulado else 0.0

    # --- POST: Guardar ---
    if request.method == "POST":
        def _num(prefix, name):
            raw = request.POST.get(f"{prefix}__{name}", "0")
            try:
                return int(raw or 0)
            except ValueError:
                try:
                    return float(raw or 0)
                except ValueError:
                    return 0

        def _num_global(name):
            raw = request.POST.get(name, "0")
            try:
                return int(raw or 0)
            except ValueError:
                try:
                    return float(raw or 0)
                except ValueError:
                    return 0

        new_rows = {}
        total_eq11_produccion_post = 0.0
        for pres, size in combos:
            k_saved  = f"{pres}|{size}"
            k_totals = _combo_key(pres, size)
            k_form   = f"{slugify(pres)}__{slugify(size)}"

            exist_prev    = int(rows_yday.get(k_saved, {}).get("exist_almacen", 0))
            exist_almacen = _num(k_form, "exist_almacen")
            debe          = _num(k_form, "debe")
            pago          = _num(k_form, "pago")
            presto        = _num(k_form, "presto")
            le_pagaron    = _num(k_form, "le_pagaron")

            total_emb = totals.get(k_totals, 0)
            prod_dia  = exist_almacen - exist_prev - debe + pago + presto - le_pagaron + total_emb

            new_rows[k_saved] = {
                "exist_prev": exist_prev,
                "exist_almacen": exist_almacen,
                "debe": debe, "pago": pago, "presto": presto, "le_pagaron": le_pagaron,
                "prod_dia": prod_dia,
            }

            cf = pres_cf.get((pres or "").strip().upper(), 1.0)
            total_eq11_produccion_post += (prod_dia or 0) * cf

        exist_piso_hoy_post        = _num_global("exist_piso_hoy")
        cajas_campo_recibidas_post = _num_global("cajas_campo_recibidas")
        exist_piso_ayer_post       = int((saved_yday or {}).get("exist_piso_hoy", 0))
        total_cajas_trabajar_post  = exist_piso_ayer_post + cajas_campo_recibidas_post
        cajas_campo_trabajadas_post = total_cajas_trabajar_post - exist_piso_hoy_post

        # Manual vs automático
        use_manual_acum = bool(request.POST.get("use_manual_acum"))
        man_c = int(_num_global("acum_cosechadas_ayer") or 0)
        man_e = float(_num_global("acum_empacadas_ayer") or 0.0)

        if use_manual_acum:
            base_c = man_c
            base_e = man_e
        else:
            prev = _load_prod(prod_date - timedelta(days=1), empresa) or _load_last_before(prod_date, empresa) or {}
            base_c = prev.get("acum_cosechadas")
            base_e = prev.get("acum_empacadas")
            if base_c is None:
                base_c = int(saved_today.get("acum_cosechadas_ayer", 0))
            if base_e is None:
                base_e = float(saved_today.get("acum_empacadas_ayer", 0.0))
            base_c = int(base_c or 0)
            base_e = float(base_e or 0.0)

        new_acum_cosechadas = base_c + int(cajas_campo_trabajadas_post or 0)
        new_acum_empacadas  = base_e + float(total_eq11_produccion_post or 0.0)

        order_number_post = (request.POST.get("order_number") or "").strip()

        payload = {
            "date": prod_date.isoformat(),
            "ship_cols": ship_cols,
            "rows": new_rows,
            "exist_piso_hoy": exist_piso_hoy_post,
            "cajas_campo_recibidas": cajas_campo_recibidas_post,

            # guardamos lo manual capturado y la bandera
            "acum_cosechadas_ayer": man_c,
            "acum_empacadas_ayer": round(man_e, 2),
            "use_manual_acum": use_manual_acum,

            # Acumulados de temporada (con HOY)
            "acum_cosechadas": int(new_acum_cosechadas),
            "acum_empacadas": round(float(new_acum_empacadas), 2),

            "order_number": order_number_post,
        }
        _save_prod(prod_date, payload, empresa)
        return redirect(f"{request.path}?date={prod_date.isoformat()}&empresa={empresa.lower()}")

    # Totales ligeros (si los usas en <tfoot>)
    totals_row = [0]*11
    totals_row_eq11 = [0.0]*11

    # Contexto
    ctx = {
        "prod_date": prod_date,
        "rows": rows,
        "ship_cols_labels": ship_cols,
        "empresa": empresa,
        "empresas": ["RC", "LACIMA", "GH", "GOURMET", "GBF", "AGRICOLA DH & G"],

        # Bloque inferior
        "exist_piso_ayer": exist_piso_ayer,
        "cajas_campo_recibidas": cajas_campo_recibidas,
        "total_cajas_trabajar": total_cajas_trabajar,
        "exist_piso_hoy": exist_piso_hoy,
        "cajas_campo_trabajadas": cajas_campo_trabajadas,

        # Manual capturado + checkbox
        "acum_cosechadas_ayer": int(saved_today.get("acum_cosechadas_ayer", 0)),
        "acum_empacadas_ayer": float(saved_today.get("acum_empacadas_ayer", 0.0)),
        "use_manual_acum": bool(saved_today.get("use_manual_acum")),

        # Calculado (base efectiva hasta AYER) para mostrar en los campos “Calculado”
        "acum_cosechadas_hasta_ayer": int(base_c_hayer),
        "acum_empacadas_hasta_ayer": round(float(base_e_hayer), 2),

        # Acumulado con HOY (la tarjeta grande)
        "total_cosechadas_acumulado": total_cosechadas_acumulado,
        "total_empacadas_acumulado": round(total_empacadas_acumulado, 2),
        "factor_global": round(factor_global, 4),

        "totals_row": totals_row,
        "totals_row_eq11": [round(x, 2) for x in totals_row_eq11],
        "order_number": order_number,
        "legal_client_name": LEGAL_CLIENT_NAME.get(empresa, empresa),
    }
    return render(request, "empaques/production_today.html", ctx)







@login_required
def production_days(request):
    raw_empresa = request.GET.get("empresa")
    empresa = canon_company(raw_empresa) or DEFAULT_COMPANY
    subdir = PROD_DIR / company_slug(empresa)
    subdir.mkdir(parents=True, exist_ok=True)

    files = []
    for fname in sorted(os.listdir(subdir)):
        if fname.endswith(".json"):
            files.append(fname[:-5])  # YYYY-MM-DD
    return render(request, "empaques/production_days.html", {
        "days": files,
        "empresa": empresa,
    })


@login_required
def production_xlsx(request, prod_date):
    """Excel de Producción del día (filtrado por empresa) — alineado con la vista."""
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.worksheet.page import PageMargins
    from openpyxl.drawing.image import Image as XLImage
    from datetime import timedelta as _timedelta
    from .models import Presentation
    import os

    # --- Empresa (canon) + nombre legal + slug de logo ---
    empresa = canon_company(request.GET.get("empresa")) or DEFAULT_COMPANY
    LEGAL_COMPANY = {
        "LACIMA":  "La Cima Produce, S.P.R. DE R.L",
        "RC":      "Empaque N.1 S. DE R.L. DE C.V.",
        "GH":      "Empaque N.1 S. DE R.L. DE C.V.",
        "GOURMET": "Gourmet Baja Farms S. DE R.L. DE C.V.",
        "GBF":     "GBF Farms S. DE R.L. DE C.V.",
        "AGRICOLA DH & G": "AGRICOLA DH & G",
    }
    LOGO_SLUG = {
        "LACIMA":  "la-cima-produce",
        "RC":      "rc-organics",
        "GH":      "gh-farms",
        "GOURMET": "gourmet-baja-farms",
        "GBF":     "gbf-farms",
        "AGRICOLA DH & G": "agricola",
    }
    legal_name = LEGAL_COMPANY.get(empresa, empresa)
    logo_slug  = LOGO_SLUG.get(empresa)

    # --- Fecha ---
    d = _date.fromisoformat(prod_date)

    # --- Estado guardado + embarques del día (FILTRADOS POR EMPRESA) ---
    saved = _load_prod(d, empresa) or {}
    ship_cols, totals, per_ship, _eq11_map = _group_shipments_by_combo(d, empresa)

    # --- Combos y factores de conversión ---
    combos = _display_combos()
    if not combos:
        combos = _all_combos_from_db() or _ordered_combos()

    pres_cf = { (n or "").strip().upper(): float(cf or 1.0)
                for n, cf in Presentation.objects.values_list("name", "conversion_factor") }

    # --- Excel base ---
    wb = Workbook()
    ws = wb.active
    ws.title = f"Prod. {empresa}"

    # Estilos básicos
    title_font = Font(name="Calibri", size=16, bold=True, color="2E67D1")
    th_font    = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    th_fill    = PatternFill("solid", fgColor="225577")
    border     = Border(
        left=Side(style='thin',  color='AAAAAA'),
        right=Side(style='thin', color='AAAAAA'),
        top=Side(style='thin',   color='AAAAAA'),
        bottom=Side(style='thin',color='AAAAAA'),
    )

    # --- Encabezado con logo/fecha/títulos/número de hoja ---
    # Logo en A1
    if logo_slug:
        logo_path = os.path.join(settings.BASE_DIR, "static", "logos", f"{logo_slug}.png")
        if os.path.exists(logo_path):
            try:
                img = XLImage(logo_path)
                target_h = 85
                scale = target_h / float(img.height)
                img.width  = int(img.width * scale)
                img.height = int(img.height * scale)
                ws.add_image(img, "A1")
            except Exception:
                pass

    # Fecha bajo el logo (A2)
    ws.cell(row=2, column=1, value=d.strftime("%d/%m/%Y")).font = Font(name="Calibri", size=11)
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="left", vertical="center")

    # H1: título
    ws.merge_cells("H1:K1")
    c_h1 = ws.cell(row=1, column=8, value="PRODUCCIÓN DIARIA DE EMPAQUE")
    c_h1.font = Font(name="Calibri", size=16, bold=True)
    c_h1.alignment = Alignment(horizontal="left", vertical="center")

    # H2: nombre legal
    ws.merge_cells("H2:K2")
    c_h2 = ws.cell(row=2, column=8, value=legal_name)
    c_h2.font = Font(name="Calibri", size=14, bold=True, color="444444")
    c_h2.alignment = Alignment(horizontal="left", vertical="center")

    # "Hoja num.:" (E1) + NÚMERO DE ORDEN (F1)
    ws.cell(row=1, column=5, value="Hoja num.:").font = Font(name="Calibri", size=11, bold=True)
    ws.cell(row=1, column=5).alignment = Alignment(horizontal="right", vertical="center")
    hoja_num = saved.get("order_number") or saved.get("hoja_num") or ""
    ws.cell(row=1, column=6, value=str(hoja_num)).font = Font(name="Calibri", size=12, bold=True, color="1F4E79")
    ws.cell(row=1, column=6).alignment = Alignment(horizontal="left", vertical="center")

    # Subtítulo
    r = 3
    ws.cell(row=r, column=1, value=f"Fecha – {d.strftime('%d/%m/%Y')}").font = title_font
    r += 2

    # --- Encabezados de tabla ---
    headers = [
        "Empaque", "Exist. anterior", "Producción del día", "Exist. almacén",
        "Emb. 1", "Emb. 2", "Emb. 3", "Emb. 4",
        "Debe", "Pago", "Presto", "Le pagaron",
    ]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.font = th_font
        cell.fill = th_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
    ws.row_dimensions[r].height = 28
    r += 1
    first_data_row = r

    # --- Totales auxiliares ---
    tot_norm = [0]*11
    tot_eq11 = [0.0]*11
    total_eq11_produccion = 0.0  # Eq. 11 lb de la PRODUCCIÓN DEL DÍA

    # --- Filas de datos ---
    for pres, size in combos:
        key_saved = f"{pres}|{size}"
        saved_row = (saved.get("rows") or {}).get(key_saved, {})
        k_norm    = _combo_key(pres, size)
        per4      = per_ship.get(k_norm, [0, 0, 0, 0])

        exist_prev   = saved_row.get("exist_prev", 0)
        prod_dia     = saved_row.get("prod_dia", 0)
        exist_alm    = saved_row.get("exist_almacen", 0)
        debe         = saved_row.get("debe", 0)
        pago         = saved_row.get("pago", 0)
        presto       = saved_row.get("presto", 0)
        le_pagaron   = saved_row.get("le_pagaron", 0)

        row_vals = [
            f"{pres} / {size}",
            exist_prev, prod_dia, exist_alm,
            per4[0], per4[1], per4[2], per4[3],
            debe, pago, presto, le_pagaron,
        ]
        for c, v in enumerate(row_vals, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[r].height = 22
        r += 1

        cf = pres_cf.get((pres or "").strip().upper(), 1.0)
        nums = row_vals[1:]
        for i, n in enumerate(nums):
            n = n or 0
            tot_norm[i] += n
            tot_eq11[i] += n * cf
        total_eq11_produccion += (prod_dia or 0) * cf

    if r == first_data_row:
        ws.cell(row=r, column=1, value="(Sin datos)").border = border
        for c in range(2, 13):
            ws.cell(row=r, column=c).border = border
        ws.row_dimensions[r].height = 22
        r += 1

    # --- Totales normales ---
    ws.cell(row=r, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=r, column=1).alignment = Alignment(horizontal="right", vertical="center")
    for i, n in enumerate(tot_norm, start=2):
        c = ws.cell(row=r, column=i, value=n)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border
    ws.row_dimensions[r].height = 24
    r += 1

    # --- Totales Eq. 11 lbs ---
    ws.cell(row=r, column=1, value="TOTAL (Eq. 11 lbs)").font = Font(bold=True)
    ws.cell(row=r, column=1).alignment = Alignment(horizontal="right", vertical="center")
    for i, n in enumerate(tot_eq11, start=2):
        c = ws.cell(row=r, column=i, value=round(n, 2))
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border
    ws.row_dimensions[r].height = 24
    r += 1

    # === Resumen inferior (incluye acumulados y factores) ===
    # Bloque del propio día (coincide con la vista)
    yday = d - _timedelta(days=1)
    saved_yday = _load_prod(yday, empresa) or {}
    exist_piso_ayer = int((saved_yday or {}).get("exist_piso_hoy", 0))
    exist_piso_hoy  = int((saved or {}).get("exist_piso_hoy", 0))
    cajas_campo_recibidas = int((saved or {}).get("cajas_campo_recibidas", 0))
    total_cajas_trabajar   = exist_piso_ayer + cajas_campo_recibidas
    cajas_campo_trabajadas = total_cajas_trabajar - exist_piso_hoy

    # MISMA base efectiva “hasta AYER” que la vista (respeta manual si use_manual_acum)
    base_c_hayer, base_e_hayer = _effective_acum_base(d, empresa, saved)

    # Si el día actual no guardó los acumulados, recalcular como en la vista usando la base efectiva + HOY
    acum_cosechadas = saved.get("acum_cosechadas")
    acum_empacadas  = saved.get("acum_empacadas")
    if acum_cosechadas is None or acum_empacadas is None:
        acum_cosechadas = int(base_c_hayer) + int(cajas_campo_trabajadas or 0)
        acum_empacadas  = float(base_e_hayer) + float(total_eq11_produccion or 0.0)

    factor_dia    = (total_eq11_produccion / cajas_campo_trabajadas) if cajas_campo_trabajadas else 0.0
    factor_global = (float(acum_empacadas) / int(acum_cosechadas)) if acum_cosechadas else 0.0

    r += 2
    label_font = Font(bold=True)
    pairs = [
        ("EXISTENCIA PISO DÍA ANTERIOR:", exist_piso_ayer),
        ("CAJAS DE CAMPO RECIBIDAS:", cajas_campo_recibidas),
        ("TOTAL CAJAS A TRABAJAR:", total_cajas_trabajar),
        ("CAJAS CAMPO TRABAJADAS:", cajas_campo_trabajadas),
        ("EXISTENCIA DE PISO HOY:", exist_piso_hoy),
        ("FACTOR DEL DÍA (Eq. 11 lb prod. / Cajas trabajadas):", round(factor_dia, 4)),
        ("ACUM. COSECHADAS (TEMPORADA):", int(acum_cosechadas or 0)),
        ("ACUM. EMPACADAS (TEMPORADA, Eq. 11 lb):", round(float(acum_empacadas or 0.0), 2)),
        ("FACTOR GLOBAL (Empacadas / Cosechadas):", round(float(factor_global or 0.0), 4)),
    ]
    for label, value in pairs:
        ws.cell(row=r, column=1, value=label).font = label_font
        ws.cell(row=r, column=2, value=value)
        ws.cell(row=r, column=1).border = border
        ws.cell(row=r, column=2).border = border
        ws.row_dimensions[r].height = 22
        r += 1

    # --- Anchos y alturas por defecto ---
    ws.sheet_format.defaultRowHeight = 22.5
    ws.row_dimensions[1].height = 40
    ws.row_dimensions[2].height = 26

    ws.column_dimensions["A"].width = 34
    for col in "BCDEFGHIJKL":
        ws.column_dimensions[col].width = 14

    # --- Setup de impresión ---
    last_col_idx = 12
    last_col_letter = chr(ord('A') + last_col_idx - 1)
    ws.print_area = f"A1:{last_col_letter}{r}"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.3, bottom=0.3, header=0.2, footer=0.2)
    ws.print_options.horizontalCentered = True

    # --- Respuesta ---
    out = BytesIO()
    wb.save(out); out.seek(0)
    resp = HttpResponse(out, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = f'attachment; filename="produccion_{empresa.lower()}_{d.isoformat()}.xlsx"'
    return resp







def production_list_view(request):
    # Alias a la lista de días guardados
    return production_days(request)

def es_capturista(user):
    return user.is_authenticated and user.groups.filter(name="capturista").exists()

@login_required
def post_login_redirect(request):
    """
    Redirige según el rol:
    - capturista -> nuevo embarque
    - demás -> lista de embarques
    """
    if es_capturista(request.user):
        return redirect("shipment_create")
    return redirect("shipment_list")

from django.utils.crypto import get_random_string
from django.db import transaction
from django.forms import inlineformset_factory
from django.contrib.auth.decorators import login_required, permission_required
@login_required
@permission_required('empaques.add_shipment', raise_exception=True)
def shipment_create(request):
    ItemFormSet = inlineformset_factory(
        Shipment,
        ShipmentItem,
        form=ShipmentItemForm,
        formset=BaseShipmentItemFormSet,
        extra=26,
        can_delete=True,
    )

    # Inicializa la lista de tokens usados en sesión
    used_tokens = request.session.get('used_form_tokens', [])
    if not isinstance(used_tokens, list):
        used_tokens = []
    request.session['used_form_tokens'] = used_tokens

    if request.method == 'POST':
        token = request.POST.get('form_token')
        used = set(request.session.get('used_form_tokens', []))

        # Si no hay token o ya se usó, no duplica
        if (not token) or (token in used):
            return redirect('shipment_list')

        form = ShipmentForm(request.POST)
        formset = ItemFormSet(request.POST)

        if form.is_valid() and formset.is_valid():
            # Marcamos token como usado ANTES de guardar (bloquea multi-clics)
            used.add(token)
            request.session['used_form_tokens'] = list(used)
            request.session.modified = True

            with transaction.atomic():
                shipment = form.save()
                formset.instance = shipment
                formset.save()

            return redirect('shipment_list')

        # Si hay errores, generamos un nuevo token
        new_token = get_random_string(32)
        request.session['current_form_token'] = new_token
        request.session.modified = True
        return render(request, 'empaques/shipment_form.html', {
            'form': form,
            'formset': formset,
            'form_token': new_token,
        })

    # GET: token nuevo
    token = get_random_string(32)
    request.session['current_form_token'] = token
    # Limpieza de tokens antiguos
    if len(used_tokens) > 100:
        request.session['used_form_tokens'] = used_tokens[-50:]
    request.session.modified = True

    form = ShipmentForm()
    formset = ItemFormSet()
    return render(request, 'empaques/shipment_form.html', {
        'form': form,
        'formset': formset,
        'form_token': token,
    })
from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from django.http import HttpResponse
from datetime import date
from collections import defaultdict

from .models import Shipment, ShipmentItem
EMPRESAS = ["RC", "LACIMA", "GH", "GOURMET", "GBF"]

def _empresa_param(request):
    raw = (request.GET.get("empresa") or "").strip()
    if not raw or raw.lower() == "general":
        return None  # None = todas
    return raw  # devuelve el nombre tal cual: "La Cima Produce", "RC Organics", etc.

@login_required
def shipment_list(request):
    from datetime import date, timedelta
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from decimal import Decimal, ROUND_HALF_UP

    def q2(x) -> Decimal:
        """Redondeo a 2 decimales con HALF_UP."""
        if not isinstance(x, Decimal):
            x = Decimal(str(x))
        return x.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

    def round_half_up_to_int(x) -> int:
        """Entero con HALF_UP (.49 ↓, .50/.51 ↑)."""
        if not isinstance(x, Decimal):
            x = Decimal(str(x))
        return int(x.to_integral_value(rounding=ROUND_HALF_UP))

    # --- Lista de embarques para la tabla ---
    shipments = Shipment.objects.order_by('-date', '-id')

    # --- Permiso para descargar/exportar ---
    can_download = request.user.has_perm('empaques.can_download_reports')

    # --- Parámetros de periodo ---
    try:
        year = int(request.GET.get('year') or date.today().year)
    except ValueError:
        year = date.today().year
    try:
        month = int(request.GET.get('month') or date.today().month)
    except ValueError:
        month = date.today().month

    # --------------------------
    # Helper: genera XLSX común
    # --------------------------
    

    def build_summary_xlsx(title_text, subtitle_text, embarques_qs):
        from collections import defaultdict
        items = ShipmentItem.objects.filter(shipment__in=embarques_qs).select_related('presentation')

        # Agregados por presentación/tamaño
        presentaciones_info = collections.defaultdict(lambda: {'cajas': 0, 'dinero': 0.0})
        for it in items:
            k = (it.presentation.name, it.size)
            presentaciones_info[k]['cajas'] += it.quantity
            presentaciones_info[k]['dinero'] += it.quantity * float(it.presentation.price)

        total_cajas = sum(i.quantity for i in items)
        total_eq_11lbs = sum(i.quantity * float(i.presentation.conversion_factor) for i in items)
        total_dinero = sum(i.quantity * float(i.presentation.price) for i in items)

        wb = Workbook()
        ws = wb.active
        ws.title = "Resumen"

        # Estilos
        title_font = Font(name="Calibri", size=16, bold=True, color="3C78D8")
        h_font = Font(name="Calibri", size=12, bold=True)
        th_font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
        th_fill = PatternFill("solid", fgColor="225577")
        border = Border(
            left=Side(style='thin', color='AAAAAA'),
            right=Side(style='thin', color='AAAAAA'),
            top=Side(style='thin', color='AAAAAA'),
            bottom=Side(style='thin', color='AAAAAA'),
        )

        r = 1
        ws.cell(row=r, column=1, value=title_text).font = title_font
        r += 1
        if subtitle_text:
            ws.cell(row=r, column=1, value=subtitle_text)
            r += 2

        # Presentaciones
        ws.cell(row=r, column=1, value="Presentaciones utilizadas").font = h_font
        r += 1
        headers_pres = ["Presentación", "Tamaño", "Total cajas", "Total dinero"]
        for c, txt in enumerate(headers_pres, start=1):
            cell = ws.cell(row=r, column=c, value=txt)
            cell.font = th_font
            cell.fill = th_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
        r += 1

        if presentaciones_info:
            for (n_pres, sz), info in sorted(presentaciones_info.items()):
                ws.cell(row=r, column=1, value=n_pres)
                ws.cell(row=r, column=2, value=sz)
                ws.cell(row=r, column=3, value=info['cajas'])
                ws.cell(row=r, column=4, value=round(info['dinero'], 2))
                for c in range(1, 5):
                    ws.cell(row=r, column=c).border = border
                    ws.cell(row=r, column=c).alignment = Alignment(horizontal="center")
                r += 1
        else:
            ws.cell(row=r, column=1, value="(Sin datos)")
            r += 1

        r += 1
        # Totales
        ws.cell(row=r, column=1, value="Número total de cajas:").font = h_font
        ws.cell(row=r, column=2, value=total_cajas)
        r += 1
        ws.cell(row=r, column=1, value="Total equivalente en 11 lbs:").font = h_font
        ws.cell(row=r, column=2, value=round(total_eq_11lbs, 2))
        r += 1
        ws.cell(row=r, column=1, value="Total de dinero:").font = h_font
        ws.cell(row=r, column=2, value=round(total_dinero, 2))
        r += 2

        # Detalle
        ws.cell(row=r, column=1, value="Detalle de embarques").font = h_font
        r += 1
        headers_det = ["Fecha", "N# Embarque", "N# Factura", "Presentación", "Tamaño", "Cantidad", "Importe"]
        for c, txt in enumerate(headers_det, start=1):
            cell = ws.cell(row=r, column=c, value=txt)
            cell.font = th_font
            cell.fill = th_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
        r += 1

        for sh in embarques_qs.order_by('date', 'tracking_number'):
            for it in sh.items.all():
                importe = it.quantity * float(it.presentation.price)
                ws.cell(row=r, column=1, value=sh.date.strftime("%d/%m/%Y"))
                ws.cell(row=r, column=2, value=sh.tracking_number)
                ws.cell(row=r, column=3, value=sh.invoice_number)
                ws.cell(row=r, column=4, value=it.presentation.name)
                ws.cell(row=r, column=5, value=it.size)
                ws.cell(row=r, column=6, value=it.quantity)
                ws.cell(row=r, column=7, value=round(importe, 2))
                for c in range(1, 7 + 1):
                    ws.cell(row=r, column=c).border = border
                    ws.cell(row=r, column=c).alignment = Alignment(horizontal="center")
                r += 1

        # Anchos
        ws.column_dimensions['A'].width = 14
        ws.column_dimensions['B'].width = 16
        ws.column_dimensions['C'].width = 16
        ws.column_dimensions['D'].width = 26
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 20.29
        ws.column_dimensions['G'].width = 14
        ws.column_dimensions['L'].width = 20.29

        return wb

    # Solo aceptamos descargas si tiene permiso
    descargar = request.GET.get('descargar') if can_download else None

    from datetime import date, timedelta
    from django.http import HttpResponse
    from django.utils.text import slugify
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    # ...
    if request.GET.get('descargar') == 'semana':
        from collections import defaultdict
        from io import BytesIO
        from datetime import date, timedelta
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from decimal import Decimal

        iso_week = request.GET.get('iso_week')  # ej: "2025-W35"
        empresa  = (request.GET.get('empresa') or 'general').strip()

        if not iso_week:
            return HttpResponse("Falta la semana (iso_week).", status=400)

        # Parsear ISO week -> lunes..domingo
        try:
            year_str, week_str = iso_week.split('-W')
            year = int(year_str); week = int(week_str)
            monday = date.fromisocalendar(year, week, 1)  # 1 = lunes
            sunday = monday + timedelta(days=6)
        except Exception:
            return HttpResponse("Semana inválida.", status=400)

        # Query base de la semana
        weeks_qs = (
            Shipment.objects
            .filter(date__range=(monday, sunday))
            .order_by('date', 'id')
            .prefetch_related('items', 'items__presentation')
        )

        # ===== Estilos comunes =====
        th_font = Font(name='Calibri', size=12, bold=True, color="FFFFFF")
        th_fill = PatternFill("solid", fgColor="225577")
        thin    = Border(
            left=Side(style='thin', color='AAAAAA'),
            right=Side(style='thin', color='AAAAAA'),
            top=Side(style='thin', color='AAAAAA'),
            bottom=Side(style='thin', color='AAAAAA'),
        )

        # ------------------------------------------------------------------
        # 1) DISEÑO EMPRESA ESPECÍFICA (con Decimal y regla AGRÍCOLA por embarque)
        # ------------------------------------------------------------------
        if empresa.lower() != 'general':
            from decimal import Decimal, ROUND_HALF_UP
            from openpyxl.drawing.image import Image as XLImage
            from openpyxl.utils import get_column_letter
            import os

            def _q2(x: Decimal) -> Decimal:
                return x.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

            def _round_half_up_to_int(x: Decimal) -> int:
                # x viene en Eq.11 del embarque (no del total)
                # Regla: .49 ↓, .50/.51 ↑ (half-up estándar)
                return int(x.to_integral_value(rounding=ROUND_HALF_UP))

            wb = Workbook()
            ws = wb.active
            ws.title = "Semana"

            emp_label = _canon_company_label(empresa)

            # --- Logo en A1 (escalado) ---
            LOGO_SLUG_MAP = {
                "La Cima Produce": "la-cima-produce",
                "RC Organics": "gh-farms",              # comparte logo con Empaque N.1/GH
                "GH Farms": "gh-farms",
                "Gourmet Baja Farms": "gourmet-baja-farms",
                "GBF Farms": "gbf-farms",
                "Agricola DH & G": "AGRICOLA",
            }
            logo_slug = LOGO_SLUG_MAP.get(emp_label)
            if logo_slug:
                logo_path = os.path.join(settings.BASE_DIR, "static", "logos", f"{logo_slug}.png")
                if os.path.exists(logo_path):
                    try:
                        img = XLImage(logo_path)
                        target_h = 120  # px aprox
                        scale = target_h / float(img.height)
                        img.width  = int(img.width * scale)
                        img.height = int(img.height * scale)
                        ws.add_image(img, "A1")
                    except Exception:
                        pass

            # --- Título en D1 (merge) + subtítulo D3 ---
            ws.merge_cells(start_row=1, start_column=4, end_row=2, end_column=9)  # D1:I2
            tcell = ws.cell(
                row=1, column=4,
                value=f"Resumen semanal {monday.strftime('%d/%m/%Y')} – {sunday.strftime('%d/%m/%Y')} – {emp_label}"
            )
            tcell.font = Font(name='Calibri', size=18, bold=True, color="3C78D8")
            tcell.alignment = Alignment(horizontal="left", vertical="center")
            scell = ws.cell(row=3, column=4, value="Detalle de embarques y totales por semana")
            scell.font = Font(name='Calibri', size=11, italic=True, color="6D6D6D")
            scell.alignment = Alignment(horizontal="left")

            # --- Encabezados (fila 7) ---
            headers = ["N° EMBARQUE", "N° FACTURA", "FECHA", "PRESENTACIÓN",
                    "TAMAÑO", "CANTIDAD", "EQUIV. 11 LBS", "IMPORTE ($)", "CLIENTE"]
            r = 7
            for c, h in enumerate(headers, start=1):
                cell = ws.cell(row=r, column=c, value=h)
                cell.font = th_font
                cell.fill = th_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin
            ws.row_dimensions[r].height = 22
            ws.auto_filter.ref = f"A{r}:I{r}"
            r += 1

            # --- Anchos + freeze panes ---
            for col_idx in range(1, 9 + 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = 24
            ws.freeze_panes = "A8"

            # --- Helper: número de orden mostrado por cliente ---
            def _client_order_for(embarque, cliente_name):
                cname = (cliente_name or "").lower()
                if "cima" in cname:
                    return getattr(embarque, "order_lacima", None)
                if "rc" in cname:
                    return getattr(embarque, "order_rc", None)
                if "gh" in cname:
                    return getattr(embarque, "order_gh", None)
                if "gourmet" in cname:
                    return getattr(embarque, "order_gourmet", None)
                if "gbf" in cname:
                    return getattr(embarque, "order_gbf", None)
                if "agricola dh & g" in cname or "agricola" in cname or "dhg" in cname:
                    return getattr(embarque, "order_dhg", None)
                return None

            # --- Filas (una por ítem; importes con Decimal) ---
            total_boxes = 0
            total_eq    = Decimal('0')
            total_amt   = Decimal('0')

            is_agricola = _canon_company_label(emp_label).upper() in SPECIAL_EQ11_ROUND_CLIENTS

            for comp, s, it in _iter_company_items(weeks_qs, emp_label):
                qty = int(it.quantity or 0)
                cf  = Decimal(str(getattr(it.presentation, "conversion_factor", 1.0)))
                eq_emb = Decimal(qty) * cf

                # Importe “por embarque”:
                if is_agricola:
                    # AGRÍCOLA: 3.40 * round_half_up(eq11_embarque), y se redondea a 2 decimales
                    amt_emb = _q2(Decimal('3.40') * Decimal(_round_half_up_to_int(eq_emb)))
                else:
                    prc = Decimal(str(getattr(it.presentation, "price", 0.0)))
                    amt_emb = _q2(Decimal(qty) * prc)

                num_emb = _client_order_for(s, emp_label) or str(s.tracking_number)

                vals = [
                    num_emb,
                    s.invoice_number,
                    s.date.strftime('%d/%m/%Y'),
                    str(it.presentation.name).strip(),
                    str(it.size).strip(),
                    qty,
                    float(_q2(eq_emb)),       # mostrar a 2 dec
                    float(amt_emb),           # mostrar a 2 dec
                    emp_label,
                ]
                for c, v in enumerate(vals, start=1):
                    cc = ws.cell(row=r, column=c, value=v)
                    cc.border = thin
                    if c in (6, 7, 8):
                        cc.alignment = Alignment(horizontal="right")
                    else:
                        cc.alignment = Alignment(horizontal="center")
                r += 1

                total_boxes += qty
                total_eq    += eq_emb
                total_amt   += amt_emb

            # --- Fila de TOTALES (banda azul claro) ---
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
            lbl = ws.cell(row=r, column=1, value="TOTALES:")
            lbl.font = Font(name='Calibri', size=12, bold=True, color="225577")
            lbl.alignment = Alignment(horizontal="right", vertical="center")

            c_cajas = ws.cell(row=r, column=6, value=total_boxes)
            c_eq    = ws.cell(row=r, column=7, value=float(_q2(total_eq)))
            c_amt   = ws.cell(row=r, column=8, value=float(_q2(total_amt)))
            for c in (6, 7, 8):
                cc = ws.cell(row=r, column=c)
                cc.font = Font(name='Calibri', size=12, bold=True)
                cc.fill = PatternFill("solid", fgColor="BBDDFF")
                cc.alignment = Alignment(horizontal="right", vertical="center")
                cc.border = thin
            c_cajas.number_format = '#,##0'
            c_eq.number_format    = '#,##0.00'
            c_amt.number_format   = '$#,##0.00'

            # --- Salida ---
            out = BytesIO()
            wb.save(out)
            out.seek(0)
            filename = f"semana_{year}-W{week}_{slugify(emp_label)}.xlsx"
            resp = HttpResponse(out, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            resp["Content-Disposition"] = f'attachment; filename="{filename}"'
            return resp



        # ================================================================
        # 2) DISEÑO GENERAL (misma hoja + matriz al final)
        # ================================================================
        # Mapa {empresa: [(shipment,item), ...]}
        company_pairs = defaultdict(list)
        for comp, s, it in _iter_company_items(weeks_qs, None):
            company_pairs[comp].append((s, it))

        wb = Workbook()

        # 2.1 Hoja por empresa
        ws = wb.active
        ws.title = "Resumen por empresa"
        ws.cell(row=1, column=1, value=f"Resumen semanal {monday.strftime('%d/%m/%Y')} – {sunday.strftime('%d/%m/%Y')} – GENERAL") \
        .font = Font(size=18, bold=True, color="3C78D8")
        r = 3

        grand_cajas = grand_eq = grand_amt = 0.0
        for comp in sorted(company_pairs.keys(), key=lambda x: x.upper()):
            pres_info, totals, detalle = _compute_company_summary(comp, company_pairs[comp])
            r = _write_company_section(ws, r, thin, th_font, th_fill, comp, pres_info, totals, detalle)
            grand_cajas += totals["total_cajas"]
            grand_eq    += totals["total_eq11"]
            grand_amt   += totals["total_dinero"]

        # Totales generales de la hoja
        ws.cell(row=r, column=1, value="TOTALES GENERALES").font = Font(size=14, bold=True); r += 1
        for label, val in [("TOTAL CAJAS", int(grand_cajas)),
                        ("TOTAL Eq. 11 lbs", round(grand_eq, 2)),
                        ("TOTAL IMPORTE", round(grand_amt, 2))]:
            ws.cell(row=r, column=1, value=label).border = thin
            ws.cell(row=r, column=2, value=val).border = thin
            r += 1

        # 2.2 Matriz (Resumen) al final de la MISMA hoja
        r += 2
        ws.cell(row=r, column=1, value="Matriz (Resumen)").font = Font(size=13, bold=True, color="225577")
        r += 1

        empresas = sorted(company_pairs.keys(), key=str.upper)
        headers = ["Semana", "Rango"] + empresas + ["TOTAL_EQ11"]
        for c, h in enumerate(headers, start=1):
            cell = ws.cell(row=r, column=c, value=h)
            cell.font = th_font; cell.fill = th_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin

        r += 1

        # Etiquetas
        if weeks_qs:
            week_label = f"{monday.isocalendar().week}"
            rango = f"{monday.strftime('%d/%m/%Y')} – {sunday.strftime('%d/%m/%Y')}"
        else:
            week_label = ""; rango = ""

        # --- Importes por empresa + total eq11 (AGRICOLA: redondeo POR EMBARQUE) ---
        from collections import defaultdict
        from decimal import Decimal, ROUND_HALF_UP

        def _q2(x: Decimal) -> Decimal:
            return x.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

        def _round_half_up_to_int(x: Decimal) -> int:
            # .49 ↓, .50/.51 ↑ (equiv a ROUND_HALF_UP)
            return int(x.to_integral_value(rounding=ROUND_HALF_UP))

        def _is_agricola(label: str) -> bool:
            """
            Detecta AGRICOLA DH & G aun si el canon queda 'AGRICOLA' a secas u otras variantes.
            En tu dataset solo hay una 'Agricola', así que es seguro usar startswith.
            """
            n = _canon_company_label(label).upper()
            # normaliza algunos símbolos por si acaso
            n = n.replace('&', '').replace('  ', ' ').strip()
            # Acepta 'AGRICOLA', 'AGRICOLA DH G', 'AGRICOLA DHG', etc.
            return n.startswith('AGRICOLA')


        Q01  = Decimal('0.01')
        P340 = Decimal('3.40')

        # ⚠️ Normaliza las llaves de 'empresas'
        empresas = sorted({_canon_company_label(e) for e in company_pairs.keys()}, key=str.upper)

        per_company_amt: dict[str, Decimal] = {e: Decimal('0') for e in empresas}
        per_company_eq:  dict[str, Decimal] = {e: Decimal('0') for e in empresas}
        total_eq11 = Decimal('0')

        # Para AGRICOLA acumulamos Eq.11 POR EMBARQUE (clave normalizada)
        ship_eq: dict[str, dict[int, Decimal]] = defaultdict(lambda: defaultdict(lambda: Decimal('0')))

        for comp_raw, s, it in _iter_company_items(weeks_qs, None):
            comp = _canon_company_label(comp_raw)  # ← clave normalizada
            qty = Decimal(str(int(it.quantity or 0)))
            cf  = Decimal(str(getattr(it.presentation, "conversion_factor", 1.0)))
            prc = Decimal(str(getattr(it.presentation, "price", 0.0)))

            eq = qty * cf
            total_eq11 += eq

            # Sumas "lista" por defecto
            per_company_eq[comp]  += eq
            per_company_amt[comp] += (qty * prc)

            # Guarda Eq.11 por embarque para clientes especiales
            ship_eq[comp][s.id]   += eq

        # Ajuste especial: AGRICOLA DH & G => sum( round_half_up(Eq.11_emb) * 3.40 )
        for comp in empresas:
            if _is_agricola(comp):
                total_importe = Decimal('0')
                for _, eq_emb in ship_eq.get(comp, {}).items():
                    bill_units = _round_half_up_to_int(eq_emb)
                    total_importe += (P340 * Decimal(bill_units))
                per_company_amt[comp] = _q2(total_importe)

        # (coherencia visual del total Eq11)
        total_eq11 = _q2(total_eq11)

        # === Escribir fila de la matriz ===
        ws.cell(row=r, column=1, value=week_label).border = thin
        ws.cell(row=r, column=2, value=rango).border = thin
        cidx = 3
        for emp in empresas:
            # Per-company amt ya trae AGRICOLA ajustado por embarque
            val = float(_q2(per_company_amt.get(emp, Decimal('0'))))
            ws.cell(row=r, column=cidx, value=val).border = thin
            cidx += 1
        ws.cell(row=r, column=cidx, value=float(total_eq11)).border = thin

        # Anchos cómodos
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 28
        for i in range(3, 3 + len(empresas) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 18

        # ======= COBRO DE MAQUILA (debajo de la Matriz) =======
        r += 2
        ws.cell(row=r, column=1, value="Cobro de maquila").font = Font(size=13, bold=True, color="225577")
        r += 1

        # Excluir La Cima y RC (por etiqueta “bonita”)
        EXCLUDE = {"LA CIMA PRODUCE", "RC ORGANICS"}
        empresas_mq = [e for e in empresas if _canon_company_label(e).upper() not in EXCLUDE]

        # Encabezados: Semana | Rango | <empresas sin LC/RC> | TOTAL_IMPORTE
        headers_mq = ["Semana", "Rango"] + empresas_mq + ["TOTAL_IMPORTE"]
        for c, h in enumerate(headers_mq, start=1):
            cell = ws.cell(row=r, column=c, value=h)
            cell.font = th_font
            cell.fill = th_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin
        r += 1

        # Fila de datos (reutiliza week_label, rango, per_company_amt calculados arriba)
        cidx = 1
        c = ws.cell(row=r, column=cidx, value=week_label); c.border = thin; c.alignment = Alignment(horizontal="center"); cidx += 1
        c = ws.cell(row=r, column=cidx, value=rango);      c.border = thin; c.alignment = Alignment(horizontal="center"); cidx += 1

        total_importe_mq = 0.0
        for emp in empresas_mq:
            amt = round(float(per_company_amt.get(emp, 0.0)), 2)
            cell = ws.cell(row=r, column=cidx, value=amt)
            cell.border = thin
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.number_format = '$#,##0.00'
            total_importe_mq += amt
            cidx += 1

        # Total importe (suma de importes de las empresas incluidas)
        cell = ws.cell(row=r, column=cidx, value=round(total_importe_mq, 2))
        cell.border = thin
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.number_format = '$#,##0.00'

        # Ajuste de anchos (opcional)
        ws.column_dimensions['A'].width = max(ws.column_dimensions.get('A').width or 0, 12)
        ws.column_dimensions['B'].width = max(ws.column_dimensions.get('B').width or 0, 28)
        from openpyxl.utils import get_column_letter
        for i in range(3, 3 + len(empresas_mq) + 1):
            col = get_column_letter(i)
            ws.column_dimensions[col].width = max(ws.column_dimensions.get(col).width or 0, 18)

        # ======= COBRO LC + RC por Eq. 11 lbs (x $6.00) =======
        r += 2
        ws.cell(row=r, column=1, value="Cobro LC + RC (Eq. 11 lbs × $6.00)").font = Font(size=13, bold=True, color="225577")
        r += 1

        # Helpers para encontrar el nombre "bonito" tal como aparece en 'empresas'
        def _find_display_name(label_list, canon_target):
            canon_target = canon_target.upper()
            for e in label_list:
                if _canon_company_label(e).upper() == canon_target:
                    return e
            # Si no lo encontró, regresa el canon como fallback
            return canon_target.title()

        LC_LABEL = _find_display_name(empresas, "LA CIMA PRODUCE")
        RC_LABEL = _find_display_name(empresas, "RC ORGANICS")

        headers_lcrc = ["Semana", "Rango", LC_LABEL, RC_LABEL, "TOTAL Eq. 11 lbs", "IMPORTE ($)"]
        for c, h in enumerate(headers_lcrc, start=1):
            cell = ws.cell(row=r, column=c, value=h)
            cell.font = th_font
            cell.fill = th_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin
        r += 1

        # Eq. 11 lbs por empresa (columna por empresa)
        eq_lc = float(per_company_eq.get(LC_LABEL, 0.0))
        eq_rc = float(per_company_eq.get(RC_LABEL, 0.0))
        total_eq_lcrc = eq_lc + eq_rc
        importe_lcrc = round(total_eq_lcrc * 6.0, 2)

        # Fila de datos
        cidx = 1
        c = ws.cell(row=r, column=cidx, value=week_label); c.border = thin; c.alignment = Alignment(horizontal="center"); cidx += 1
        c = ws.cell(row=r, column=cidx, value=rango);      c.border = thin; c.alignment = Alignment(horizontal="center"); cidx += 1

        c = ws.cell(row=r, column=cidx, value=round(eq_lc, 2)); c.border = thin; c.alignment = Alignment(horizontal="right"); c.number_format = '#,##0.00'; cidx += 1
        c = ws.cell(row=r, column=cidx, value=round(eq_rc, 2)); c.border = thin; c.alignment = Alignment(horizontal="right"); c.number_format = '#,##0.00'; cidx += 1

        c = ws.cell(row=r, column=cidx, value=round(total_eq_lcrc, 2)); c.border = thin; c.alignment = Alignment(horizontal="right"); c.number_format = '#,##0.00'; cidx += 1
        c = ws.cell(row=r, column=cidx, value=importe_lcrc);             c.border = thin; c.alignment = Alignment(horizontal="right"); c.number_format = '$#,##0.00'

        # Anchos cómodos (opcional)
        from openpyxl.utils import get_column_letter
        ws.column_dimensions['A'].width = max(ws.column_dimensions.get('A').width or 0, 12)
        ws.column_dimensions['B'].width = max(ws.column_dimensions.get('B').width or 0, 28)
        for i in range(3, 7):  # columnas C..F
            col = get_column_letter(i)
            ws.column_dimensions[col].width = max(ws.column_dimensions.get(col).width or 0, 18)



        # Salida
        out = BytesIO(); wb.save(out); out.seek(0)
        filename = f"semana_{year}-W{week}_general.xlsx"
        resp = HttpResponse(out, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        resp["Content-Disposition"] = f'attachment; filename="{filename}"'
        return resp




    # --------------------------
    # Descarga por RANGO (XLSX)
    # --------------------------
    if descargar == 'rango':
        if not request.user.has_perm('empaques.can_download_reports'):
            return HttpResponse("No tienes permiso para descargar reportes.", status=403)

        start_str = request.GET.get('start')
        end_str   = request.GET.get('end')
        if not start_str or not end_str:
            return HttpResponse("Debes indicar 'start' y 'end' en formato YYYY-MM-DD.", status=400)

        try:
            start_d = date.fromisoformat(start_str)
            end_d   = date.fromisoformat(end_str)
        except ValueError:
            return HttpResponse("Fechas inválidas. Usa formato YYYY-MM-DD.", status=400)

        if end_d < start_d:
            return HttpResponse("El fin de rango no puede ser menor que el inicio.", status=400)
        else:

            embarques = Shipment.objects.filter(date__range=(start_d, end_d))

        wb = build_summary_xlsx(
            "Resumen de Embarques – Rango de fechas",
            f"Rango: {start_d.strftime('%d/%m/%Y')} – {end_d.strftime('%d/%m/%Y')}",
            embarques
        )

        output = BytesIO(); wb.save(output); output.seek(0)
        filename = f"resumen_rango_{start_d.strftime('%Y%m%d')}_{end_d.strftime('%Y%m%d')}.xlsx"
        resp = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        resp['Content-Disposition'] = f'attachment; filename="{filename}"'
        return resp

    # ================================
    # Descarga mensual o anual (XLSX) con formato de "semana"
    # ================================
    from django.utils.text import slugify
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    if descargar in ('mes', 'ano'):
        from collections import defaultdict
        empresa = _empresa_param(request)  # None => general (todas)


        # --- Helper para número de orden por cliente (igual que en semana) ---
        def get_client_order_number_for(embarque, cliente_name):
            if not embarque or not cliente_name:
                return None
            cname = str(cliente_name).lower()
            if "cima" in cname:
                return getattr(embarque, "order_lacima", None)
            if "rc" in cname:
                return getattr(embarque, "order_rc", None)
            if "gh" in cname:
                return getattr(embarque, "order_gh", None)
            if "gourmet" in cname:
                return getattr(embarque, "order_gourmet", None)
            if "gbf" in cname:
                return getattr(embarque, "order_gbf", None)
            if "agricola dh & g" in cname or "agricola" in cname or "dhg" in cname:
                return getattr(embarque, "order_dhg", None)
            return None

        # --- Query base ---
        if descargar == 'mes':
            embarques = (
                Shipment.objects
                .filter(date__year=year, date__month=month)
                .order_by('date', 'id')
                .prefetch_related('items', 'items__presentation')
            )
            base_filename = f"resumen_mes_{year}_{month:02d}"
            titulo = f"Resumen mensual {year}-{month:02d}"
        else:
            embarques = (
                Shipment.objects
                .filter(date__year=year)
                .order_by('date', 'id')
                .prefetch_related('items', 'items__presentation')
            )
            base_filename = f"resumen_anual_{year}"
            titulo = f"Resumen anual {year}"

        # Si viene empresa específica, ajústalo en título/archivo
        emp_label = (empresa or "general")
        if empresa:
            titulo = f"{titulo} – {empresa}"
            base_filename = f"{base_filename}_{slugify(empresa)}"

        # ===================== Construir Excel (MES/AÑO) – MULTIEMPRESA =====================
        empresa_filter = empresa  # ya viene de _empresa_param(request) (None => general)

        # Arma mapa {empresa: [(s,it), ...]}
        company_map = defaultdict(list)
        for comp, s, it in _iter_company_items(embarques, empresa_filter):
            label = _canon_company_label(comp)
            company_map[label].append((s, it))

        # Si es empresa específica, fuerza a una sola sección
        if empresa_filter:
            only = _canon_company_label(empresa_filter)
            company_map = { only: company_map.get(only, []) }

        wb = Workbook()
        ws = wb.active
        ws.title = "Resumen por empresa"

        title_font   = Font(name='Calibri', size=18, bold=True, color="3C78D8")
        th_font      = Font(name='Calibri', size=12, bold=True, color="FFFFFF")
        th_fill      = PatternFill("solid", fgColor="225577")
        thin_border  = Border(
            left=Side(style='thin', color='AAAAAA'),
            right=Side(style='thin', color='AAAAAA'),
            top=Side(style='thin', color='AAAAAA'),
            bottom=Side(style='thin', color='AAAAAA'),
        )

        r = 1
        ws.cell(row=r, column=1, value=titulo if empresa_filter else f"{titulo} – General por empresa").font = title_font
        r += 2

        grand_cajas = 0
        grand_eq11  = 0.0
        grand_dinero = 0.0

        for comp in sorted(company_map.keys(), key=lambda x: x.upper()):
            tuples_cs = company_map[comp]
            pres_info, totals, detalle = _compute_company_summary(comp, tuples_cs)

            r = _write_company_section(ws, r, thin_border, th_font, th_fill, comp,
                                    pres_info, totals, detalle)

            grand_cajas  += totals['total_cajas']
            grand_eq11   += float(totals['total_eq11'] or 0.0)
            grand_dinero += float(totals['total_dinero'] or 0.0)

        # ===== TOTALES GENERALES =====
        ws.cell(row=r, column=1, value="TOTALES GENERALES").font = Font(name="Calibri", size=14, bold=True)
        r += 1
        headers_tot = ["Métrica", "Total"]
        for c, txt in enumerate(headers_tot, start=1):
            cell = ws.cell(row=r, column=c, value=txt)
            cell.font = th_font; cell.fill = th_fill
            cell.alignment = Alignment(horizontal="center", vertical="center"); cell.border = thin_border
        r += 1
        pairs = [
            ("TOTAL CAJAS", int(grand_cajas)),
            ("TOTAL Eq. 11 lbs", round(grand_eq11, 2)),
            ("TOTAL IMPORTE", round(grand_dinero, 2)),
        ]
        for label, val in pairs:
            ws.cell(row=r, column=1, value=label).border = thin_border
            ws.cell(row=r, column=2, value=val).border = thin_border
            ws.cell(row=r, column=1).alignment = Alignment(horizontal="right")
            ws.cell(row=r, column=2).alignment = Alignment(horizontal="center")
            r += 1

        # Anchos
        ws.column_dimensions['A'].width = 32
        ws.column_dimensions['B'].width = 16
        ws.column_dimensions['C'].width = 16
        ws.column_dimensions['D'].width = 26
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 14
        ws.column_dimensions['G'].width = 16

        thin = thin_border

        # ======= MATRIZ (RESUMEN) — al final de la hoja =======
        from openpyxl.utils import get_column_letter
        from decimal import Decimal

        r += 2
        ws.cell(row=r, column=1, value="Matriz (Resumen)").font = Font(size=13, bold=True, color="225577")
        r += 1

        # Empresas detectadas en el periodo
        empresas = sorted(company_map.keys(), key=str.upper)

        # Headers
        headers = ["Periodo", "Rango"] + empresas + ["TOTAL_EQ11"]
        for c, h in enumerate(headers, start=1):
            cell = ws.cell(row=r, column=c, value=h)
            cell.font = th_font; cell.fill = th_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin
        r += 1

        # Etiquetas de periodo y rango
        if descargar == 'mes':
            periodo_label = f"{year}-{month:02d}"
            # Rango mes: primer día..último día del mes
            from calendar import monthrange
            d1 = date(year, month, 1)
            d2 = date(year, month, monthrange(year, month)[1])
            rango_label = f"{d1.strftime('%d/%m/%Y')} – {d2.strftime('%d/%m/%Y')}"
        else:
            periodo_label = f"{year}"
            d1 = date(year, 1, 1)
            d2 = date(year, 12, 31)
            rango_label = f"{d1.strftime('%d/%m/%Y')} – {d2.strftime('%d/%m/%Y')}"

        # Importes por empresa + total eq11 del periodo
        per_company_amt = {e: 0.0 for e in empresas}   # 'empresas' ya viene normalizado desde company_map.keys()
        per_company_eq  = {e: 0.0 for e in empresas}
        total_eq11 = 0.0

        # Recorremos TODO el queryset (embarques) usando SIEMPRE etiqueta normalizada
        for comp, s, it in _iter_company_items(embarques, None):
            label = _canon_company_label(comp)  # <-- NORMALIZA LA CLAVE AQUÍ
            qty = int(it.quantity or 0)
            cf  = float(getattr(it.presentation, "conversion_factor", 1.0))
            prc = float(getattr(it.presentation, "price", 0.0))

            total_eq11 += qty * cf
            # Usa .get(...) para blindarte si aparece una empresa no listada en 'empresas'
            per_company_eq[label]  = per_company_eq.get(label, 0.0) + qty * cf
            per_company_amt[label] = per_company_amt.get(label, 0.0) + qty * prc

        # Ajuste especial por empresa (AGRICOLA DH & G)
        for emp in empresas:  # 'empresas' ya son etiquetas normalizadas
            if _canon_company_label(emp).upper() in SPECIAL_EQ11_ROUND_CLIENTS:
                per_company_amt[emp] = float(
                    Decimal('3.40') * Decimal(_round_half_up_to_int(per_company_eq.get(emp, 0.0)))
                )

        # Escribir fila de la matriz
        ws.cell(row=r, column=1, value=periodo_label).border = thin
        ws.cell(row=r, column=2, value=rango_label).border  = thin
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=r, column=2).alignment = Alignment(horizontal="center", vertical="center")

        cidx = 3
        for emp in empresas:
            val = round(per_company_amt.get(emp, 0.0), 2)
            cell = ws.cell(row=r, column=cidx, value=val)
            cell.border = thin
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.number_format = '$#,##0.00'
            cidx += 1

        cell = ws.cell(row=r, column=cidx, value=round(total_eq11, 2))
        cell.border = thin
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.number_format = '#,##0.00'

        # Anchos cómodos
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 28
        for i in range(3, 3 + len(empresas) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 18



        # Salida
        output = BytesIO(); wb.save(output); output.seek(0)
        filename = f"{base_filename}.xlsx"
        resp = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        resp['Content-Disposition'] = f'attachment; filename="{filename}"'
        return resp




    # ================================
    # Render normal (sin descargas)
    # ================================
    return render(request, 'empaques/shipment_list.html', {
        'shipments': shipments,
        'today': date.today(),
        'can_download': can_download,
    })




from datetime import date
from django.utils.text import slugify
from io import BytesIO
import os
from django.conf import settings 
from django.http import HttpResponse
@login_required

def daily_report(request, shipment_id=None):
    ship_id  = request.GET.get('shipment_id') or shipment_id
    tracking = request.GET.get('tracking')
    if not (request.user.has_perm('empaques.view_shipment') or request.user.has_perm('empaques.export_reports')):
        return HttpResponseForbidden("No tienes permiso para ver reportes.")  # <-- usa HttpResponseForbidden

    fmt = request.GET.get('format')
    if fmt and not request.user.has_perm('empaques.export_reports'):
        return HttpResponseForbidden("No tienes permiso para exportar reportes.") 
    from io import BytesIO
    import os
    from datetime import date
    from django.conf import settings
    from django.utils.text import slugify
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    ORDER_FIELD_BY_SLUG = {
        # nombres cortos
        'la-cima-produce':                 'order_lacima',
        'rc-organics':                     'order_rc',
        'gh-farms':                     'order_gh',
        'gourmet-baja-farms':              'order_gourmet',
        'gbf-farms':                       'order_gbf',

        # variantes largas (razón social completa)
        'la-cima-produce-s-p-r-de-r-l':                'order_lacima',
        'rc-organics-s-de-r-l-de-c-v':                 'order_rc',
        'empaque-n-1-s-de-r-l-de-c-v':                 'order_gh',
        'gourmet-baja-farms-s-de-r-l-de-c-v':          'order_gourmet',
        'gbf-farms-s-de-r-l-de-c-v':                   'order_gbf',
    }

    def get_client_order_number(embarque, cliente):
        """Devuelve el número de orden específico del cliente si existe."""
        cname = (cliente or "").lower()
        if "cima" in cname:
            return getattr(embarque, "order_lacima", None)
        if "rc" in cname:
            return getattr(embarque, "order_rc", None)
        if "gourmet" in cname:
            return getattr(embarque, "order_gourmet", None)
        if "gbf" in cname:
            return getattr(embarque, "order_gbf", None)
        if "gh" in cname:
            return getattr(embarque, "order_gh", None)
        return None
    def fecha_es(d):
        """Devuelve la fecha en español: LUNES 31 DE AGOSTO DEL 2025 (en mayúsculas)."""
        if not d:
            return ""
    # Opción A: Babel (recomendado)
        try:
            from babel.dates import format_date
            txt = format_date(d, format="EEEE d 'DE' MMMM 'DEL' y", locale='es_MX')
            return txt.upper()
        except Exception:
            # Opción B: fallback manual si Babel no está disponible
            dias = ["LUNES","MARTES","MIÉRCOLES","JUEVES","VIERNES","SÁBADO","DOMINGO"]
            meses = ["ENERO","FEBRERO","MARZO","ABRIL","MAYO","JUNIO","JULIO","AGOSTO","SEPTIEMBRE","OCTUBRE","NOVIEMBRE","DICIEMBRE"]
            return f"{dias[d.weekday()]} {d.day:02d} DE {meses[d.month-1]} DEL {d.year}"
    

 


    # ---- Fecha a reportar ----
    qdate = request.GET.get('date')
    try:
        report_date = date.fromisoformat(qdate) if qdate else date.today()
    except ValueError:
        report_date = date.today()

    if ship_id:
        try:
            sid = int(ship_id)
            qs = (
                Shipment.objects
                .filter(id=sid)
                .prefetch_related('items', 'items__presentation')
            )
        except ValueError:
            qs = Shipment.objects.none()
    elif tracking:
        qs = (
            Shipment.objects
            .filter(tracking_number=tracking)
            .prefetch_related('items', 'items__presentation')
        )
    else:
        qs = (
            Shipment.objects
            .filter(date=report_date)
            .order_by('-id')
            .prefetch_related('items', 'items__presentation')
        )

    # Si vino selección específica, ajusta report_date para mostrarlo bonito
    if (ship_id or tracking) and qs.exists():
        report_date = qs.first().date
     # Si me pasaron shipment_id o tracking, fuerzo a 1 solo embarque
    if shipment_id or tracking:
        base_qs = Shipment.objects.all().prefetch_related('items', 'items__presentation')
        if shipment_id:
            qs = base_qs.filter(id=shipment_id)
        else:
            qs = base_qs.filter(tracking_number=tracking).order_by('-id')[:1]
        # ajusta la fecha para que los títulos/encabezados muestren el día real del embarque
        if qs:
            report_date = qs[0].date
    

    # Totales del día (para el general) - precomputar total_boxes = sum(item.quantity)
    total_boxes = sum(item.quantity for s in qs for item in s.items.all()) 
    total_eq_11lbs = sum(
        item.quantity * float(item.presentation.conversion_factor)
        for s in qs for item in s.items.all()
    )
    total_amount = sum(
        item.quantity * float(item.presentation.price)
        for s in qs for item in s.items.all()
    )

    # ---------------- Helpers ----------------
    def _str(v):
        if v is None:
            return ""
        if isinstance(v, bool):
            return "Sí" if v else "No" 
        return str(v)

    def write_shipment_info(ws, start_row, start_col, embarque, include_peco=False):
        """Escribe bloque con datos del embarque. Devuelve el último row usado."""
        label_font = Font(name='Calibri', size=18, bold=True, color="666666")
        value_font = Font(name='Calibri', size=16)
        seals = ", ".join([s for s in [embarque.seal_1, embarque.seal_2, embarque.seal_3, embarque.seal_4] if s])
        info = [
            ("Núm. Orden",     _str(embarque.tracking_number)),
            ("FECHA", fecha_es(getattr(embarque, "date", None))),
            ("Transportista",  _str(embarque.carrier)),
            ("Placas Tractor", _str(embarque.tractor_plates)),
            ("Placas Caja",    _str(embarque.box_plates)),
            ("Operador",       _str(embarque.driver)),
            ("Hora Salida",    _str(embarque.departure_time)),
            ("Caja",           _str(embarque.box)),
            ("Cond. de Caja",  _str(embarque.box_conditions)),
            ("Caja sin Olores",_str(embarque.box_free_of_odors)),
            ("Ryan",           _str(embarque.ryan)),
            ("Sellos",         seals),
            ("Chismógrafo",    _str(embarque.chismografo)),
            ("Firma Entrega",  _str(embarque.delivery_signature)),
            ("Firma Operador", _str(embarque.driver_signature)),
            ("Factura",        _str(embarque.invoice_number)),
            ("Dirección",      "H. GALEANA N. 85 LOC A-B, C. ZARAGOZA"),
            ("Y R. ZAPATA COL. CENTRO, 23600", ""),
        ]
         # Insertar Tarimas PECO SOLO si include_peco=True 
        if include_peco:
            info.append(("Tarimas PECO", _str(getattr(embarque, "tarimas_peco", None))))


         # Escribir en hoja
        r = start_row
        for label, value in info:
            ws.cell(row=r, column=start_col,     value=label + ":").font = label_font
            ws.cell(row=r, column=start_col + 1, value=value).font     = value_font
            r += 1

        return r - 1

    def score_shipment(s):
        """Para el general: prioriza más ítems, más campos llenos, más reciente."""
        items_count = len(s.items.all())
        header_values = [
            s.box_conditions, s.box_free_of_odors, s.ryan,
            s.seal_1, s.seal_2, s.seal_3, s.seal_4,
            s.chismografo, s.delivery_signature, s.driver_signature, s.invoice_number
        ]
        filled_count = sum(1 for v in header_values if v not in (None, "", False))
        return (items_count, filled_count, s.id)

    def tarima_temp_text(items_list, tarima):
        """Busca la primera temperatura no vacía en esa tarima y devuelve texto pretty."""
        for it in items_list:
            if it.tarima == tarima and it.temperatura not in (None, ""):
                try:
                    return f"{float(it.temperatura):.1f} °F"
                except Exception:
                    return str(it.temperatura)
        return ""
    

    def pintar_bloque_tarima(ws_, top_row, left_col, temp_col, items_, temp_text):
        
        from openpyxl.styles import Alignment, Border, Side, Font

        thin  = Side(style='thin',   color='999999')
        thick = Side(style='medium', color='000000')
        thick_all = Border(top=thick, bottom=thick, left=thick, right=thick)

        # Marco 2x4 (bordes internos finos, exteriores gruesos)
        for rr in (top_row, top_row + 1):
            for cc in range(left_col, left_col + 4):
                cell = ws_.cell(row=rr, column=cc, value="")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                top_side    = thick if rr == top_row else thin
                bottom_side = thick if rr == (top_row + 1) else thin
                left_side   = thick if cc == left_col else thin
                right_side  = thick if cc == (left_col + 3) else thin
                cell.border = Border(top=top_side, bottom=bottom_side, left=left_side, right=right_side)

        # Celda única de temperatura (fusionada verticalmente)
        ws_.merge_cells(start_row=top_row, start_column=temp_col, end_row=top_row + 1, end_column=temp_col)
        for rr in (top_row, top_row + 1):
            c = ws_.cell(row=rr, column=temp_col, value=(temp_text or "") if rr == top_row else None)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = thick_all

        # Hasta 4 ítems (slots: (fila, col_label, col_qty))
        slots = [
            (top_row,     left_col,     left_col + 1),
            (top_row,     left_col + 2, left_col + 3),
            (top_row + 1, left_col,     left_col + 1),
            (top_row + 1, left_col + 2, left_col + 3),
        ]

        # Un poco más alto para que quepan 2 líneas
        ws_.row_dimensions[top_row].height     = 32
        ws_.row_dimensions[top_row + 1].height = 32

        for it, (r, c_label, c_qty) in zip(items_[:4], slots):
            # Texto con salto de línea: Tipo + Tamaño
            label_text = f"{_str(it.presentation.name)}\n{_str(it.size)}"
            lbl = ws_.cell(row=r, column=c_label, value=label_text)
            lbl.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            qty = ws_.cell(row=r, column=c_qty, value=it.quantity)
            qty.alignment = Alignment(horizontal="center", vertical="center")
            qty.font = Font(bold=True)

    def write_datos(ws_, start_row, embarque, order_override=None):
        from openpyxl.styles import Font, Alignment
        lf = Font(name='Calibri', size=12, bold=True, color="000000")
        vf = Font(name='Calibri', size=13)

        r = start_row
        if not embarque:
            return r - 1

        # 1) elegir el número de orden a mostrar (cliente > general)
        #    usamos "no vacío" (None o "" cae al general)
        if order_override is not None and str(order_override).strip() != "":
            orden_val = str(order_override).strip()
        else:
            orden_val = _str(embarque.tracking_number)

        # 2) texto de fecha en español (usa tu helper si existe)
        try:
            fecha_txt = fecha_es(getattr(embarque, "date", None))
        except NameError:
            fecha_txt = embarque.date.strftime("%A %d DE %B DEL %Y").upper() if getattr(embarque, "date", None) else ""

        seals = ", ".join([s for s in [embarque.seal_1, embarque.seal_2, embarque.seal_3, embarque.seal_4] if s])

        info = [
            ("NUM. DE ORDEN", orden_val),
            ("FECHA",         fecha_txt),
            ("TRANSPORTISTA", _str(embarque.carrier)),
            ("PLACAS TRACTOR", _str(embarque.tractor_plates)),
            ("PLACAS CAJA",    _str(embarque.box_plates)),
            ("OPERADOR",       _str(embarque.driver)),
            ("HORA DE SALIDA", _str(embarque.departure_time)),
            ("CAJA",           _str(embarque.box)),
            ("CONDICIONES DE LA CAJA", _str(embarque.box_conditions)),
            ("CAJA LIBRE DE OLORES",   _str(embarque.box_free_of_odors)),
            ("RYAN",           _str(embarque.ryan)),
            ("SELLOS",         seals),
            ("CHISMÓGRAFO",    _str(embarque.chismografo)),
            ("FIRMA DEL QUE ENTREGA", _str(embarque.delivery_signature)),
            ("FIRMA DEL OPERADOR",    _str(embarque.driver_signature)),
            ("DEBERÁ MANTENERSE UNA TEMPERATURA CONTINUA DE 35°F", ""),
            ("T. PECO",        _str(getattr(embarque, "tarimas_peco", None))),
            ("DIRECCIÓN",      "H. GALEANA N. 85 LOC A-B, C. ZARAGOZA"),
            ("Y R. ZAPATA COL. CENTRO, 23600", ""),
            ("TELÉFONO",       "01 (613) 132-19-08"),
            ("CEL",            "613 111-71-87, 613 122-01-05"),
        ]

        for label, value in info:
            lbl_up = (label or "").upper()

            # Dirección: valor en B..C (fusionado) y etiqueta en A
            if lbl_up.startswith("DIRECCIÓN"):
                ws_.cell(row=r, column=1, value=f"{label}:").font = lf  # A
                ws_.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
                vcell = ws_.cell(row=r, column=2, value=value)         # B (superior-izq del merge)
                vcell.font = vf
                vcell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                r += 1
                continue

            # 2ª línea de dirección sin etiqueta, solo valor en B..C
            if lbl_up.startswith("Y R. ZAPATA"):
                ws_.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
                vcell = ws_.cell(row=r, column=2, value=value)
                vcell.font = vf
                vcell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                r += 1
                continue

            # Resto normal: etiqueta en A, valor en C
            ws_.cell(row=r, column=1, value=label + ":").font = lf  # A
            ws_.cell(row=r, column=3, value=value).font = vf        # C
            r += 1

        return r - 1


    # =================== EXCEL GENERAL (diseño) ===================
    if request.GET.get('format') == 'xlsx':
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte General"

        header_font = Font(name='Calibri', size=16, bold=True, color="3C78D8")
        table_header_font = Font(name='Calibri', size=14, bold=True, color="FFFFFF")
        th_fill = PatternFill("solid", fgColor="225577")
        border = Border(
            left=Side(style='thin', color='AAAAAA'),
            right=Side(style='thin', color='AAAAAA'),
            top=Side(style='thin', color='AAAAAA'),
            bottom=Side(style='thin', color='AAAAAA'),
        )

        row = 1
        ws.cell(row=row, column=1, value=f"Reporte General – {report_date.strftime('%d/%m/%Y')}")
        ws.cell(row=row, column=1).font = header_font
        row += 2

        # Elegir embarque representativo
        ships = list(qs)
        if ships:
            rep = max(ships, key=score_shipment)
            last = write_shipment_info(ws, start_row=row, start_col=1, embarque=rep)
            row = last + 2
            
        # Encabezados tabla
        headers = [
            "N# EMBARQUE", "N# FACTURA", "PRESENTACIÓN", "TAMAÑO",
            "CANTIDAD", "EQUIV. 11 LBS", "IMPORTE ($)", "CLIENTE"
        ]
        for col, val in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = table_header_font
            cell.fill = th_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
        row += 1
         

        # Filas
        for s in qs:
            for item in s.items.all():
                eq = item.quantity * float(item.presentation.conversion_factor)
                amt = item.quantity * float(item.presentation.price)
                ws.cell(row=row, column=1, value=_str(s.tracking_number))
                ws.cell(row=row, column=2, value=_str(s.invoice_number))
                ws.cell(row=row, column=3, value=_str(item.presentation.name))
                ws.cell(row=row, column=4, value=_str(item.size))
                ws.cell(row=row, column=5, value=item.quantity)
                ws.cell(row=row, column=6, value=round(eq, 2))
                ws.cell(row=row, column=7, value=round(amt, 2))
                ws.cell(row=row, column=8, value=_str(item.cliente))
                for c in range(1, 9):
                    ws.cell(row=row, column=c).alignment = Alignment(horizontal="center")
                    ws.cell(row=row, column=c).border = border
                row += 1
                

        # Totales
        ws.merge_cells(start_row=row+1, start_column=1, end_row=row+1, end_column=4)
        ws.cell(row=row+1, column=1, value="TOTALES:").alignment = Alignment(horizontal="right")
        ws.cell(row=row+1, column=1).font = Font(bold=True, color="225577")
        ws.cell(row=row+1, column=5, value=total_boxes)
        ws.cell(row=row+1, column=6, value=round(total_eq_11lbs, 2))
        ws.cell(row=row+1, column=7, value=round(total_amount, 2))
        for c in range(1, 9):
            ws.cell(row=row+1, column=c).font = Font(bold=True)
            ws.cell(row=row+1, column=c).fill = PatternFill("solid", fgColor="BBDDFF")
            ws.cell(row=row+1, column=c).alignment = Alignment(horizontal="center")

        # Anchos
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['H'].width = 20
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"reporte_{report_date}.xlsx"
        response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
         
    fmt = (request.GET.get('format') or "").strip().lower()
    if fmt.startswith('xlsx_'):
        cliente_slug = fmt[5:]
        cliente = next((c for c in clientes if slugify(c) == cliente_slug), None)
        if not cliente:
            return HttpResponse("Cliente no válido para este reporte.", status=400)

        wb = Workbook()
        ws = wb.active
        ws.title = slugify(cliente)[:31] or "Cliente"

        # --- Anchos A..R (exactos) ---
        widths = {
            'A': 11.43, 'B': 8.57,  'C': 8.57,  'D':10.29, 'E': 7.14,
            'F': 5.43,  'G': 6.14,  'H': 5.57,  'I': 6.86, 'J': 2.86,
            'K': 5.86,  'L': 5.71,  'M': 5.43,  'N': 6.29, 'O': 6.86,
            'P': 3.57,  'Q': 5.00,  'R': 1.71,
        }
        for col, w in widths.items():
            ws.column_dimensions[col].width = w

        # --- Alturas (todas 18.75) ---
        for r in range(1, 80):
            ws.row_dimensions[r].height = 18.75

        # --- Logo en A1 (escalado) ---
        logo_path = os.path.join(settings.BASE_DIR, 'static', 'logos', f'{slugify(cliente)}.png')
        if os.path.exists(logo_path):
            img = XLImage(logo_path)
            target_h = 140
            scale = target_h / img.height
            img.width  = int(img.width * scale)
            img.height = int(img.height * scale)
            ws.add_image(img, "A1")

        # --- Título y subtítulo en I1/I2 ---
        title_font = Font(name="Calibri", size=12, bold=True)
        subtitle_font = Font(name="Calibri", size=11, bold=True)
        display_name = LEGAL_CLIENT_NAME.get(cliente, cliente)

        c1 = ws.cell(row=1, column=9, value=display_name)  # I1
        c1.font = title_font
        c1.alignment = Alignment(horizontal="left", vertical="center")
        c2 = ws.cell(row=2, column=9, value="ESPÁRRAGO MANIFIESTO DE EMBARQUE")  # I2
        c2.font = subtitle_font
        c2.alignment = Alignment(horizontal="left", vertical="center")

        # === Selección del embarque del CLIENTE (solo UNO) ===
        shipments_cliente = [s for s in qs if s.items.filter(cliente=cliente).exists()]

        ship_id  = request.GET.get('shipment_id')
        tracking = request.GET.get('tracking')

        rep_cli = None
        # 1) Prioriza ?shipment_id= (si viene numérico)
        if ship_id:
            try:
                sid = int(ship_id)
            except (TypeError, ValueError):
                sid = None
            if sid is not None:
                rep_cli = next((s for s in shipments_cliente if s.id == sid), None)

        # 2) Si no hubo match por id y viene ?tracking=, intenta por tracking_number
        if rep_cli is None and tracking:
            rep_cli = next((s for s in shipments_cliente if _str(s.tracking_number) == tracking), None)

        # 3) Fallback: si no se especificó nada, usa el primero de ese cliente en la fecha
        if rep_cli is None and shipments_cliente:
            rep_cli = shipments_cliente[0]


        # --- Cabecera "EMBARQUE:" (P2:Q2 etiqueta, R2 valor) con NÚMERO POR CLIENTE ---
        label_font = Font(name="Calibri", size=10, bold=True)
        num_font   = Font(name="Calibri", size=18, bold=True, color="FF0000")
        ws.merge_cells(start_row=2, start_column=16, end_row=2, end_column=17)  # P2:Q2
        ws.cell(row=2, column=16, value="EMBARQUE:").font = label_font

        numero_final = ""
        if rep_cli:
            num_cli = get_client_order_number(rep_cli, cliente)  # puede ser None/""
            numero_final = num_cli or _str(rep_cli.tracking_number)
        ws.cell(row=2, column=18, value=numero_final).font = num_font  # R2

        numero_final = ""
        if rep_cli:
            num_cli = get_client_order_number(rep_cli, cliente)  # puede ser None/""
            numero_final = num_cli or _str(rep_cli.tracking_number)
        ws.cell(row=2, column=18, value=numero_final).font = num_font  # R2

        # --- Datos del embarque (izquierda), SOLO UNA VEZ ---
        datos_start_row = 7
        rptr = datos_start_row
        if rep_cli:
            orden_cliente = get_client_order_number(rep_cli, cliente)
            rptr = write_datos(ws, rptr, rep_cli, order_override=orden_cliente) + 2

        # --- GRID de tarimas ---
        grid_start_row = 5
        number_font = Font(name='Calibri', size=8, bold=True, color="444444")
        thick = Side(style='medium', color='000000')
        thick_all = Border(top=thick, bottom=thick, left=thick, right=thick)

        # Ítems SOLO del embarque rep_cli
        items_cliente = list(rep_cli.items.filter(cliente=cliente)) if rep_cli else []

        def temp_txt(tarima_n):
            for it in items_cliente:
                if it.tarima == tarima_n and it.temperatura not in (None, ""):
                    try:
                        return f"{float(it.temperatura):.1f}°F"
                    except Exception:
                        return str(it.temperatura)
            return ""

        # Ajustes de columnas “recorridos 1 a la izquierda”
        ws.column_dimensions['T'].width = 9.57
        ws.column_dimensions['B'].width = 14.86
        ws.column_dimensions['G'].width = 5.86
        ws.column_dimensions['F'].width = 8.86
        ws.column_dimensions['K'].width = 9.57
        ws.column_dimensions['A'].width = 16
        ws.column_dimensions['Q'].width = 9.57
        ws.column_dimensions['J'].width = 22
        ws.column_dimensions['H'].width = 22
        ws.column_dimensions['M'].width = 22
        ws.column_dimensions['I'].width = 10.29
        ws.column_dimensions['P'].width = 10.29
        ws.column_dimensions['L'].width = 9.57
        ws.column_dimensions['N'].width = 10.29
        ws.column_dimensions['R'].width = 5.86
        ws.column_dimensions['S'].width = 5.71
        ws.column_dimensions['O'].width = 22
        for rr in range(5, 57):
            ws.row_dimensions[rr].height = 32.25

        for i in range(13):
            top = grid_start_row + i * 2
            t_impar = 1 + 2*i
            t_par   = 2 + 2*i

            num_left_col  = 7   # G
            block_left    = 8   # H..K
            temp_left_l   = 12  # L..M
            block_right   = 13  # N..Q
            temp_right_l  = 17  # R..S
            num_right_col = 18  # T

            # número izquierdo
            for rr in (top, top + 1):
                c = ws.cell(row=rr, column=num_left_col)
                c.font = number_font
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = thick_all
            ws.merge_cells(start_row=top, start_column=num_left_col, end_row=top + 1, end_column=num_left_col)
            ws.cell(row=top, column=num_left_col, value=str(t_impar))

            # bloque izquierdo + temp
            items_impar = [it for it in items_cliente if it.tarima == t_impar][:4]
            pintar_bloque_tarima(ws, top, block_left, temp_left_l, items_impar, temp_txt(t_impar))

            # bloque derecho + temp
            items_par   = [it for it in items_cliente if it.tarima == t_par][:4]
            pintar_bloque_tarima(ws, top, block_right, temp_right_l, items_par, temp_txt(t_par))

            # número derecho
            for rr in (top, top + 1):
                c = ws.cell(row=rr, column=num_right_col)
                c.font = number_font
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = thick_all
            ws.merge_cells(start_row=top, start_column=num_right_col, end_row=top + 1, end_column=num_right_col)
            ws.cell(row=top, column=num_right_col, value=str(t_par))

        # ===== Resumen inferior =====
        from collections import defaultdict
        th_font = Font(name='Calibri', size=16, bold=True, color="FFFFFF")
        th_fill = PatternFill("solid", fgColor="225577")
        thin_border = Border(
            left=Side(style='thin', color='AAAAAA'),
            right=Side(style='thin', color='AAAAAA'),
            top=Side(style='thin', color='AAAAAA'),
            bottom=Side(style='thin', color='AAAAAA'),
        )
        body_font = Font(name='Calibri', size=14)

        grid_last_row = grid_start_row + 13*2 - 1
        data_block_last_row = (rptr - 1) if rep_cli else (datos_start_row - 1)
        summary_top = max(grid_last_row, data_block_last_row) + 2

        # Agregar/agrup ar solo items del rep_cli:
        presentaciones_info = defaultdict(lambda: {'cajas': 0, 'eq11': 0.0})
        total_cajas = 0
        total_eq11  = 0.0

        for it in items_cliente:
            k = (it.presentation.name, it.size)
            presentaciones_info[k]['cajas'] += it.quantity
            eq = it.quantity * float(it.presentation.conversion_factor)
            presentaciones_info[k]['eq11']  += eq
            total_cajas += it.quantity
            total_eq11  += eq

        def merge_pair(row, c1, c2, value=None, *, font=None, fill=None, border=None):
            for cc in range(c1, c2 + 1):
                cell = ws.cell(row=row, column=cc)
                if font:  cell.font = font
                if fill:  cell.fill = fill
                if border: cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
            if value is not None:
                ws.cell(row=row, column=c1, value=value)

        # Encabezados (2 col por campo)
        header_pairs = [(1,2), (3,4), (5,6), (7,8)]
        headers = ["Presentación", "Tamaño", "Cantidad", "Equiv. 11 lbs"]
        for (c1, c2), txt in zip(header_pairs, headers):
            merge_pair(summary_top, c1, c2, txt, font=th_font, fill=th_fill, border=thin_border)

        # Filas
        r = summary_top + 1
        if presentaciones_info:
            for (pres, size), info in sorted(presentaciones_info.items(),
                                     key=lambda kv: (kv[0][0].lower(), str(kv[0][1]).lower())):
                merge_pair(r, 1, 2, pres,               font=body_font, border=thin_border)
                merge_pair(r, 3, 4, size,               font=body_font, border=thin_border)
                merge_pair(r, 5, 6, info['cajas'],      font=body_font, border=thin_border)
                merge_pair(r, 7, 8, round(info['eq11'],2), font=body_font, border=thin_border)
                r += 1
        else:
            merge_pair(r, 1, 8, "(Sin datos)", font=body_font, border=thin_border)
            r += 1

        # Totales
        r += 1
        tot_fill = PatternFill("solid", fgColor="BBDDFF")
        bold = Font(bold=True)
        tot_label_font = Font(name='Calibri', size=16, bold=True)  # etiqueta "TOTALES"
        tot_num_font   = Font(name='Calibri', size=18, bold=True)  # números grandes
        merge_pair(r, 1, 2, "TOTALES",               font=tot_label_font, fill=tot_fill, border=thin_border)
        merge_pair(r, 3, 4, "",                      font=tot_label_font, fill=tot_fill, border=thin_border)
        merge_pair(r, 5, 6, total_cajas,             font=tot_num_font,   fill=tot_fill, border=thin_border)
        merge_pair(r, 7, 8, round(total_eq11, 2),    font=tot_num_font,   fill=tot_fill, border=thin_border)



        # ===== Totales por Presentación (a la derecha) =====
        # Arranca en J32 (columna 10), 2 columnas por campo, mismo look & feel
        tot_pres = defaultdict(lambda: {'cajas': 0, 'eq11': 0.0})
        for it in items_cliente:  # <- SOLO del embarque rep_cli
            tot_pres[it.presentation.name]['cajas'] += it.quantity
            tot_pres[it.presentation.name]['eq11']  += it.quantity * float(it.presentation.conversion_factor)

        right_top_row = 32
        col_left = 10  # J

        def merge_pair_abs(row, c1, c2, value=None, *, font=None, fill=None, border=None):
            # Igual que merge_pair pero con columnas absolutas
            for cc in range(c1, c2 + 1):
                cell = ws.cell(row=row, column=cc)
                if font:  cell.font = font
                if fill:  cell.fill = fill
                if border: cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
            if value is not None:
                ws.cell(row=row, column=c1, value=value)

        # Encabezados: Presentación (J–K), Total (L–M), Equiv. 11 lbs (N–O)
        merge_pair_abs(right_top_row,     col_left,     col_left+1, "Presentación",   font=th_font, fill=th_fill, border=thin_border)
        merge_pair_abs(right_top_row,     col_left+2,   col_left+3, "Total",          font=th_font, fill=th_fill, border=thin_border)
        merge_pair_abs(right_top_row,     col_left+4,   col_left+5, "Equiv. 11 lbs",  font=th_font, fill=th_fill, border=thin_border)

        rr = right_top_row + 1
        if tot_pres:
            for pres, agg in sorted(tot_pres.items(), key=lambda kv: kv[0].lower()):
                merge_pair_abs(rr, col_left,   col_left+1, pres,                 font=body_font, border=thin_border)
                merge_pair_abs(rr, col_left+2, col_left+3, agg['cajas'],         font=body_font, border=thin_border)
                merge_pair_abs(rr, col_left+4, col_left+5, round(agg['eq11'],2), font=body_font, border=thin_border)
                rr += 1
        else:
            merge_pair_abs(rr, col_left, col_left+5, "(Sin datos)", font=body_font, border=thin_border)
            rr += 1

        # Fila de total general (mismo estilo que la izquierda)
        rr += 1
        tot_fill_right = PatternFill("solid", fgColor="BBDDFF")
        merge_pair_abs(rr, col_left,   col_left+1, "TOTAL", font=tot_label_font, fill=tot_fill_right, border=thin_border)
        merge_pair_abs(rr, col_left+2, col_left+3, sum(v['cajas'] for v in tot_pres.values()), font=tot_num_font, fill=tot_fill_right, border=thin_border)
        merge_pair_abs(rr, col_left+4, col_left+5, round(sum(v['eq11'] for v in tot_pres.values()), 2), font=tot_num_font, fill=tot_fill_right, border=thin_border)


        # --- Exportar ---
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"reporte_{report_date}_{slugify(cliente)}.xlsx"
        response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


    # ---- HTML normal ----- 


    for s in qs:
        for item in s.items.all():
            item.eq_11lbs_calc = item.quantity * float(item.presentation.conversion_factor)
            item.amount_calc   = item.quantity * float(item.presentation.price)

    return render(request, 'empaques/daily_report.html', {
        'report_date': report_date,
        'shipments': qs,
        'total_boxes': total_boxes,
        'total_eq_11lbs': total_eq_11lbs,
        'total_amount': total_amount,
        'clientes': clientes_slug,
    })
