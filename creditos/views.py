from collections import OrderedDict
from datetime import timedelta
from decimal import Decimal

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Prefetch
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.utils.text import slugify
from django.views.decorators.http import require_http_methods

from .forms import AbonoForm, CreditoForm, GarantiaForm
from .models import (
    BANCO_CHOICES, EMPRESA_CHOICES, MONEDA_CHOICES, SIMBOLO_MONEDA,
    TIPO_CREDITO_CHOICES, Abono, Credito, Garantia,
)

# Menú de proximidad a vencer. clave -> (etiqueta, días); None = sin tope.
VENCE_FILTROS = OrderedDict([
    ("todos",    ("Todos",            None)),
    ("vencidos", ("Vencidos",         -1)),
    ("30",       ("Vencen en 30 días", 30)),
    ("60",       ("Vencen en 60 días", 60)),
    ("90",       ("Vencen en 90 días", 90)),
])

ORDEN_OPCIONES = OrderedDict([
    ("vencimiento", "Más próximos a vencer"),
    ("saldo",       "Mayor saldo"),
    ("monto",       "Mayor monto"),
    ("reciente",    "Disposición más reciente"),
])


def _totales_por_moneda(creditos):
    """Suma monto/abonado/saldo agrupado por moneda (no se mezclan divisas)."""
    acc = {}
    for c in creditos:
        d = acc.setdefault(c.moneda, {
            "moneda": c.moneda,
            "simbolo": SIMBOLO_MONEDA.get(c.moneda, "$"),
            "monto": Decimal("0"),
            "interes": Decimal("0"),
            "total": Decimal("0"),
            "abonado": Decimal("0"),
            "saldo": Decimal("0"),
            "n": 0,
        })
        d["monto"]   += Decimal(c.monto or 0)
        d["interes"] += c.interes
        d["total"]   += c.total_a_pagar
        d["abonado"] += c.total_abonado
        d["saldo"]   += c.saldo
        d["n"]       += 1

    salida = []
    for k in sorted(acc):
        d = acc[k]
        # Ya formateados con separador de miles (floatformat no agrupa)
        for campo in ("monto", "interes", "total", "abonado", "saldo"):
            d[campo + "_fmt"] = f"{d['simbolo']}{d[campo]:,.2f}"
        salida.append(d)
    return salida


def _aplicar_filtros(request):
    """
    Resuelve los filtros de la lista de créditos.

    La usan tanto la pantalla como la exportación a Excel, para que el archivo
    descargado sea exactamente lo que el usuario está viendo.

    Devuelve (creditos_ordenados, dict_de_filtros).
    """
    hoy = timezone.localdate()

    empresa   = (request.GET.get("empresa") or "").strip()
    banco     = (request.GET.get("banco") or "").strip()
    moneda    = (request.GET.get("moneda") or "").strip()
    vence     = (request.GET.get("vence") or "todos").strip()
    orden     = (request.GET.get("orden") or "vencimiento").strip()
    ocultar_liq = request.GET.get("ocultar_liquidados") == "1"

    if vence not in VENCE_FILTROS:
        vence = "todos"
    if orden not in ORDEN_OPCIONES:
        orden = "vencimiento"

    qs = Credito.objects.select_related("garantia").prefetch_related(
        Prefetch("abonos", queryset=Abono.objects.order_by("-fecha", "-id"))
    )

    if empresa:
        qs = qs.filter(empresa=empresa)
    if banco:
        qs = qs.filter(banco=banco)
    if moneda:
        qs = qs.filter(moneda=moneda)

    # Filtro por proximidad de vencimiento (a nivel de base de datos)
    dias = VENCE_FILTROS[vence][1]
    if vence == "vencidos":
        qs = qs.filter(fecha_vencimiento__lt=hoy)
    elif dias is not None:
        qs = qs.filter(fecha_vencimiento__gte=hoy,
                       fecha_vencimiento__lte=hoy + timedelta(days=dias))

    creditos = list(qs)

    # 'liquidado' y 'saldo' dependen de los abonos: se resuelven en Python
    if ocultar_liq:
        creditos = [c for c in creditos if not c.liquidado]

    if orden == "saldo":
        creditos.sort(key=lambda c: c.saldo, reverse=True)
    elif orden == "monto":
        creditos.sort(key=lambda c: c.monto or 0, reverse=True)
    elif orden == "reciente":
        creditos.sort(key=lambda c: c.fecha_disposicion or hoy, reverse=True)
    else:
        # Más próximos a vencer primero; los liquidados al final
        creditos.sort(key=lambda c: (c.liquidado, c.fecha_vencimiento or hoy))

    filtros = {
        "empresa": empresa, "banco": banco, "moneda": moneda,
        "vence": vence, "orden": orden, "ocultar_liquidados": ocultar_liq,
    }
    return creditos, filtros


@login_required
def credito_list(request):
    hoy = timezone.localdate()
    creditos, filtros = _aplicar_filtros(request)

    # Conteos para las pestañas del menú
    todos = list(Credito.objects.all())
    pendientes = [c for c in todos if not c.liquidado]
    conteos = {
        "todos":    len(todos),
        "vencidos": sum(1 for c in pendientes if (c.dias_para_vencer or 0) < 0),
        "30":       sum(1 for c in pendientes if 0 <= (c.dias_para_vencer or -1) <= 30),
        "60":       sum(1 for c in pendientes if 0 <= (c.dias_para_vencer or -1) <= 60),
        "90":       sum(1 for c in pendientes if 0 <= (c.dias_para_vencer or -1) <= 90),
    }

    # Querystring actual, para que el botón de Excel exporte lo mismo que se ve
    qs_actual = request.GET.urlencode()

    return render(request, "creditos/credito_list.html", {
        "creditos": creditos,
        "totales": _totales_por_moneda(creditos),
        "hoy": hoy,
        "f": filtros,
        "qs_actual": qs_actual,
        "empresas": EMPRESA_CHOICES,
        "bancos": BANCO_CHOICES,
        "monedas": MONEDA_CHOICES,
        "vence_filtros": [(k, v[0], conteos.get(k, 0)) for k, v in VENCE_FILTROS.items()],
        "orden_opciones": list(ORDEN_OPCIONES.items()),
    })


@login_required
@require_http_methods(["GET", "POST"])
def credito_create(request):
    if request.method == "POST":
        form = CreditoForm(request.POST)
        if form.is_valid():
            credito = form.save()
            messages.success(request, "Crédito registrado correctamente.")
            return redirect("creditos:credito_detail", pk=credito.pk)
        messages.error(request, "Revisa los campos marcados.")
    else:
        form = CreditoForm()

    return render(request, "creditos/credito_form.html", {
        "form": form, "credito": None, "titulo": "Nuevo crédito",
    })


@login_required
@require_http_methods(["GET", "POST"])
def credito_edit(request, pk):
    credito = get_object_or_404(Credito, pk=pk)

    if request.method == "POST":
        form = CreditoForm(request.POST, instance=credito)
        if form.is_valid():
            form.save()
            messages.success(request, "Crédito actualizado.")
            return redirect("creditos:credito_detail", pk=credito.pk)
        messages.error(request, "Revisa los campos marcados.")
    else:
        form = CreditoForm(instance=credito)

    return render(request, "creditos/credito_form.html", {
        "form": form, "credito": credito, "titulo": "Editar crédito",
    })


@login_required
@require_http_methods(["GET", "POST"])
def credito_detail(request, pk):
    credito = get_object_or_404(
        Credito.objects.select_related("garantia"), pk=pk,
    )

    if request.method == "POST":
        abono_form = AbonoForm(request.POST, credito=credito)
        if abono_form.is_valid():
            abono = abono_form.save(commit=False)
            abono.credito = credito
            abono.save()
            messages.success(request, f"Abono de {abono.monto_fmt} registrado.")
            return redirect("creditos:credito_detail", pk=credito.pk)
        messages.error(request, "No se pudo registrar el abono.")
    else:
        abono_form = AbonoForm(credito=credito)

    return render(request, "creditos/credito_detail.html", {
        "credito": credito,
        "abonos": credito.abonos.all(),
        "abono_form": abono_form,
    })


@login_required
@require_http_methods(["POST"])
def abono_delete(request, pk):
    abono = get_object_or_404(Abono, pk=pk)
    credito_pk = abono.credito_id
    abono.delete()
    messages.success(request, "Abono eliminado.")
    return redirect("creditos:credito_detail", pk=credito_pk)


@login_required
def credito_export_xlsx(request):
    """
    Exporta a Excel EXACTAMENTE los créditos que se están viendo en la lista:
    mismos filtros, mismo orden y mismas columnas.
    """
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    creditos, filtros = _aplicar_filtros(request)
    hoy = timezone.localdate()

    wb = Workbook()
    ws = wb.active
    ws.title = "Créditos"

    thin = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    th_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    th_fill = PatternFill("solid", fgColor="1E3A5F")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center")

    # ── Título y filtros aplicados ──
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=16)
    c = ws.cell(row=1, column=1, value="Créditos")
    c.font = Font(name="Calibri", size=16, bold=True, color="1E3A5F")
    c.alignment = Alignment(horizontal="left", vertical="center")

    partes = [f"Generado {hoy.strftime('%d/%m/%Y')}"]
    if filtros["empresa"]:
        partes.append("Empresa: " + dict(EMPRESA_CHOICES).get(filtros["empresa"], filtros["empresa"]))
    if filtros["banco"]:
        partes.append("Banco: " + dict(BANCO_CHOICES).get(filtros["banco"], filtros["banco"]))
    if filtros["moneda"]:
        partes.append("Moneda: " + filtros["moneda"])
    partes.append("Vencimiento: " + VENCE_FILTROS[filtros["vence"]][0])
    partes.append("Orden: " + ORDEN_OPCIONES[filtros["orden"]])
    if filtros["ocultar_liquidados"]:
        partes.append("Sin liquidados")

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=16)
    c = ws.cell(row=2, column=1, value="  ·  ".join(partes))
    c.font = Font(name="Calibri", size=10, italic=True, color="6D6D6D")
    c.alignment = Alignment(horizontal="left", vertical="center")

    # ── Encabezados (mismas columnas que la tabla en pantalla) ──
    headers = [
        ("Empresa", 20), ("Banco", 20), ("Tipo", 24), ("Garantía", 22),
        ("Moneda", 9), ("Monto", 16), ("Tasa (%)", 11), ("Interés", 16),
        ("Abonado", 16), ("Saldo", 16),
        ("Pagos", 9), ("Frecuencia", 14), ("Plazo (meses)", 13),
        ("Disposición", 13), ("Vencimiento", 13), ("Estado", 13),
    ]
    r = 4
    for col, (h, w) in enumerate(headers, start=1):
        cell = ws.cell(row=r, column=col, value=h)
        cell.font = th_font
        cell.fill = th_fill
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[r].height = 24
    r += 1

    fmt_money = '#,##0.00'
    fill_vencido = PatternFill("solid", fgColor="FDECEC")

    for cr in creditos:
        vals = [
            cr.get_empresa_display(),
            cr.get_banco_display(),
            cr.tipo_credito_label,
            cr.garantia.nombre if cr.garantia else "",
            cr.moneda,
            float(cr.monto or 0),
            float(cr.tasa) if cr.tasa is not None else None,
            float(cr.interes),
            float(cr.total_abonado),
            float(cr.saldo),
            cr.cantidad_pagos,
            cr.frecuencia_label if cr.frecuencia_pagos else "",
            cr.plazo_meses,
            cr.fecha_disposicion,
            cr.fecha_vencimiento,
            cr.estado_label,
        ]
        for col, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=col, value=v)
            cell.border = border
            if col in (6, 8, 9, 10):
                cell.number_format = fmt_money
                cell.alignment = right
            elif col == 7:
                cell.number_format = '0.000'
                cell.alignment = right
            elif col in (11, 12, 13, 16):
                cell.alignment = center
            elif col in (14, 15):
                cell.number_format = 'dd/mm/yyyy'
                cell.alignment = center
            if cr.estado == "vencido":
                cell.fill = fill_vencido
        r += 1

    if not creditos:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=16)
        c = ws.cell(row=r, column=1, value="No hay créditos que coincidan con el filtro.")
        c.alignment = center
        c.font = Font(italic=True, color="9CA3AF")
        r += 1

    # ── Totales por moneda (no se mezclan divisas) ──
    r += 1
    for t in _totales_por_moneda(creditos):
        ws.cell(row=r, column=1, value=f"TOTAL {t['moneda']}").font = Font(bold=True)
        ws.cell(row=r, column=5, value=f"{t['n']} créditos").alignment = center
        for col, key in ((6, "monto"), (8, "interes"),
                         (9, "abonado"), (10, "saldo")):
            cell = ws.cell(row=r, column=col, value=float(t[key]))
            cell.font = Font(bold=True)
            cell.number_format = fmt_money
            cell.alignment = right
            cell.border = border
        r += 1

    ws.freeze_panes = "A5"

    out = BytesIO()
    wb.save(out)
    out.seek(0)

    partes_nombre = ["creditos"]
    if filtros["empresa"]:
        partes_nombre.append(slugify(filtros["empresa"]))
    if filtros["banco"]:
        partes_nombre.append(slugify(filtros["banco"]))
    if filtros["vence"] != "todos":
        partes_nombre.append(filtros["vence"])
    partes_nombre.append(hoy.isoformat())
    fname = "_".join(partes_nombre) + ".xlsx"

    resp = HttpResponse(
        out.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{fname}"'
    return resp


@login_required
def credito_plan_xlsx(request, pk):
    """
    Calendario de pagos de UN crédito, en Excel.

    ?solo=pendientes  -> únicamente los pagos que faltan por cubrir
    (por omisión salen todos, pagados y pendientes).
    """
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    credito = get_object_or_404(Credito.objects.select_related("garantia"), pk=pk)
    solo_pendientes = (request.GET.get("solo") or "").lower() == "pendientes"

    pagos = credito.pagos_pendientes() if solo_pendientes else credito.plan_pagos()

    if not credito.tiene_plan:
        return HttpResponse(
            "Este crédito no tiene calendario: falta capturar la cantidad de pagos "
            "y cada cuánto se paga.", status=400)

    wb = Workbook()
    ws = wb.active
    ws.title = "Pendientes" if solo_pendientes else "Calendario"

    thin = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    th_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    th_fill = PatternFill("solid", fgColor="1E3A5F")
    anio_fill = PatternFill("solid", fgColor="2E67D1")
    pagado_fill = PatternFill("solid", fgColor="DCFCE7")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center")
    fmt_money = '#,##0.00'

    # ── Título ──
    titulo = ("Pagos pendientes" if solo_pendientes else "Calendario de pagos")
    ws.cell(row=1, column=1,
            value=f"{titulo} · {credito.get_empresa_display()} · {credito.get_banco_display()}"
            ).font = Font(name="Calibri", size=16, bold=True, color="1E3A5F")

    sub = (f"{credito.cantidad_pagos} pagos {credito.frecuencia_label.lower()} de "
           f"{credito.monto_por_pago_fmt}  ·  Generado {timezone.localdate().strftime('%d/%m/%Y')}")
    ws.cell(row=2, column=1, value=sub).font = Font(name="Calibri", size=10,
                                                    italic=True, color="6D6D6D")

    # ── Columnas fijas (encabezado en filas 4-5, valores en fila 6) ──
    r_anio, r_head, r_data = 4, 5, 6

    fijas = [
        ("Empresa",                  credito.get_empresa_display(),                 20),
        ("Crédito",                  credito.tipo_credito_label,                    24),
        ("Garantía",                 credito.garantia.nombre if credito.garantia else "", 22),
        ("Tasa (%)",                 float(credito.tasa) if credito.tasa is not None else None, 11),
        ("Moneda",                   credito.moneda,                                 9),
        ("Plazo (meses)",            credito.plazo_meses,                           13),
        ("Fecha de crédito",         credito.fecha_contratacion,                    15),
        ("Monto del crédito",        float(credito.monto or 0),                     17),
        ("Fecha de disposición",     credito.fecha_disposicion,                     17),
    ]

    for col, (etiqueta, valor, ancho) in enumerate(fijas, start=1):
        ws.merge_cells(start_row=r_anio, start_column=col, end_row=r_head, end_column=col)
        h = ws.cell(row=r_anio, column=col, value=etiqueta)
        h.font, h.fill, h.alignment, h.border = th_font, th_fill, center, border
        ws.cell(row=r_head, column=col).border = border

        c = ws.cell(row=r_data, column=col, value=valor)
        c.border = border
        if etiqueta == "Monto del crédito":
            c.number_format = fmt_money
            c.alignment = right
        elif etiqueta == "Tasa (%)":
            c.number_format = '0.000'
            c.alignment = right
        elif etiqueta.startswith("Fecha"):
            c.number_format = 'dd/mm/yyyy'
            c.alignment = center
        else:
            c.alignment = center
        ws.column_dimensions[get_column_letter(col)].width = ancho

    # ── Bloques de pago: por cada pago, columna de fecha + columna de importe,
    #    agrupados bajo el año (celda combinada arriba) ──
    col = len(fijas) + 1
    inicio_anio = col
    anio_actual = pagos[0]["anio"] if pagos else None

    def cerrar_anio(hasta_col, anio):
        if anio is None or hasta_col < inicio_anio:
            return
        ws.merge_cells(start_row=r_anio, start_column=inicio_anio,
                       end_row=r_anio, end_column=hasta_col)
        c = ws.cell(row=r_anio, column=inicio_anio, value=anio)
        c.font, c.fill, c.alignment, c.border = th_font, anio_fill, center, border
        for cc in range(inicio_anio, hasta_col + 1):
            ws.cell(row=r_anio, column=cc).border = border

    for p in pagos:
        if p["anio"] != anio_actual:
            cerrar_anio(col - 1, anio_actual)
            anio_actual = p["anio"]
            inicio_anio = col

        # Columna de la fecha
        c = ws.cell(row=r_head, column=col, value=p["fecha_texto"])
        c.font, c.fill, c.alignment, c.border = th_font, th_fill, center, border
        ws.column_dimensions[get_column_letter(col)].width = 14
        marca = ws.cell(row=r_data, column=col,
                        value=("Pagado" if p["pagado"] else ""))
        marca.alignment, marca.border = center, border
        if p["pagado"]:
            marca.fill = pagado_fill
            marca.font = Font(size=10, bold=True, color="166534")

        # Columna del importe
        c = ws.cell(row=r_head, column=col + 1, value="$")
        c.font, c.fill, c.alignment, c.border = th_font, th_fill, center, border
        ws.column_dimensions[get_column_letter(col + 1)].width = 14
        v = ws.cell(row=r_data, column=col + 1, value=float(p["monto"]))
        v.number_format = fmt_money
        v.alignment, v.border = right, border
        if p["pagado"]:
            v.fill = pagado_fill

        col += 2

    cerrar_anio(col - 1, anio_actual)

    if not pagos:
        ws.merge_cells(start_row=r_anio, start_column=len(fijas) + 1,
                       end_row=r_data, end_column=len(fijas) + 3)
        c = ws.cell(row=r_anio, column=len(fijas) + 1,
                    value="No quedan pagos pendientes")
        c.alignment = center
        c.font = Font(italic=True, color="9CA3AF")

    # ── Total ──
    total = sum((p["monto"] for p in pagos), Decimal("0"))
    r_tot = r_data + 2
    ws.cell(row=r_tot, column=1,
            value=("Total pendiente" if solo_pendientes else "Total del calendario")
            ).font = Font(bold=True)
    c = ws.cell(row=r_tot, column=8, value=float(total))
    c.font = Font(bold=True)
    c.number_format = fmt_money
    c.alignment = right
    c.border = border

    ws.row_dimensions[r_anio].height = 20
    ws.row_dimensions[r_head].height = 26
    ws.row_dimensions[r_data].height = 22
    ws.freeze_panes = ws.cell(row=r_data, column=len(fijas) + 1)

    out = BytesIO()
    wb.save(out)
    out.seek(0)

    sufijo = "pendientes" if solo_pendientes else "calendario"
    fname = f"credito_{credito.pk}_{slugify(credito.get_empresa_display())}_{sufijo}.xlsx"
    resp = HttpResponse(
        out.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{fname}"'
    return resp


@login_required
@require_http_methods(["POST"])
def garantia_delete(request, pk):
    garantia = get_object_or_404(Garantia, pk=pk)

    n = garantia.creditos.count()
    if n:
        messages.error(
            request,
            f"No se puede quitar «{garantia.nombre}»: está usada por {n} "
            f"crédito{'s' if n != 1 else ''}."
        )
    else:
        nombre = garantia.nombre
        garantia.delete()
        messages.success(request, f"Garantía «{nombre}» eliminada.")

    return redirect("creditos:garantia_list")


@login_required
@require_http_methods(["GET", "POST"])
def garantia_list(request):
    """Catálogo de campos/terrenos dados en garantía."""
    if request.method == "POST":
        form = GarantiaForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Garantía agregada.")
            return redirect("creditos:garantia_list")
        messages.error(request, "Revisa los campos marcados.")
    else:
        form = GarantiaForm()

    return render(request, "creditos/garantia_list.html", {
        "form": form,
        "garantias": Garantia.objects.all().order_by("nombre"),
    })
