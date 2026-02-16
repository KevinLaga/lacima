# arandano/views.py (encabezado limpio)
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.db import transaction
from django.utils.text import slugify
from django.utils import timezone
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.contrib import messages

from .models import Campo, CampoVariedad, Variedad, ProduccionDia, ProduccionItem, Salida
from .forms import ProduccionDiaForm, ProduccionItemFormSet, ProduccionItemForm, SalidaItemForm, SalidaDiaForm, SalidaForm
from django.forms import inlineformset_factory
from django.urls import reverse
from decimal import Decimal, ROUND_HALF_UP
from django.db.models import Prefetch
from django.forms import formset_factory
from .models import SalidaDia, SalidaItem, DestinoSalida
from django.db.models import Sum
from collections import defaultdict
from django.contrib import messages
from decimal import Decimal
from .utils_inv import post_ledger, ScopeInv, q5
from django.conf import settings
from django.utils import timezone
from empaques.models import Presentation
from empaques.utils_inv import post_in_from_field



@login_required
def salidas_list(request):
    """
    /arandano/salidas/?campo=<id>
    Lista de salidas por día (SalidaDia) con sus renglones (SalidaItem).
    Permite filtrar por campo.
    """
    campo_id = request.GET.get("campo")
    campos = Campo.objects.filter(activo=True).order_by("nombre")

    base_qs = (
        SalidaDia.objects
        .select_related("campo")
        .prefetch_related(
            Prefetch(
                "items",
                queryset=SalidaItem.objects.select_related("variedad").order_by("variedad__nombre"),
            )
        )
        .order_by("-fecha", "-id")
    )
    if campo_id and campo_id != "all":
        base_qs = base_qs.filter(campo_id=campo_id)
        campo_sel = Campo.objects.filter(pk=campo_id).first()
    else:
        campo_sel = None

    # Totales de kg por día (suma de items) y total global (en pantalla)
    filas = []
    total_global = 0
    for s in base_qs:
        total_dia = sum((it.kg or 0) for it in s.items.all())
        total_global += float(total_dia or 0)
        filas.append({
            "obj": s,
            "total_dia": total_dia,
            "items": list(s.items.all()),
        })

    return render(request, "arandano/salidas_list.html", {
        "filas": filas,
        "campos": campos,
        "campo_sel": campo_sel,
        "total_global": total_global,
    })


@login_required
def salidas_create(request):
    """
    Crear SalidaDia + SalidaItem (formset).
    Valida que no se exceda el stock por variedad:
      disponible = (Producido kg/cs hasta fecha) - (Salidas kg/cs hasta fecha).
    """
    from .utils_inv import post_ledger, ScopeInv, q5
    # 1) Campo preseleccionado por GET ?campo=<id>
    campo_actual = None
    campo_id_get = request.GET.get("campo")
    if campo_id_get:
        campo_actual = Campo.objects.filter(pk=campo_id_get).first()
    if not campo_actual:
        campo_actual = Campo.objects.filter(activo=True).order_by("nombre").first()

    # Factory del formset
    ItemsFormSet = formset_factory(SalidaItemForm, extra=0, can_delete=False)

    if request.method == "POST":
        form = SalidaDiaForm(request.POST)
        formset = ItemsFormSet(request.POST, prefix="items")

        # Campo en POST (para limitar variedades del formset)
        campo_post = None
        if form.is_valid():
            campo_post = form.cleaned_data.get("campo")
        else:
            try:
                cid = int(request.POST.get("campo") or 0)
                campo_post = Campo.objects.filter(pk=cid).first()
            except Exception:
                campo_post = None

        # Limita variedades visibles según el campo del POST
        if campo_post:
            vids = list(
                CampoVariedad.objects.filter(
                    campo=campo_post, activo=True, variedad__activo=True
                ).values_list("variedad_id", flat=True)
            )
        else:
            vids = []

        for f in formset.forms:
            if "variedad" in f.fields:
                f.fields["variedad"].queryset = Variedad.objects.filter(pk__in=vids).order_by("nombre")
                f.fields["variedad"].empty_label = None

        if form.is_valid() and formset.is_valid():
            fecha = form.cleaned_data["fecha"]
            campo = form.cleaned_data["campo"]

            # --- Disponible por variedad (kg y clamshells) ---
            # a) Producido hasta fecha
            prod_rows = (
                ProduccionItem.objects
                .filter(produccion__campo=campo, produccion__fecha__lte=fecha)
                .values("variedad")
                .annotate(
                    kg_total=Sum("kg"),
                    cs6_total=Sum("cs_6oz"),
                    cs98_total=Sum("cs_9_8oz"),
                    cs18_total=Sum("cs_18oz"),
                )
            )
            producido_por_var = {
                row["variedad"]: {
                    "kg":   Decimal(row["kg_total"] or 0),
                    "cs6":  int(row["cs6_total"] or 0),
                    "cs98": int(row["cs98_total"] or 0),
                    "cs18": int(row["cs18_total"] or 0),
                }
                for row in prod_rows
            }

            # b) Salidas previas hasta fecha
            sal_rows = (
                SalidaItem.objects
                .filter(salida__campo=campo, salida__fecha__lte=fecha)
                .values("variedad")
                .annotate(
                    kg_total=Sum("kg"),
                    cs6_total=Sum("cs_6oz"),
                    cs98_total=Sum("cs_9_8oz"),
                    cs18_total=Sum("cs_18oz"),
                )
            )
            salidas_por_var = {
                row["variedad"]: {
                    "kg":   Decimal(row["kg_total"] or 0),
                    "cs6":  int(row["cs6_total"] or 0),
                    "cs98": int(row["cs98_total"] or 0),
                    "cs18": int(row["cs18_total"] or 0),
                }
                for row in sal_rows
            }

            # c) Disponible = producido - salidas
            disp_por_var = {}
            for vid in set(list(producido_por_var.keys()) + list(salidas_por_var.keys())):
                prod = producido_por_var.get(vid, {"kg":0, "cs6":0, "cs98":0, "cs18":0})
                sal  = salidas_por_var.get(vid,  {"kg":0, "cs6":0, "cs98":0, "cs18":0})
                disp_por_var[vid] = {
                    "kg":   Decimal(prod["kg"]) - Decimal(sal["kg"]),
                    "cs6":  int(prod["cs6"]) - int(sal["cs6"]),
                    "cs98": int(prod["cs98"]) - int(sal["cs98"]),
                    "cs18": int(prod["cs18"]) - int(sal["cs18"]),
                }

            # d) Consolidar lo solicitado en este POST por variedad
            req_por_var = {}
            for cd in formset.cleaned_data:
                if not cd:
                    continue
                v = cd.get("variedad")
                if not v:
                    continue
                req = req_por_var.setdefault(v.id, {"kg":Decimal("0"), "cs6":0, "cs98":0, "cs18":0})
                # kg
                try:
                    req["kg"] += Decimal(str(cd.get("kg") or "0"))
                except Exception:
                    pass
                # clamshells
                req["cs6"]  += int(cd.get("cs_6oz")   or 0)
                req["cs98"] += int(cd.get("cs_9_8oz") or 0)
                req["cs18"] += int(cd.get("cs_18oz")  or 0)

            # e) Validar por variedad
            hay_error = False
            for f in formset.forms:
                if not getattr(f, "cleaned_data", None):
                    continue
                v = f.cleaned_data.get("variedad")
                if not v:
                    continue

                d = disp_por_var.get(v.id, {"kg":Decimal("0"), "cs6":0, "cs98":0, "cs18":0})
                r = req_por_var.get(v.id, {"kg":Decimal("0"), "cs6":0, "cs98":0, "cs18":0})

                msgs = []
                if r["kg"] > d["kg"]:
                    msgs.append(f"kg: disp {d['kg']:.5f}, sol {r['kg']:.2f}")
                if r["cs6"] > d["cs6"]:
                    msgs.append(f"6 oz: disp {d['cs6']}, sol {r['cs6']}")
                if r["cs98"] > d["cs98"]:
                    msgs.append(f"9.8 oz: disp {d['cs98']}, sol {r['cs98']}")
                if r["cs18"] > d["cs18"]:
                    msgs.append(f"18 oz: disp {d['cs18']}, sol {r['cs18']}")

                if msgs:
                    f.add_error("kg", "No hay stock suficiente (" + "; ".join(msgs) + ").")
                    hay_error = True

            if hay_error:
                messages.error(request, "Revisa las cantidades: exceden el stock disponible.")
                return render(request, "arandano/salida_form.html", {
                    "form": form,
                    "formset": formset,
                    "campo_actual": campo_post or campo_actual,
                    "campos": Campo.objects.filter(activo=True).order_by("nombre"),
                })
            


            # Guardar
            with transaction.atomic():
                salida_dia = form.save()

                # Si destino = EMPAQUE, vamos a acumular totals de clamshells para el almacén de empaques
                tot_cs6 = tot_cs98 = tot_cs18 = 0
                is_empaque = (salida_dia.destino == DestinoSalida.EMPAQUE)  # o == "EMPAQUE"

                for cd in formset.cleaned_data:
                    if not cd:
                        continue

                    v = cd.get("variedad")
                    if not v:
                        continue

                    # parseos seguros
                    try:
                        kg = Decimal(str(cd.get("kg") or "0"))
                    except Exception:
                        kg = Decimal("0")

                    cs6  = int(cd.get("cs_6oz")   or 0)
                    cs98 = int(cd.get("cs_9_8oz") or 0)
                    cs18 = int(cd.get("cs_18oz")  or 0)

                    # 1) Guardar renglón
                    if kg > 0 or cs6 > 0 or cs98 > 0 or cs18 > 0:
                        SalidaItem.objects.create(
                            salida=salida_dia,
                            variedad=v,
                            kg=kg,
                            cs_6oz=cs6,
                            cs_9_8oz=cs98,
                            cs_18oz=cs18,
                        )

                    # 2) Salida del CAMPO (negativo en CAMPO)
                    post_ledger(
                        scope=ScopeInv.CAMPO,
                        fecha=salida_dia.fecha,
                        variedad=v,
                        kg=q5(kg),
                        cs6=cs6,
                        cs98=cs98,
                        cs18=cs18,
                        campo=salida_dia.campo,
                        ref_app="SALIDA",
                        ref_id=str(salida_dia.id),
                    )

                    # 3) Si va a EMPAQUE: espejo positivo al inventario EMPAQUE (por variedad)
                    if is_empaque:
                        post_ledger(
                            scope=ScopeInv.EMPAQUE,
                            fecha=salida_dia.fecha,
                            variedad=v,
                            kg=q5(kg),
                            cs6=cs6,
                            cs98=cs98,
                            cs18=cs18,
                            campo=None,
                            ref_app="MOV_CAMPO_A_EMPAQUE",
                            ref_id=str(salida_dia.id),
                        )

                        # acumular clamshells para el almacén de empaques (solo por tamaño, sin variedad)
                        tot_cs6  += cs6
                        tot_cs98 += cs98
                        tot_cs18 += cs18

                # 4) Registrar ENTRADA (IN) al almacén de empaques por SKU (una vez por salida)
                if is_empaque:
                    from empaques.models import InventoryItem, InventoryMovement

                    def _in_sku(sku: str, qty: int):
                        if qty <= 0:
                            return
                        item = InventoryItem.objects.filter(sku=sku).first()
                        if not item:
                            raise InventoryItem.DoesNotExist(
                                f"No existe InventoryItem con sku={sku}. Crea ese artículo en Almacén."
                            )
                        InventoryMovement.objects.create(
                            item=item,
                            date=salida_dia.fecha,
                            type="IN",
                            quantity=Decimal(qty),
                            reference=f"Salida Campo→Empaque (Salida {salida_dia.id})",
                            notes=(salida_dia.notas or ""),
                            created_by=request.user if request.user.is_authenticated else None,
                        )

                    _in_sku("AR-CS6",  tot_cs6)
                    _in_sku("AR-CS98", tot_cs98)
                    _in_sku("AR-CS18", tot_cs18)


            messages.success(request, "Salida registrada correctamente.")
            # Redirige conservando ?campo=<id>
            url = reverse("arandano:salidas_list")
            if campo_post:
                url += f"?campo={campo_post.id}"
            return redirect(url)

        # POST inválido → re-render
        return render(request, "arandano/salida_form.html", {
            "form": form,
            "formset": formset,
            "campo_actual": campo_post or campo_actual,
            "campos": Campo.objects.filter(activo=True).order_by("nombre"),
        })

    # GET: construir initial por variedades del campo actual
    form = SalidaDiaForm(initial={"campo": campo_actual.id if campo_actual else None})

    vids = []
    if campo_actual:
        vids = list(
            CampoVariedad.objects
            .filter(campo=campo_actual, activo=True, variedad__activo=True)
            .values_list("variedad_id", flat=True)
        )

    # filas iniciales: una por variedad del campo, kg y clamshells en 0
    initial_rows = [
        {"variedad": vid, "kg": Decimal("0.00"), "cs_6oz": 0, "cs_9_8oz": 0, "cs_18oz": 0}
        for vid in vids
    ]

    formset = ItemsFormSet(initial=initial_rows, prefix="items")

    # Limitar queryset de variedad por campo
    for f in formset.forms:
        if "variedad" in f.fields:
            f.fields["variedad"].queryset = Variedad.objects.filter(pk__in=vids).order_by("nombre")
            f.fields["variedad"].empty_label = None

    return render(request, "arandano/salida_form.html", {
        "form": form,
        "formset": formset,
        "campo_actual": campo_actual,
        "campos": Campo.objects.filter(activo=True).order_by("nombre"),
    })

from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.shortcuts import render, redirect
from django.db import transaction
from django.forms import formset_factory
from django.forms.widgets import HiddenInput
from decimal import Decimal

from .models import Campo, CampoVariedad, Variedad, ProduccionDia, ProduccionItem
from .forms import ProduccionDiaForm, ProduccionItemForm



@login_required
def produccion_create(request):
    """
    /arandano/nuevo/
    GET: filas fijas por variedad del campo; 'variedad' va oculta y amarrada a esa fila.
    POST: guarda ProduccionDia + renglones (kg y clamshells) con Decimal.
    """

    # --- 1) Resolver campo_actual ---
    campo_actual = None
    campo_id_get = request.GET.get("campo")

    if request.method == "POST":
        tmp = ProduccionDiaForm(request.POST)
        if tmp.is_valid():
            campo_actual = tmp.cleaned_data.get("campo")
        else:
            cid = request.POST.get("campo")
            campo_actual = Campo.objects.filter(pk=cid).first() if cid else None
    else:
        if campo_id_get:
            campo_actual = Campo.objects.filter(pk=campo_id_get).first()
        if not campo_actual:
            campo_actual = Campo.objects.filter(activo=True).order_by("nombre").first()

    # --- 2) Orden y lista de variedades del campo (fija y estable) ---
    variedades_qs = Variedad.objects.filter(
        id__in=CampoVariedad.objects.filter(
            campo=campo_actual, activo=True, variedad__activo=True
        ).values_list("variedad_id", flat=True)
    ).order_by("nombre")

    variedad_ids = list(variedades_qs.values_list("id", flat=True))
    # Para mostrar el nombre junto a cada formulario:
    variedades_ctx = list(variedades_qs)  # misma posición que el formset

    # --- 3) initial rows para GET (una por variedad) ---
    qs_var = Variedad.objects.filter(pk__in=variedad_ids).order_by("nombre")
    variedades_ctx = list(qs_var)  # en el MISMO orden que initial_rows
    initial_rows = [{"variedad": vid, "kg": Decimal("0.00000")} for vid in variedad_ids]

    # --- 4) Formset sin extras ---
    ItemsFormSet = formset_factory(ProduccionItemForm, extra=0, can_delete=False)

    def lock_variedad_field(fs):
        """
        Convierte 'variedad' en HiddenInput y restringe las opciones a la variedad de esa fila.
        Esto asegura que items-0-* siempre corresponde a la misma variedad que se muestra.
        """
        for idx, f in enumerate(fs.forms):
            if "variedad" in f.fields:
                f.fields["variedad"].widget = HiddenInput()
                # Opción única: la variedad de la misma posición
                if idx < len(variedad_ids):
                    only_id = variedad_ids[idx]
                    f.fields["variedad"].queryset = Variedad.objects.filter(id=only_id)
                    f.fields["variedad"].empty_label = None
                    # Asegura initial por si el navegador no envía el hidden (o JS raro)
                    f.fields["variedad"].initial = only_id

    # =======================
    # POST
    # =======================
    if request.method == "POST":
        form = ProduccionDiaForm(request.POST)
        formset = ItemsFormSet(request.POST, prefix="items")

        # Fijar 'variedad' como hidden y con opción única
        lock_variedad_field(formset)

        if form.is_valid() and formset.is_valid():
            with transaction.atomic():
                produccion = form.save()  # fecha, campo, rezaga_kg, notas

                # Limpia por si reenvían
                ProduccionItem.objects.filter(produccion=produccion).delete()

                for idx, cd in enumerate(formset.cleaned_data):
                    if not cd:
                        continue
                    variedad = cd.get("variedad")  # ya viene fijada por hidden
                    # Normaliza a Decimal
                    try:
                        kg = Decimal(str(cd.get("kg") or "0"))
                    except Exception:
                        kg = Decimal("0")

                    cs6  = int(cd.get("cs_6oz")   or 0)
                    cs98 = int(cd.get("cs_9_8oz") or 0)
                    cs18 = int(cd.get("cs_18oz")  or 0)

                    if variedad:
                        ProduccionItem.objects.create(
                            produccion=produccion,
                            variedad=variedad,
                            kg=kg,                 # Decimal
                            cs_6oz=int(cs6),
                            cs_9_8oz=int(cs98),
                            cs_18oz=int(cs18),
                        )

                        # 👉 Apunta al ledger (entrada positiva en CAMPO)
                        post_ledger(
                            scope=ScopeInv.CAMPO,
                            fecha=produccion.fecha,
                            variedad=variedad,
                            kg=q5(kg),
                            cs6=int(cs6),
                            cs98=int(cs98),
                            cs18=int(cs18),
                            campo=produccion.campo,
                            ref_app="PROD",
                            ref_id=str(produccion.id),
                        )

            messages.success(request, "Producción guardada correctamente.")
            return redirect("arandano:arandano_produccion_list")

        # POST inválido → mostrar errores y re-render
        if not form.is_valid():
            messages.error(request, f"Errores en cabecera: {form.errors.as_text()}")
        if not formset.is_valid():
            messages.error(request, f"Errores en renglones: {formset.errors}")

        # Re-render conservando el zip para mostrar nombres
        filas = list(zip(formset.forms, variedades_ctx))
        return render(request, "arandano/produccion_form.html", {
            "form": form,
            "formset": formset,
            "campo_actual": campo_actual,
            "campos": Campo.objects.filter(activo=True).order_by("nombre"),
            "filas": filas,
        })

    # =======================
    # GET
    # =======================
    form = ProduccionDiaForm(initial={"campo": campo_actual.id if campo_actual else None})
    formset = ItemsFormSet(initial=initial_rows, prefix="items")
    lock_variedad_field(formset)

    filas = list(zip(formset.forms, variedades_ctx))
    return render(request, "arandano/produccion_form.html", {
        "form": form,
        "formset": formset,
        "campo_actual": campo_actual,
        "campos": Campo.objects.filter(activo=True).order_by("nombre"),
        "filas": filas,
    })  



# arandano/views.py
from django.db.models import Prefetch, Sum
from decimal import Decimal, ROUND_HALF_UP

@login_required
def produccion_list(request):
    """
    Lista de producción:
    - Filtro por campo opcional (?campo=<id> o 'all')
    - Totales por campo = Entradas (ProduccionItem.kg) - Salidas (SalidaItem.kg)
    - Además construye 'filas' con totales por día para la tabla inferior.
    """
    from decimal import Decimal, ROUND_HALF_UP
    from django.db.models import Prefetch
    from .utils_inv import stock_por_campo, ScopeInv

    Q2 = Decimal('0.00001') 
    def q2(x):
        if x is None:
            x = Decimal('0')
        if not isinstance(x, Decimal):
            x = Decimal(str(x))
        return x.quantize(Q2, rounding=ROUND_HALF_UP)
    
    # --- resolver campo_sel desde ?campo= ---
    campo_param = request.GET.get("campo")  # 'all' | None | '<id>'
    campo_sel = None
    if campo_param and campo_param != "all":
        try:
            campo_sel = Campo.objects.get(pk=int(campo_param))
        except (ValueError, Campo.DoesNotExist):
            campo_sel = None  # fallback

    # --- stock de inventario (si tienes util) ---
    stock_campo = None
    if campo_sel:
        # Si tu helper acepta hasta=None, deja así:
        stock_campo = stock_por_campo(campo=campo_sel, hasta_fecha=None)

    campos = Campo.objects.filter(activo=True).order_by("nombre")

    

    # Filtro de campo
    campo_param = request.GET.get("campo")
    campo_sel = None
    if campo_param and campo_param != "all":
        try:
            campo_sel = Campo.objects.get(pk=int(campo_param))
        except Exception:
            campo_sel = None

    # -------- Producciones (tabla inferior) --------
    qs = (
        ProduccionDia.objects
        .select_related("campo")
        .prefetch_related(Prefetch("items", queryset=ProduccionItem.objects.select_related("variedad")))
        .order_by("-fecha", "-id")
    )
    if campo_sel:
        qs = qs.filter(campo=campo_sel)

    # Construye filas con totales por producción (para usar en template si quieres)
    filas = []
    for p in qs:
        total = q2(sum(Decimal(str(it.kg or 0)) for it in p.items.all()))
        rezaga = q2(getattr(p, "rezaga_kg", 0))
        empacado = q2(max(total - rezaga, Decimal('0')))
        porc = q2((rezaga * Decimal('100') / total) if total > 0 else Decimal('0'))
        filas.append({
            "obj": p,
            "total": total,
            "rezaga": rezaga,
            "empacado": empacado,
            "porc": porc,
        })

    # -------- Totales capturados (arriba): producido / salidas / neto por campo --------
    tot_prod_por_campo = defaultdict(lambda: Decimal('0'))
    tot_sal_por_campo = defaultdict(lambda: Decimal('0'))

    # Entradas
    for p in qs:
        total_dia = sum(Decimal(str(it.kg or 0)) for it in p.items.all())
        tot_prod_por_campo[p.campo.nombre] += q2(total_dia)

    # Salidas (respeta filtro)
    salidas_qs = (
        SalidaDia.objects
        .select_related("campo")
        .prefetch_related(Prefetch("items", queryset=SalidaItem.objects.select_related("variedad")))
    )
    if campo_sel:
        salidas_qs = salidas_qs.filter(campo=campo_sel)

    for s in salidas_qs:
        total_salida = sum(Decimal(str(it.kg or 0)) for it in s.items.all())
        tot_sal_por_campo[s.campo.nombre] += q2(total_salida)

    # Tabla de totales por campo
    nombres_campos = sorted({*tot_prod_por_campo.keys(), *tot_sal_por_campo.keys()})
    tabla_totales = []
    total_prod_general = Decimal('0')
    total_sal_general = Decimal('0')

    for nombre in nombres_campos:
        prod = q2(tot_prod_por_campo.get(nombre, Decimal('0')))
        sal = q2(tot_sal_por_campo.get(nombre, Decimal('0')))
        neto = q2(prod - sal)
        tabla_totales.append({
            "campo": nombre,
            "producido": prod,
            "salidas": sal,
            "neto": neto,
        })
        total_prod_general += prod
        total_sal_general += sal

    total_neto_general = q2(total_prod_general - total_sal_general)

    # ===== NUEVO: Inventario por variedad =====
    # Producido por variedad
    prod_items = ProduccionItem.objects.select_related("variedad", "produccion", "produccion__campo")
    if campo_sel:
        prod_items = prod_items.filter(produccion__campo=campo_sel)
    prod_agg = (
        prod_items.values("variedad__nombre")
                  .annotate(kg_total=Sum("kg"))
    )
    prod_map = {row["variedad__nombre"]: Decimal(row["kg_total"] or 0) for row in prod_agg}

    # Salidas por variedad
    sal_items = SalidaItem.objects.select_related("variedad", "salida", "salida__campo")
    if campo_sel:
        sal_items = sal_items.filter(salida__campo=campo_sel)
    sal_agg = (
        sal_items.values("variedad__nombre")
                 .annotate(kg_total=Sum("kg"))
    )
    sal_map = {row["variedad__nombre"]: Decimal(row["kg_total"] or 0) for row in sal_agg}

    # Stock = producido - salidas
    var_nombres = sorted(set(prod_map.keys()) | set(sal_map.keys()))
    inventario_variedades = []
    total_stock_general = Decimal('0')
    for vn in var_nombres:
        stock = q2(prod_map.get(vn, Decimal('0')) - sal_map.get(vn, Decimal('0')))
        inventario_variedades.append({"variedad": vn, "stock": stock})
        total_stock_general += stock


        


    # Render
    return render(request, "arandano/produccion_list.html", {
        "campos": campos,
        "campo_sel": campo_sel,
        "producciones": qs,   # por si tu template aún usa este queryset
        "filas": filas,       # si prefieres mostrar totales pre-calculados por fila
        "tabla_totales": tabla_totales,
        "total_prod_general": total_prod_general,
        "total_sal_general": total_sal_general,
        "total_neto_general": total_neto_general,
        "inventario_variedades": inventario_variedades,
        "total_stock_general": total_stock_general,
    })


from .models import Campo, CampoVariedad, Variedad, SalidaDia, SalidaItem, ProduccionDia, ProduccionItem
@login_required
def salidas_excel_rango(request):
    """
    GET:
      - start, end  (YYYY-MM-DD)
      - campo       ('all' | id). 'all'/ausente => General (una hoja por campo + hoja Totales)
                      id         => Solo ese campo (una hoja).
    """
    start_str = request.GET.get("start") or request.GET.get("desde")
    end_str   = request.GET.get("end")   or request.GET.get("hasta")
    campo_id  = request.GET.get("campo")  # 'all', None o id

    d1 = parse_date(start_str) if start_str else None
    d2 = parse_date(end_str)   if end_str   else None
    if not d1 or not d2:
        return HttpResponse("Faltan parámetros start/end.", status=400)

    # Redondeo
    Q2 = Decimal('0.00001')
    def q2(x):
        if x is None: x = Decimal('0')
        if not isinstance(x, Decimal): x = Decimal(str(x))
        return x.quantize(Q2, rounding=ROUND_HALF_UP)

    # Base: salidas en rango
    base_qs = (
        SalidaDia.objects
        .filter(fecha__gte=d1, fecha__lte=d2)
        .select_related("campo")
        .prefetch_related(Prefetch("items", queryset=SalidaItem.objects.select_related("variedad")))
        .order_by("fecha", "campo__nombre", "id")
    )

    campo_sel = None
    if campo_id and campo_id != "all":
        try:
            campo_sel = Campo.objects.get(pk=int(campo_id))
        except Exception:
            return HttpResponse("Campo no existe.", status=404)
        base_qs = base_qs.filter(campo=campo_sel)

    if not base_qs.exists():
        return HttpResponse("Sin salidas en ese rango.", status=404)

    # Excel
    wb = Workbook()
    wb.remove(wb.active)

    th_font = Font(bold=True, color="FFFFFF")
    th_fill = PatternFill("solid", fgColor="225577")
    thin = Border(
        left=Side(style="thin", color="AAAAAA"),
        right=Side(style="thin", color="AAAAAA"),
        top=Side(style="thin", color="AAAAAA"),
        bottom=Side(style="thin", color="AAAAAA"),
    )

    def hoja_campo(campo, registros):
        ws = wb.create_sheet(title=str(campo)[:28])
        ws.cell(row=1, column=1, value=f"Salidas – {campo} – {d1.strftime('%d/%m/%Y')} a {d2.strftime('%d/%m/%Y')}") \
        .font = Font(size=16, bold=True, color="3C78D8")

        # ==== NUEVO: catálogo de variedades del campo ====
        variedades_campo = list(
            Variedad.objects
            .filter(activo=True, campovariedad__campo=campo, campovariedad__activo=True)
            .order_by("nombre")
            .values_list("nombre", flat=True)
        )
        # Si no hay mapeo, igual construimos desde lo que haya en items (fallback)
        if not variedades_campo:
            vset = set()
            for s in registros:
                for it in s.items.all():
                    if it.variedad:
                        vset.add(it.variedad.nombre)
            variedades_campo = sorted(vset)

        # Cabecera
        r = 3
        headers = ["Semana", "Fecha", "Variedad", "Kg", "Notas"]
        for c, h in enumerate(headers, start=1):
            cell = ws.cell(row=r, column=c, value=h)
            cell.font = th_font; cell.fill = th_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin
        r += 1

        # Totales
        tot_var = {vn: Decimal('0') for vn in variedades_campo}  # inicia todas en 0
        tot_general = Decimal('0')

        for s in registros:
            semana = s.fecha.isocalendar().week

            if s.items.all():
                for it in s.items.all():
                    c = 1
                    ws.cell(row=r, column=c, value=int(semana)).border = thin; c += 1
                    ws.cell(row=r, column=c, value=s.fecha.strftime("%d/%m/%Y")).border = thin; c += 1
                    ws.cell(row=r, column=c, value=str(it.variedad) if it.variedad else "(Sin variedad)").border = thin; c += 1
                    cv = ws.cell(row=r, column=c, value=float(q2(it.kg or 0)))
                    cv.alignment = Alignment(horizontal="right", vertical="center"); cv.border = thin; c += 1
                    ws.cell(row=r, column=c, value=s.notas or "").border = thin
                    r += 1

                    # Suma al total por variedad solo si tiene variedad definida
                    if it.variedad:
                        nombre = it.variedad.nombre
                        if nombre not in tot_var:
                            tot_var[nombre] = Decimal('0')
                        tot_var[nombre] += q2(it.kg or 0)
                    tot_general += q2(it.kg or 0)
            else:
                # salida sin items
                c = 1
                ws.cell(row=r, column=c, value=int(semana)).border = thin; c += 1
                ws.cell(row=r, column=c, value=s.fecha.strftime("%d/%m/%Y")).border = thin; c += 1
                ws.cell(row=r, column=c, value="(Sin renglones)").border = thin; c += 1
                ws.cell(row=r, column=c, value=0.0).border = thin; c += 1
                ws.cell(row=r, column=c, value=s.notas or "").border = thin
                r += 1

        # Totales por variedad (muestra TODO el catálogo, aunque quede 0)
        r += 1
        ws.cell(row=r, column=1, value="Totales por variedad").font = Font(size=13, bold=True, color="225577"); r += 1
        for c, h in enumerate(["Variedad", "Kg"], start=1):
            cell = ws.cell(row=r, column=c, value=h)
            cell.font = th_font; cell.fill = th_fill; cell.border = thin
        r += 1
        for vn in variedades_campo:
            ws.cell(row=r, column=1, value=vn).border = thin
            cv = ws.cell(row=r, column=2, value=float(q2(tot_var.get(vn, Decimal('0')))))
            cv.alignment = Alignment(horizontal="right", vertical="center"); cv.border = thin
            r += 1

        r += 1
        ws.cell(row=r, column=1, value="TOTAL SALIDAS (kg)").border = thin
        ws.cell(row=r, column=2, value=float(q2(tot_general))).border = thin

        # Anchos
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 14
        ws.column_dimensions['C'].width = 28
        ws.column_dimensions['D'].width = 14
        ws.column_dimensions['E'].width = 40


    # agrupar por campo
    por_campo = defaultdict(list)
    for s in base_qs:
        por_campo[s.campo].append(s)

    for campo, regs in por_campo.items():
        hoja_campo(campo, regs)

    # Totales (solo general)
    if not campo_sel:
        ws = wb.create_sheet(title="Totales")
        ws.cell(row=1, column=1, value=f"Totales – {d1.strftime('%d/%m/%Y')} a {d2.strftime('%d/%m/%Y')}") \
          .font = Font(size=16, bold=True, color="3C78D8")
        r = 3
        for c, h in enumerate(["Campo", "Salidas (kg)"], start=1):
            cell = ws.cell(row=r, column=c, value=h)
            cell.font = th_font; cell.fill = th_fill; cell.border = thin
        r += 1
        for campo, regs in por_campo.items():
            total = Decimal('0')
            for s in regs:
                total += q2(sum(q2(it.kg or 0) for it in s.items.all()))
            ws.cell(row=r, column=1, value=str(campo)).border = thin
            cv = ws.cell(row=r, column=2, value=float(q2(total)))
            cv.alignment = Alignment(horizontal="right", vertical="center"); cv.border = thin
            r += 1

        ws.column_dimensions['A'].width = 28
        ws.column_dimensions['B'].width = 18

    # salida
    out = BytesIO(); wb.save(out); out.seek(0)
    fname = f"salidas_rango_{d1}_{d2}"
    if campo_sel:
        fname += f"_{slugify(campo_sel.nombre)}"
    resp = HttpResponse(out.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = f'attachment; filename=\"{fname}.xlsx\"'
    return resp

@login_required
def produccion_excel_dia(request, pk):
    # Obtén la Producción del día (con el campo)
    prod = get_object_or_404(
        ProduccionDia.objects.select_related("campo"),
        pk=pk
    )

    # Helpers de redondeo a 2 decimales
    Q2 = Decimal('0.00001')
    def q2(x):
        if x is None:
            x = Decimal('0')
        if not isinstance(x, Decimal):
            x = Decimal(str(x))
        return x.quantize(Q2, rounding=ROUND_HALF_UP)

    # Renglones reales (ProduccionItem)
    items = list(
        prod.items.select_related("variedad").order_by("variedad__nombre")
    )

    # Totales (en kg)
    total_general = q2(sum(q2(it.kg or 0) for it in items))
    rezaga = q2(getattr(prod, "rezaga_kg", 0))
    empacado = q2(max(total_general - rezaga, Decimal('0')))
    porcentaje_rezaga = (q2(rezaga * Decimal('100')) / total_general) if total_general > 0 else Decimal('0')

    # ===== Excel =====
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen"

    th_font = Font(bold=True, color="FFFFFF")
    th_fill = PatternFill("solid", fgColor="225577")
    thin = Border(
        left=Side(style="thin", color="AAAAAA"),
        right=Side(style="thin", color="AAAAAA"),
        top=Side(style="thin", color="AAAAAA"),
        bottom=Side(style="thin", color="AAAAAA"),
    )

    # Encabezado
    ws.cell(row=1, column=1, value=f"Producción diaria – {prod.campo} – {prod.fecha}").font = Font(size=16, bold=True, color="3C78D8")

    # Tabla de variedades (Variedad | Kg cosechado)
    ws.cell(row=3, column=1, value="Variedad")
    ws.cell(row=3, column=2, value="Kg cosechado")
    for c in (1, 2):
        cell = ws.cell(row=3, column=c)
        cell.font = th_font
        cell.fill = th_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin

    r = 4
    if items:
        for it in items:
            ws.cell(row=r, column=1, value=str(it.variedad)).border = thin
            ckg = ws.cell(row=r, column=2, value=float(q2(it.kg or 0)))
            ckg.border = thin
            ckg.alignment = Alignment(horizontal="right", vertical="center")
            r += 1
    else:
        ws.cell(row=r, column=1, value="(Sin variedades capturadas)").border = thin
        r += 1

    # Totales
    r += 1
    ws.cell(row=r, column=1, value="Total general (kg):")
    ws.cell(row=r, column=2, value=float(total_general)); r += 1

    ws.cell(row=r, column=1, value="Rezaga (kg):")
    ws.cell(row=r, column=2, value=float(rezaga)); r += 1

    ws.cell(row=r, column=1, value="Empacado (kg):")
    ws.cell(row=r, column=2, value=float(empacado)); r += 1

    ws.cell(row=r, column=1, value="% Rezaga:")
    ws.cell(row=r, column=2, value=float(q2(porcentaje_rezaga))); r += 1

    # Anchos
    ws.column_dimensions['A'].width = 36
    ws.column_dimensions['B'].width = 18

    # Salida
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    fname = f"produccion_{slugify(prod.campo.nombre)}_{prod.fecha}.xlsx"
    resp = HttpResponse(out.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = f'attachment; filename="{fname}"'
    return resp
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.utils.text import slugify
from django.db.models import Prefetch
from django.utils.dateparse import parse_date
from decimal import Decimal, ROUND_HALF_UP
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO

Q2 = Decimal('0.00001')
def q2(x):
    if x is None:
        x = Decimal('0')
    if not isinstance(x, Decimal):
        x = Decimal(str(x))
    return x.quantize(Q2, rounding=ROUND_HALF_UP)


@login_required
def arandano_produccion_excel_rango(request):
    """
    GET:
      - start, end  (YYYY-MM-DD)
      - campo       ('all' | id). 'all' o ausente => General (una hoja por campo + hoja Totales)
                      id         => Solo ese campo (una hoja).
    """
    # --- Parámetros ---
    start_str = request.GET.get("start") or request.GET.get("desde")
    end_str   = request.GET.get("end")   or request.GET.get("hasta")
    campo_id  = request.GET.get("campo")  # 'all', None o id

    d1 = parse_date(start_str) if start_str else None
    d2 = parse_date(end_str)   if end_str   else None
    if not d1 or not d2:
        return HttpResponse("Faltan parámetros start/end.", status=400)

    # Helper redondeo a 2 decimales
    Q2 = Decimal('0.00001')
    def q2(x):
        if x is None: x = Decimal('0')
        if not isinstance(x, Decimal): x = Decimal(str(x))
        return x.quantize(Q2, rounding=ROUND_HALF_UP)

    # ===== Query base (producciones dentro del rango) =====
    base_qs = (
        ProduccionDia.objects
        .filter(fecha__gte=d1, fecha__lte=d2)
        .select_related("campo")
        .prefetch_related(
            Prefetch("items", queryset=ProduccionItem.objects.select_related("variedad"))
        )
        .order_by("fecha", "campo__nombre", "id")
    )

    campo_sel = None
    if campo_id and campo_id != "all":
        try:
            campo_sel = Campo.objects.get(pk=int(campo_id))
        except Exception:
            return HttpResponse("Campo no existe.", status=404)
        base_qs = base_qs.filter(campo=campo_sel)

    if not base_qs.exists():
        return HttpResponse("Sin registros en ese rango.", status=404)

    # ===== Excel: libro y estilos =====
    wb = Workbook()
    wb.remove(wb.active)  # quitar la hoja vacía por defecto

    th_font = Font(bold=True, color="FFFFFF")
    th_fill = PatternFill("solid", fgColor="225577")
    thin = Border(
        left=Side(style="thin", color="AAAAAA"),
        right=Side(style="thin", color="AAAAAA"),
        top=Side(style="thin", color="AAAAAA"),
        bottom=Side(style="thin", color="AAAAAA"),
    )

    def crear_hoja_campo(campo, registros):
        """
        Hoja por campo con columnas:
        Semana | Fecha | <Var1..VarN> | Total | Empacado | Rezaga | % Rezaga | Notas | CS 6 oz | CS 9.8 oz | CS 18 oz
        y una tabla de totales por variedad + Producido/Salidas/Neto.
        """
        ws = wb.create_sheet(title=str(campo)[:28])

        # Título
        ws.cell(row=1, column=1,
                value=f"{campo} – {d1.strftime('%d/%m/%Y')} a {d2.strftime('%d/%m/%Y')}"
               ).font = Font(size=16, bold=True, color="3C78D8")

        # Variedades presentes en el rango de este campo
        variedades, vset = [], set()
        for p in registros:
            for it in p.items.all():
                if it.variedad:
                    vn = it.variedad.nombre
                    if vn not in vset:
                        vset.add(vn)
                        variedades.append(vn)
        variedades.sort()

        # Cabecera
        r = 3
        headers = (
            ["Semana", "Fecha"] +
            variedades +
            ["Total", "Empacado", "Rezaga", "% Rezaga", "Notas", "CS 6 oz", "CS 9.8 oz", "CS 18 oz"]
        )
        for c, h in enumerate(headers, start=1):
            cell = ws.cell(row=r, column=c, value=h)
            cell.font = th_font; cell.fill = th_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin
        r += 1

        # Acumuladores
        tot_por_var = defaultdict(lambda: Decimal('0'))
        tot_total   = Decimal('0')
        tot_rez     = Decimal('0')
        tot_emp     = Decimal('0')
        tot_cs6     = 0
        tot_cs98    = 0
        tot_cs18    = 0

        # Filas por producción
        for p in registros:
            # kg por variedad en la fecha
            fila = {vn: Decimal('0') for vn in variedades}
            cs6_dia = cs98_dia = cs18_dia = 0

            for it in p.items.all():
                if it.variedad:
                    fila[it.variedad.nombre] += q2(it.kg or 0)
                # clamshells del día (sumamos por renglones)
                cs6_dia  += int(it.cs_6oz   or 0)
                cs98_dia += int(it.cs_9_8oz or 0)
                cs18_dia += int(it.cs_18oz  or 0)

            total  = q2(sum(fila.values()))
            rezaga = q2(getattr(p, "rezaga_kg", 0))
            emp    = q2(max(total - rezaga, Decimal('0')))
            porc   = q2((rezaga * Decimal('100') / total) if total > 0 else Decimal('0'))

            c = 1
            ws.cell(row=r, column=c, value=int(p.fecha.isocalendar().week)).border = thin; c += 1
            ws.cell(row=r, column=c, value=p.fecha.strftime("%d/%m/%Y")).border = thin; c += 1

            for vn in variedades:
                cv = ws.cell(row=r, column=c, value=float(q2(fila[vn])))
                cv.alignment = Alignment(horizontal="right", vertical="center"); cv.border = thin
                c += 1

            # Total, Empacado, Rezaga, % y Notas
            for val in (float(total), float(emp), float(rezaga), float(porc)):
                cv = ws.cell(row=r, column=c, value=val)
                cv.alignment = Alignment(horizontal="right", vertical="center"); cv.border = thin
                c += 1
            ws.cell(row=r, column=c, value=p.notas or "").border = thin; c += 1

            # Clamshells (día)
            for vcs in (cs6_dia, cs98_dia, cs18_dia):
                cv = ws.cell(row=r, column=c, value=int(vcs))
                cv.alignment = Alignment(horizontal="right", vertical="center"); cv.border = thin
                c += 1

            # Acumular totales
            for vn in variedades:
                tot_por_var[vn] += fila[vn]
            tot_total += total
            tot_rez   += rezaga
            tot_emp   += emp
            tot_cs6   += cs6_dia
            tot_cs98  += cs98_dia
            tot_cs18  += cs18_dia

            r += 1

        # Totales por variedad
        r += 1
        ws.cell(row=r, column=1, value="Totales por variedad").font = Font(size=13, bold=True, color="225577")
        r += 1
        ws.cell(row=r, column=1, value="Variedad").font = th_font; ws.cell(row=r, column=1).fill = th_fill; ws.cell(row=r, column=1).border = thin
        ws.cell(row=r, column=2, value="Kg").font = th_font; ws.cell(row=r, column=2).fill = th_fill; ws.cell(row=r, column=2).border = thin
        r += 1
        for vn in variedades:
            ws.cell(row=r, column=1, value=vn).border = thin
            cv = ws.cell(row=r, column=2, value=float(q2(tot_por_var[vn])))
            cv.alignment = Alignment(horizontal="right", vertical="center"); cv.border = thin
            r += 1

        # Totales del rango (producido/salidas/neto y clamshells)
        r += 1
        ws.cell(row=r, column=1, value="TOTAL PRODUCIDO (kg)").border = thin
        ws.cell(row=r, column=2, value=float(q2(tot_total))).border = thin; r += 1

        # Salidas del campo en el rango
        sal_qs = (
            SalidaDia.objects
            .filter(campo=campo, fecha__gte=d1, fecha__lte=d2)
            .prefetch_related("items")
        )
        tot_sal_kg  = q2(sum(q2(it.kg or 0) for s in sal_qs for it in s.items.all()))
        tot_sal_cs6 = sum(int(it.cs_6oz   or 0) for s in sal_qs for it in s.items.all())
        tot_sal_cs98= sum(int(it.cs_9_8oz or 0) for s in sal_qs for it in s.items.all())
        tot_sal_cs18= sum(int(it.cs_18oz  or 0) for s in sal_qs for it in s.items.all())

        ws.cell(row=r, column=1, value="TOTAL SALIDAS (kg)").border = thin
        ws.cell(row=r, column=2, value=float(tot_sal_kg)).border = thin; r += 1

        ws.cell(row=r, column=1, value="TOTAL NETO (kg)").border = thin
        ws.cell(row=r, column=2, value=float(q2(tot_total - tot_sal_kg))).border = thin; r += 1

        # Totales clamshells del rango (producido vs salidas)
        ws.cell(row=r, column=1, value="TOTAL CS 6 oz (prod / sal / net)").border = thin
        ws.cell(row=r, column=2, value=f"{tot_cs6} / {tot_sal_cs6} / {max(tot_cs6 - tot_sal_cs6, 0)}").border = thin; r += 1
        ws.cell(row=r, column=1, value="TOTAL CS 9.8 oz (prod / sal / net)").border = thin
        ws.cell(row=r, column=2, value=f"{tot_cs98} / {tot_sal_cs98} / {max(tot_cs98 - tot_sal_cs98, 0)}").border = thin; r += 1
        ws.cell(row=r, column=1, value="TOTAL CS 18 oz (prod / sal / net)").border = thin
        ws.cell(row=r, column=2, value=f"{tot_cs18} / {tot_sal_cs18} / {max(tot_cs18 - tot_sal_cs18, 0)}").border = thin

        # Anchos de columna
        ws.column_dimensions['A'].width = 10  # Semana
        ws.column_dimensions['B'].width = 14  # Fecha
        col_count = 2 + len(variedades) + 5 + 1 + 3  # 2 fijas + variedades + (total, emp, rez, %, notas) + 3 cs
        for i in range(3, col_count + 1):
            ws.column_dimensions[get_column_letter(i)].width = 16

    # --- Agrupar por campo y crear hojas ---
    por_campo = defaultdict(list)
    for p in base_qs:
        por_campo[p.campo].append(p)

    for campo, registros in por_campo.items():
        crear_hoja_campo(campo, registros)

    # Hoja “Totales” (solo si es general)
    if not campo_sel:
        ws = wb.create_sheet(title="Totales")
        ws.cell(row=1, column=1,
                value=f"Totales – {d1.strftime('%d/%m/%Y')} a {d2.strftime('%d/%m/%Y')}"
               ).font = Font(size=16, bold=True, color="3C78D8")
        r = 3
        for c, h in enumerate(["Campo", "Producido (kg)", "Salidas (kg)", "Neto (kg)"], start=1):
            cell = ws.cell(row=r, column=c, value=h)
            cell.font = th_font; cell.fill = th_fill
            cell.alignment = Alignment(horizontal="center", vertical="center"); cell.border = thin
        r += 1

        totP = totS = Decimal('0')
        for campo, registros in por_campo.items():
            prod = Decimal('0')
            for p in registros:
                prod += q2(sum(q2(it.kg or 0) for it in p.items.all()))
            sal_qs = (
                SalidaDia.objects
                .filter(campo=campo, fecha__gte=d1, fecha__lte=d2)
                .prefetch_related("items")
            )
            sal = q2(sum(q2(it.kg or 0) for s in sal_qs for it in s.items.all()))
            neto = q2(prod - sal)

            ws.cell(row=r, column=1, value=str(campo)).border = thin
            for idx, val in enumerate((float(q2(prod)), float(q2(sal)), float(q2(neto))), start=2):
                cv = ws.cell(row=r, column=idx, value=val)
                cv.alignment = Alignment(horizontal="right", vertical="center"); cv.border = thin
            r += 1
            totP += prod; totS += sal

        ws.cell(row=r, column=1, value="TOTALES").font = Font(bold=True)
        ws.cell(row=r, column=2, value=float(q2(totP))).font = Font(bold=True)
        ws.cell(row=r, column=3, value=float(q2(totS))).font = Font(bold=True)
        ws.cell(row=r, column=4, value=float(q2(totP - totS))).font = Font(bold=True)

        ws.column_dimensions['A'].width = 28
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18

    # Descarga
    out = BytesIO(); wb.save(out); out.seek(0)
    fname = f"arandano_rango_{d1}_{d2}"
    if campo_sel:
        fname += f"_{slugify(campo_sel.nombre)}"
    resp = HttpResponse(out.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = f'attachment; filename="{fname}.xlsx"'
    return resp

@login_required
def salidas_excel_inventario(request):
    """
    Excel tipo 'inventario' (por campo o general):
      - Detalle de ENTRADAS (ProduccionDia/ProduccionItem) con clamshells
      - Detalle de SALIDAS (SalidaDia/SalidaItem) con clamshells
      - Resumen por variedad (entradas/salidas/neto en kg)
      - Resumen clamshells (prod/sal/net por tamaño)
    GET:
      - start, end  (YYYY-MM-DD)
      - campo       ('all' | id). 'all'/ausente => General (una hoja por campo + hoja Totales)
                      id         => Solo ese campo (una hoja).
    """
    # --- Parámetros ---
    start_str = request.GET.get("start") or request.GET.get("desde")
    end_str   = request.GET.get("end")   or request.GET.get("hasta")
    campo_id  = request.GET.get("campo")  # 'all', None o id

    d1 = parse_date(start_str) if start_str else None
    d2 = parse_date(end_str)   if end_str   else None
    if not d1 or not d2:
        return HttpResponse("Faltan parámetros start/end.", status=400)

    Q5 = Decimal('0.00001')
    def q5(x):
        if x is None: x = Decimal('0')
        if not isinstance(x, Decimal): x = Decimal(str(x))
        return x.quantize(Q5, rounding=ROUND_HALF_UP)

    # ===== Query base (rango) =====
    # Entradas
    entradas_qs = (
        ProduccionDia.objects
        .filter(fecha__gte=d1, fecha__lte=d2)
        .select_related("campo")
        .prefetch_related(Prefetch("items", queryset=ProduccionItem.objects.select_related("variedad")))
        .order_by("fecha", "campo__nombre", "id")
    )
    # Salidas
    salidas_qs = (
        SalidaDia.objects
        .filter(fecha__gte=d1, fecha__lte=d2)
        .select_related("campo")
        .prefetch_related(Prefetch("items", queryset=SalidaItem.objects.select_related("variedad")))
        .order_by("fecha", "campo__nombre", "id")
    )

    campo_sel = None
    if campo_id and campo_id != "all":
        try:
            campo_sel = Campo.objects.get(pk=int(campo_id))
        except Exception:
            return HttpResponse("Campo no existe.", status=404)
        entradas_qs = entradas_qs.filter(campo=campo_sel)
        salidas_qs  = salidas_qs.filter(campo=campo_sel)

    if not entradas_qs.exists() and not salidas_qs.exists():
        return HttpResponse("Sin registros (entradas/salidas) en ese rango.", status=404)

    # ===== Excel =====
    wb = Workbook()
    wb.remove(wb.active)  # quitar hoja por defecto

    th_font = Font(bold=True, color="FFFFFF")
    th_fill = PatternFill("solid", fgColor="225577")
    thin = Border(
        left=Side(style="thin", color="AAAAAA"),
        right=Side(style="thin", color="AAAAAA"),
        top=Side(style="thin", color="AAAAAA"),
        bottom=Side(style="thin", color="AAAAAA"),
    )

    def hoja_campo(campo, registros_ent, registros_sal):
        """
        Una hoja por campo con:
        - Tabla detalle ENTRADAS
        - Tabla detalle SALIDAS
        - Resumen por variedad (kg)
        - Resumen clamshells (prod/sal/net)
        """
        ws = wb.create_sheet(title=str(campo)[:28])
        ws.cell(row=1, column=1,
                value=f"Inventario – {campo} – {d1.strftime('%d/%m/%Y')} a {d2.strftime('%d/%m/%Y')}"
               ).font = Font(size=16, bold=True, color="3C78D8")

        r = 3

        # -------------------- ENTRADAS (detalle) --------------------
        ws.cell(row=r, column=1, value="ENTRADAS (Producción)").font = Font(size=13, bold=True, color="225577")
        r += 1
        headers_ent = ["Semana", "Fecha", "Variedad", "Kg", "CS 6 oz", "CS 9.8 oz", "CS 18 oz", "Notas"]
        for c, h in enumerate(headers_ent, start=1):
            cell = ws.cell(row=r, column=c, value=h)
            cell.font = th_font; cell.fill = th_fill
            cell.alignment = Alignment(horizontal="center", vertical="center"); cell.border = thin
        r += 1

        # Acumular para resumen
        ent_por_var = defaultdict(Decimal)   # kg por variedad
        ent_cs6 = ent_cs98 = ent_cs18 = 0

        for p in registros_ent:
            semana = p.fecha.isocalendar().week
            # sumatorio de clamshells por día (toma de todos los renglones)
            cs6_dia = cs98_dia = cs18_dia = 0
            for it in p.items.all():
                cs6_dia  += int(getattr(it, "cs_6oz", 0) or 0)
                cs98_dia += int(getattr(it, "cs_9_8oz", 0) or 0)
                cs18_dia += int(getattr(it, "cs_18oz", 0) or 0)
                # detalle por renglón
                c = 1
                ws.cell(row=r, column=c, value=int(semana)).border = thin; c += 1
                ws.cell(row=r, column=c, value=p.fecha.strftime("%d/%m/%Y")).border = thin; c += 1
                ws.cell(row=r, column=c, value=str(it.variedad) if it.variedad else "(sin variedad)").border = thin; c += 1
                cv = ws.cell(row=r, column=c, value=float(q5(it.kg or 0)))
                cv.alignment = Alignment(horizontal="right", vertical="center"); cv.border = thin; c += 1
                ws.cell(row=r, column=c, value=int(getattr(it, "cs_6oz", 0) or 0)).border = thin; c += 1
                ws.cell(row=r, column=c, value=int(getattr(it, "cs_9_8oz", 0) or 0)).border = thin; c += 1
                ws.cell(row=r, column=c, value=int(getattr(it, "cs_18oz", 0) or 0)).border = thin; c += 1
                ws.cell(row=r, column=c, value=p.notas or "").border = thin
                r += 1

                if it.variedad:
                    ent_por_var[it.variedad.nombre] += q5(it.kg or 0)

            ent_cs6  += cs6_dia
            ent_cs98 += cs98_dia
            ent_cs18 += cs18_dia

        # Anchos (entradas)
        for i, w in enumerate((10, 14, 28, 14, 12, 12, 12, 40), start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        r += 1

        # -------------------- SALIDAS (detalle) --------------------
        ws.cell(row=r, column=1, value="SALIDAS").font = Font(size=13, bold=True, color="225577"); r += 1
        headers_sal = ["Semana", "Fecha", "Variedad", "Kg", "CS 6 oz", "CS 9.8 oz", "CS 18 oz", "Notas"]
        for c, h in enumerate(headers_sal, start=1):
            cell = ws.cell(row=r, column=c, value=h)
            cell.font = th_font; cell.fill = th_fill
            cell.alignment = Alignment(horizontal="center", vertical="center"); cell.border = thin
        r += 1

        sal_por_var = defaultdict(Decimal)
        sal_cs6 = sal_cs98 = sal_cs18 = 0

        for s in registros_sal:
            semana = s.fecha.isocalendar().week
            cs6_dia = cs98_dia = cs18_dia = 0
            if s.items.all():
                for it in s.items.all():
                    cs6_dia  += int(getattr(it, "cs_6oz", 0) or 0)
                    cs98_dia += int(getattr(it, "cs_9_8oz", 0) or 0)
                    cs18_dia += int(getattr(it, "cs_18oz", 0) or 0)

                    c = 1
                    ws.cell(row=r, column=c, value=int(semana)).border = thin; c += 1
                    ws.cell(row=r, column=c, value=s.fecha.strftime("%d/%m/%Y")).border = thin; c += 1
                    ws.cell(row=r, column=c, value=str(it.variedad) if it.variedad else "(Sin variedad)").border = thin; c += 1
                    cv = ws.cell(row=r, column=c, value=float(q5(it.kg or 0)))
                    cv.alignment = Alignment(horizontal="right", vertical="center"); cv.border = thin; c += 1
                    ws.cell(row=r, column=c, value=int(getattr(it, "cs_6oz", 0) or 0)).border = thin; c += 1
                    ws.cell(row=r, column=c, value=int(getattr(it, "cs_9_8oz", 0) or 0)).border = thin; c += 1
                    ws.cell(row=r, column=c, value=int(getattr(it, "cs_18oz", 0) or 0)).border = thin; c += 1
                    ws.cell(row=r, column=c, value=s.notas or "").border = thin
                    r += 1

                    if it.variedad:
                        sal_por_var[it.variedad.nombre] += q5(it.kg or 0)
            else:
                # salida sin items
                c = 1
                ws.cell(row=r, column=c, value=int(semana)).border = thin; c += 1
                ws.cell(row=r, column=c, value=s.fecha.strftime("%d/%m/%Y")).border = thin; c += 1
                ws.cell(row=r, column=c, value="(Sin renglones)").border = thin; c += 1
                ws.cell(row=r, column=c, value=0.0).border = thin; c += 1
                ws.cell(row=r, column=c, value=0).border = thin; c += 1
                ws.cell(row=r, column=c, value=0).border = thin; c += 1
                ws.cell(row=r, column=c, value=0).border = thin; c += 1
                ws.cell(row=r, column=c, value=s.notas or "").border = thin
                r += 1

            sal_cs6  += cs6_dia
            sal_cs98 += cs98_dia
            sal_cs18 += cs18_dia

        # Anchos (salidas) — ya definidos por arriba, se mantienen

        r += 1

        # -------------------- RESUMEN POR VARIEDAD --------------------
        ws.cell(row=r, column=1, value="Resumen por variedad (kg)").font = Font(size=13, bold=True, color="225577"); r += 1
        for c, h in enumerate(["Variedad", "Entradas (kg)", "Salidas (kg)", "Neto (kg)"], start=1):
            cell = ws.cell(row=r, column=c, value=h)
            cell.font = th_font; cell.fill = th_fill
            cell.alignment = Alignment(horizontal="center", vertical="center"); cell.border = thin
        r += 1

        todas_var = sorted(set(ent_por_var.keys()) | set(sal_por_var.keys()))
        total_ent = total_sal = Decimal('0')
        for vn in todas_var:
            ent = q5(ent_por_var.get(vn, Decimal('0')))
            sal = q5(sal_por_var.get(vn, Decimal('0')))
            net = q5(ent - sal)
            ws.cell(row=r, column=1, value=vn).border = thin
            for idx, val in enumerate((float(ent), float(sal), float(net)), start=2):
                cv = ws.cell(row=r, column=idx, value=val)
                cv.alignment = Alignment(horizontal="right", vertical="center"); cv.border = thin
            r += 1
            total_ent += ent; total_sal += sal

        ws.cell(row=r, column=1, value="TOTALES").font = Font(bold=True)
        ws.cell(row=r, column=2, value=float(q5(total_ent))).font = Font(bold=True)
        ws.cell(row=r, column=3, value=float(q5(total_sal))).font = Font(bold=True)
        ws.cell(row=r, column=4, value=float(q5(total_ent - total_sal))).font = Font(bold=True)
        r += 2

        # -------------------- RESUMEN CLAMSHELLS --------------------
        ws.cell(row=r, column=1, value="Resumen clamshells (prod / sal / net)").font = Font(size=13, bold=True, color="225577"); r += 1
        for c, h in enumerate(["Tamaño", "Prod", "Sal", "Neto"], start=1):
            cell = ws.cell(row=r, column=c, value=h)
            cell.font = th_font; cell.fill = th_fill
            cell.alignment = Alignment(horizontal="center", vertical="center"); cell.border = thin
        r += 1

        filas_cs = [
            ("CS 6 oz",  ent_cs6,  sal_cs6,  max(ent_cs6  - sal_cs6,  0)),
            ("CS 9.8 oz",ent_cs98, sal_cs98, max(ent_cs98 - sal_cs98, 0)),
            ("CS 18 oz", ent_cs18, sal_cs18, max(ent_cs18 - sal_cs18, 0)),
        ]
        for nombre, prod, sal, net in filas_cs:
            ws.cell(row=r, column=1, value=nombre).border = thin
            for idx, val in enumerate((prod, sal, net), start=2):
                cv = ws.cell(row=r, column=idx, value=int(val))
                cv.alignment = Alignment(horizontal="right", vertical="center"); cv.border = thin
            r += 1

        # Ajustes de ancho ya hechos arriba.

    # --- Agrupar por campo y crear hojas ---
    por_campo_ent = defaultdict(list)
    por_campo_sal = defaultdict(list)
    for p in entradas_qs: por_campo_ent[p.campo].append(p)
    for s in salidas_qs:  por_campo_sal[s.campo].append(s)

    todos_campos = sorted(set(por_campo_ent.keys()) | set(por_campo_sal.keys()), key=lambda c: c.nombre)

    for campo in todos_campos:
        regs_ent = por_campo_ent.get(campo, [])
        regs_sal = por_campo_sal.get(campo, [])
        hoja_campo(campo, regs_ent, regs_sal)

    # Hoja “Totales” si es general
    if not campo_sel:
        ws = wb.create_sheet(title="Totales")
        ws.cell(row=1, column=1,
                value=f"Totales – {d1.strftime('%d/%m/%Y')} a {d2.strftime('%d/%m/%Y')}"
               ).font = Font(size=16, bold=True, color="3C78D8")
        r = 3
        for c, h in enumerate(["Campo", "Entradas (kg)", "Salidas (kg)", "Neto (kg)"], start=1):
            cell = ws.cell(row=r, column=c, value=h)
            cell.font = th_font; cell.fill = th_fill
            cell.alignment = Alignment(horizontal="center", vertical="center"); cell.border = thin
        r += 1

        for campo in todos_campos:
            regs_ent = por_campo_ent.get(campo, [])
            regs_sal = por_campo_sal.get(campo, [])

            ent_total = Decimal('0')
            for p in regs_ent:
                for it in p.items.all():
                    ent_total += q5(it.kg or 0)

            sal_total = Decimal('0')
            for s in regs_sal:
                for it in s.items.all():
                    sal_total += q5(it.kg or 0)

            ws.cell(row=r, column=1, value=str(campo)).border = thin
            for idx, val in enumerate((float(q5(ent_total)), float(q5(sal_total)), float(q5(ent_total - sal_total))), start=2):
                cv = ws.cell(row=r, column=idx, value=val)
                cv.alignment = Alignment(horizontal="right", vertical="center"); cv.border = thin
            r += 1

        ws.column_dimensions['A'].width = 28
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18

    # Descarga
    out = BytesIO(); wb.save(out); out.seek(0)
    fname = f"inventario_arandano_{d1}_{d2}"
    if campo_sel:
        fname += f"_{slugify(campo_sel.nombre)}"
    resp = HttpResponse(out.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = f'attachment; filename="{fname}.xlsx"'
    return resp